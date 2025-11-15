#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Advanced CryptoBot w/ Honeypot + Renounce + Passing Refresh + Silent Checks + Advanced Contract Verification
------------------------------------------------------------------------------------------------------------
Features:
1) Paper trading (inactive).
2) Honeypot check on both token0 & token1 (via Honeypot.is).
3) If DexScreener liquidity ≤ $1 => fail as "Rugpull".
4) Re-check failing pairs using RECHECK_DELAYS.
5) Passing pairs get scheduled REFRESH checks (30s up to 15m).
6) Passing pairs also do SILENT checks every 10m for 2h to watch MC growth.
7) 1-minute “quick honeypot check” for passing pairs => if dev toggles honeypot.
8) Optionally check for “renounce” => if not renounced, fail or penalize.
9) Minimizes repeated "No more scheduled attempts" spam.
10) **NEW**: Advanced Contract Verification system (fetch from Etherscan, parse for high-risk patterns).
"""

import os
import time
import json
import logging
import re
from datetime import datetime, timedelta
try:
    from solidity_parser import parser
except ImportError:  # pragma: no cover - optional dependency
    parser = None
import contextlib
import io
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from typing import Optional, Dict, Tuple, List, Union, Set, Any
from openpyxl import Workbook, load_workbook
import traceback
import requests
import numpy as np

# Web3 & dependencies
from web3 import Web3, HTTPProvider
from web3.types import HexBytes
from eth_abi.abi import decode
from eth_utils import to_checksum_address

# Requests
import asyncio
import aiohttp
import threading
from collections import Counter, deque
from concurrent.futures import Future
from itertools import cycle, product
# Ensure local imports work even when script executed from a different directory
import sys
from pathlib import Path
import urllib.parse

# Add the script directory to sys.path for relative imports
SCRIPT_DIR = Path(__file__).resolve().parent
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

# Import wallet tracker with fallback for legacy filename
tracker_etherscan_get_async = None

try:
    from wallet_tracker import (
        SmartWalletTracker,
        wallet_activity_callback,
        get_shared_tracker,
        set_notifier,
        set_etherscan_lookup_enabled as set_tracker_etherscan_enabled,
        _etherscan_get_async as tracker_etherscan_get_async,
    )
except (ModuleNotFoundError, ImportError, AttributeError):
    try:  # pragma: no cover - backward compatibility
        from wallet_tracker_system import (
            SmartWalletTracker,
            wallet_activity_callback,
            get_shared_tracker,
            set_notifier,
            set_etherscan_lookup_enabled as set_tracker_etherscan_enabled,
            _etherscan_get_async as tracker_etherscan_get_async,
        )
    except ModuleNotFoundError as exc:
        raise ModuleNotFoundError(
            "wallet_tracker module not found; ensure wallet_tracker.py is present"
        ) from exc

# Shared configuration helpers
from etherscan_config import (
    load_etherscan_base_urls,
    load_etherscan_keys,
    make_key_getter,
)

# Fallback no-op to preserve compatibility if tracker module lacks the helper.
try:
    set_tracker_etherscan_enabled
except NameError:  # pragma: no cover - defensive
    def set_tracker_etherscan_enabled(enabled: bool, reason: str = "") -> None:
        return None

# On Windows the default ProactorEventLoop can emit "Event loop is closed"
# messages when asyncio.run() is used repeatedly. Switching to the
# Selector policy avoids those spurious warnings.
if os.name == "nt":
    asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())


_TRANSFER_NAME_REGEX = re.compile(
    r"\b(safeTransferFrom|transferFrom|safeTransfer|transfer)\s*\(",
    re.IGNORECASE,
)
_SUSPICIOUS_DEST_TERMS = (
    "owner",
    "dev",
    "marketing",
    "team",
    "treasury",
    "wallet",
    "tax",
)
_SUSPICIOUS_SOURCE_TERMS = ("msg.sender", "tx.origin", "_msgsender", "_msgsender()")
_CONTRACT_SELF_TERM = "address(this)"


def _extract_call_arguments(source_text: str, open_paren_index: int) -> str:
    depth = 0
    args_chars: List[str] = []
    for idx in range(open_paren_index + 1, len(source_text)):
        ch = source_text[idx]
        if ch == "(":
            depth += 1
            args_chars.append(ch)
        elif ch == ")":
            if depth == 0:
                return "".join(args_chars)
            depth -= 1
            args_chars.append(ch)
        else:
            args_chars.append(ch)
    return ""


def _has_wallet_drainer_pattern(source_text: str) -> bool:
    """Detect suspicious token transfer patterns that siphon wallets."""

    for match in _TRANSFER_NAME_REGEX.finditer(source_text):
        prefix = source_text[: match.start()].rstrip()
        lower_prefix = prefix.lower()
        if lower_prefix.endswith("emit") or lower_prefix.endswith("event") or lower_prefix.endswith(
            "function"
        ):
            continue
        open_paren_index = match.end() - 1
        raw_args = _extract_call_arguments(source_text, open_paren_index)
        if not raw_args:
            continue
        args = [a.strip().lower() for a in raw_args.split(",")]
        if len(args) < 2:
            continue
        from_arg = args[0]
        to_arg = args[1]

        if any(term in to_arg for term in (_CONTRACT_SELF_TERM, *_SUSPICIOUS_SOURCE_TERMS)):
            return True

        keyword_to = any(term in to_arg for term in _SUSPICIOUS_DEST_TERMS)
        from_matches_source = any(term in from_arg for term in _SUSPICIOUS_SOURCE_TERMS)
        from_is_contract = _CONTRACT_SELF_TERM in from_arg
        to_is_contract = _CONTRACT_SELF_TERM in to_arg

        if to_is_contract and not from_is_contract:
            return True
        if keyword_to and (from_matches_source or from_is_contract):
            return True
    return False


def create_aiohttp_session(**kwargs: Dict) -> aiohttp.ClientSession:
    """Return an ``aiohttp`` session that honours environment proxy settings."""

    if "trust_env" not in kwargs:
        kwargs["trust_env"] = True
    return aiohttp.ClientSession(**kwargs)

###########################################################
# 1. GLOBAL CONFIG & CONSTANTS
###########################################################

INFURA_URL = os.getenv(
    "INFURA_URL", "https://mainnet.infura.io/v3/92702160421d44dd8551754e78633549"
)
INFURA_URL_V3 = os.getenv(
    "INFURA_URL_V3", "https://mainnet.infura.io/v3/7236b8ffcce1451caaf08b275fb6dfc7"
)
# backup provider for rate limit situations
INFURA_URL_BACKUP = os.getenv(
    "INFURA_URL_BACKUP",
    "https://mainnet.infura.io/v3/ec4f1dd756644dcb9eb46762c4c4c9c0",
)
INFURA_URL_V3_BACKUP = os.getenv("INFURA_URL_V3_BACKUP", INFURA_URL_BACKUP)
INFURA_URL_EMERGENCY_1 = os.getenv(
    "INFURA_URL_EMERGENCY_1",
    "https://mainnet.infura.io/v3/81a1564705e54d6bb52d3a98bc1767fc",
)
INFURA_URL_EMERGENCY_2 = os.getenv(
    "INFURA_URL_EMERGENCY_2",
    "https://mainnet.infura.io/v3/bd68c7ff3d58401ea02179b15a9efd0f",
)
ALCHEMY_URL = os.getenv(
    "ALCHEMY_URL",
    "https://eth-mainnet.g.alchemy.com/v2/ICzV00BkkR9g70gaOJrx0O80fO_c2oPB",
)
ALCHEMY_URL_BACKUP = os.getenv(
    "ALCHEMY_URL_BACKUP",
    "https://eth-mainnet.g.alchemy.com/v2/_KmbgabXYB1pttUpXP5Ls",
)

TELEGRAM_BOT_TOKEN = os.getenv(
    "TELEGRAM_BOT_TOKEN", "8274484247:AAEoiTgXb6xLDmmSU3yLbqQaMOW81v541pY"
)
TELEGRAM_CHAT_ID = os.getenv("TELEGRAM_CHAT_ID", "-4895948667")
# Pre-compute base URL to avoid repetition
TELEGRAM_BASE_URL = (
    f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}" if TELEGRAM_BOT_TOKEN else ""
)

# The Graph configuration
GRAPH_URL = "https://gateway.thegraph.com/api/subgraphs/id/EYCKATKGBKLWvSfwvBjzfCBmGwYNdVkduYXVivCsLR"
GRAPH_BEARER = "6ab18515ae540220006db77a4472de7a"

ETHPLORER_BASE_URL = os.getenv("ETHPLORER_BASE_URL", "https://api.ethplorer.io")
_DEFAULT_ETHPLORER_KEYS = [
    "EK-siGhL-4qx9Cy7-Uhqwh",
    "EK-sAJiB-RAC8sYw-bNwAE",
]
_ethplorer_keys_env = os.getenv("ETHPLORER_API_KEYS")
if _ethplorer_keys_env:
    ETHPLORER_API_KEYS = [key.strip() for key in _ethplorer_keys_env.split(",") if key.strip()]
else:
    single_key = os.getenv("ETHPLORER_API_KEY", "").strip()
    if single_key:
        ETHPLORER_API_KEYS = [single_key]
    else:
        ETHPLORER_API_KEYS = _DEFAULT_ETHPLORER_KEYS.copy()

if not ETHPLORER_API_KEYS:
    ETHPLORER_API_KEYS = _DEFAULT_ETHPLORER_KEYS.copy()

ETHPLORER_API_KEY = ETHPLORER_API_KEYS[0]
_ETHPLORER_KEY_LOCK = threading.Lock()
_ETHPLORER_KEY_CYCLE = cycle(ETHPLORER_API_KEYS)


def get_next_ethplorer_key() -> str:
    with _ETHPLORER_KEY_LOCK:
        return next(_ETHPLORER_KEY_CYCLE)

UNISWAP_V2_FACTORY_ADDRESS = to_checksum_address(
    "0x5C69bEe701ef814a2B6a3EDD4B1652CB9cc5aA6f"
)
UNISWAP_V3_FACTORY_ADDRESS = to_checksum_address(
    "0x1F98431c8aD98523631AE4a59f267346ea31F984"
)

PAIR_CREATED_TOPIC_V2 = (
    "0x0d3648bd0f6ba80134a33ba9275ac585d9d315f0ad8355cddefde31afa28d0e9"
)
POOL_CREATED_TOPIC_V3 = (
    "0x783cca1c0412dd0d695e784568c96da2e9c22ff989357a2e8b1d9b2b4e6b7118"
)

# backwards compatibility
UNISWAP_FACTORY_ADDRESS = UNISWAP_V2_FACTORY_ADDRESS
PAIR_CREATED_TOPIC = PAIR_CREATED_TOPIC_V2

MIN_PASS_THRESHOLD = 6

# Criteria thresholds
MIN_LIQUIDITY_USD = 20_000
MIN_VOLUME_USD = 25_000
MIN_FDV_USD = 40_000
MIN_MARKETCAP_USD = 40_000
MIN_BUYS_FIRST_HOUR = 20
MIN_TRADES_REQUIRED = 20

VERIFICATION_RETRY_DELAY = int(os.getenv("VERIFICATION_RETRY_DELAY", "300"))

# Allow freshly-created pairs some time to appear on DexScreener before we give up
DEXSCREENER_NOT_LISTED_REQUEUE_WINDOW = int(
    os.getenv("DEXSCREENER_NOT_LISTED_REQUEUE_WINDOW", "1800")
)


def critical_verification_failure(extra: Dict) -> Tuple[bool, str]:
    """Return True if contract verification indicates a critical failure."""

    status = str(extra.get("contractCheckStatus", "")).upper()
    risk = extra.get("riskScore")
    if status == "ERROR":
        return True, "verification error"
    if risk == 9999:
        return True, "risk score 9999"
    return False, ""


def _format_retry_delay(seconds: int) -> str:
    if seconds <= 0:
        return "soon"
    if seconds % 60 == 0:
        minutes = seconds // 60
        return f"{minutes} minute{'s' if minutes != 1 else ''}"
    return f"{seconds} seconds"


def _unique_urls(urls: List[Optional[str]]) -> List[str]:
    seen = set()
    result: List[str] = []
    for url in urls:
        if url and url not in seen:
            result.append(url)
            seen.add(url)
    return result


def _schedule_verification_retry(
    pair_addr: str,
    store: Dict[str, Any],
    extra: Dict,
    context: str,
    attempt_num: Optional[int] = None,
) -> None:
    """Record a verification warning, schedule a retry, and notify Telegram."""

    now_ts = time.time()
    store["verification_warning"] = True
    store["verification_retry_at"] = now_ts + VERIFICATION_RETRY_DELAY
    if "attempt_index" in store:
        store["attempt_index"] = max(store.get("attempt_index", 0) - 1, 0)

    delay_text = _format_retry_delay(VERIFICATION_RETRY_DELAY)
    token_name = extra.get("tokenName") or "Unnamed"
    status = extra.get("contractCheckStatus") or "ERROR"
    risk_score = extra.get("riskScore", "unknown")
    attempt_txt = f" (Attempt #{attempt_num})" if attempt_num else ""

    warning_msg = (
        f"⚠️ <b>{token_name}</b> verification warning{attempt_txt}\n"
        f"Pair: <code>{pair_addr}</code>\n"
        f"Context: {context}\n"
        f"Status: {status}\n"
        f"Risk Score: {risk_score}\n"
        f"Rechecking contract verification in {delay_text}."
    )
    send_telegram_message(warning_msg)


INFURA_EMERGENCY_URLS = _unique_urls(
    [INFURA_URL_EMERGENCY_1, INFURA_URL_EMERGENCY_2]
)

ALCHEMY_PROVIDER_URLS = _unique_urls([ALCHEMY_URL, ALCHEMY_URL_BACKUP])
_READ_PROVIDER_INDEX = -1


def _init_read_provider(urls: List[str]) -> Tuple[Web3, int]:
    """Initialise the primary read-only provider with fallback support."""

    if not urls:
        raise RuntimeError("No Alchemy provider URLs configured")

    last_exc: Optional[Exception] = None
    for idx, url in enumerate(urls):
        try:
            provider = Web3(HTTPProvider(url))
            if provider.is_connected():
                if idx > 0:
                    log_event(
                        logging.WARNING,
                        "rpc_provider_init",
                        "reader provider fallback engaged",
                        context={"url": url, "index": idx},
                    )
                return provider, idx
            log_event(
                logging.ERROR,
                "rpc_provider_init",
                "reader provider unreachable",
                context={"url": url, "index": idx},
            )
        except Exception as exc:  # pragma: no cover - defensive connection guard
            last_exc = exc
            log_event(
                logging.ERROR,
                "rpc_provider_init",
                "reader provider connection error",
                error=str(exc),
                context={"url": url, "index": idx},
            )

    log_event(
        logging.ERROR,
        "rpc_provider_init",
        "reader provider fallback failed; defaulting to primary",
        error=str(last_exc) if last_exc else "unreachable",
        context={"url": urls[0], "index": 0},
    )
    return Web3(HTTPProvider(urls[0])), 0


def _current_read_provider_url() -> str:
    if 0 <= _READ_PROVIDER_INDEX < len(ALCHEMY_PROVIDER_URLS):
        return ALCHEMY_PROVIDER_URLS[_READ_PROVIDER_INDEX]
    return ALCHEMY_URL

# Wallet tracker settings
WALLET_REPORT_TTL = 600  # 10 minutes
WALLET_REPORT_CACHE: Dict[str, dict] = {}
wallet_tracker = None  # initialized after Web3 setup
wallet_monitor_tasks: Dict[str, Future] = {}
wallet_monitor_stops: Dict[str, asyncio.Event] = {}
MAX_WALLET_MONITORS = 5
wallet_event_loop = asyncio.new_event_loop()
threading.Thread(target=wallet_event_loop.run_forever, daemon=True).start()

import threading as _threading


# Known profitable wallets for smart money detection
SMART_MONEY_WALLETS = {
    "0x742d35cc6634c0532925a3b844bc454e4438f44e",
    "0xfe9e8709d3215310075d67e3ed32a380ccf451c8",
}

# Track overall market regime to tune detection thresholds
def detect_market_mode() -> str:
    try:
        resp = requests.get("https://api.coingecko.com/api/v3/global", timeout=10)
        change = resp.json()["data"]["market_cap_change_percentage_24h_usd"]
        return "bull" if change and change > 0 else "bear"
    except Exception:
        return "bull"

MARKET_MODE = detect_market_mode()

def start_market_mode_monitor():
    def _run():
        global MARKET_MODE
        while True:
            MARKET_MODE = detect_market_mode()
            time.sleep(3600)
    threading.Thread(target=_run, daemon=True).start()

# Common constants
ZERO_ADDRESS = "0x0000000000000000000000000000000000000000"

# Tokens considered base assets (ignored for certain checks)
BASE_TOKENS = {
    "0xC02aaA39b223FE8D0A0e5C4F27eAD9083C756Cc2",  # WETH
    "0xA0b86991C6218B36C1d19D4A2E9Eb0CE3606eB48",  # USDC
    "0xdAC17F958D2ee523a2206206994597C13D831ec7",  # USDT
    ZERO_ADDRESS,
}
# WETH token address used to identify main token
WETH_ADDRESS = to_checksum_address("0xC02aaA39b223FE8D0A0E5C4F27eAD9083C756Cc2")
# EIP-1967 implementation slot for proxies
EIP1967_IMPL_SLOT = bytes.fromhex(
    "360894a13ba1a3210667c828492db98dca3e2076cc3735a920a3ca505d382bbc"
)

# Uniswap pair ABI fragment
PAIR_ABI = [
    {
        "constant": True,
        "inputs": [],
        "name": "totalSupply",
        "outputs": [{"name": "", "type": "uint256"}],
        "type": "function",
    }
]

# Burn event topic for liquidity removal
BURN_TOPIC = Web3.keccak(text="Burn(address,uint256,uint256,address)").hex()
# Swap event topics used for first-sell detection
SWAP_TOPIC_V2 = Web3.keccak(
    text="Swap(address,uint256,uint256,uint256,uint256,address)"
).hex()
SWAP_TOPIC_V3 = Web3.keccak(
    text="Swap(address,address,int256,int256,uint160,uint128,int24)"
).hex()


def get_non_weth_token(token0: str, token1: str) -> str:
    """Return the token address that is not WETH."""
    if token0.lower() == WETH_ADDRESS.lower():
        return token1
    return token0

# Re-check intervals for failing pairs
RECHECK_DELAYS = [60, 180, 300, 600, 1800]  # 1m,3m,5m,10m,30m

# Passing refresh intervals
PASSING_REFRESH_DELAYS = [
    30,
    45,
    45,
    30,
    60,
    300,
    600,
    900,
]  # 30s,45s,45s,30s,1m,5m,10m,15m

# Additional silent MC check
SILENT_CHECK_INTERVAL = 600  # 10 min
SILENT_CHECK_DURATION = 7200  # 2 hours

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
LAST_BLOCK_FILE_V2 = os.path.join(SCRIPT_DIR, "last_block_v2.json")
LAST_BLOCK_FILE_V3 = os.path.join(SCRIPT_DIR, "last_block_v3.json")
EXCEL_FILE = os.path.join(SCRIPT_DIR, "pairs.xlsx")
MAIN_LOOP_SLEEP = 2

LOG_LEVEL = os.getenv("LOG_LEVEL", "INFO").upper()
logging.basicConfig(
    level=getattr(logging, LOG_LEVEL, logging.INFO),
    format="[%(levelname)s] %(asctime)s - %(message)s",
    handlers=[logging.StreamHandler()],
    force=True,
)
logger = logging.getLogger("trading_bot")


def log_event(
    level: int,
    action: str,
    message: str,
    *,
    pair: Optional[str] = None,
    latency_ms: Optional[float] = None,
    error: Optional[str] = None,
    context: Optional[dict] = None,
) -> None:
    """Emit a structured log entry with standard metadata."""

    parts = []
    if action:
        parts.append(f"[{action}]")
    parts.append(message)

    meta: List[str] = []
    if pair:
        meta.append(f"pair={pair}")
    if latency_ms is not None:
        meta.append(f"latency={latency_ms:.2f}ms")
    if error:
        meta.append(f"error={error}")
    if context:
        try:
            context_str = json.dumps(context, sort_keys=True, default=str)
        except TypeError:
            context_str = str(context)
        meta.append(f"context={context_str}")

    text = " ".join(parts)
    if meta:
        text = f"{text} | {' '.join(meta)}"

    logger.log(level, text)


class MetricsCollector:
    """Collect runtime metrics and emit periodic summaries."""

    def __init__(self) -> None:
        self.lock = threading.Lock()
        self.minute_counts: Counter = Counter()
        self.rpc_latencies: List[float] = []
        self.api_calls: int = 0
        self.api_errors: int = 0
        self.last_emit = time.time()
        self.last_pair_seen = time.time()
        self.error_events: deque = deque()
        self.error_window = 300  # 5 minutes
        self.error_threshold = 5
        self.last_error_alert = 0.0
        self.zero_throughput_window = 600  # 10 minutes
        self.zero_throughput_alerted = False
        self.queue_depth_callback = None
        self.stop_event = threading.Event()
        self.event_history = {name: deque() for name in ("pairs_scanned", "passes")}
        self.daily_totals: Counter = Counter()
        self.daily_period_start = datetime.utcnow().replace(
            hour=0, minute=0, second=0, microsecond=0
        )
        threading.Thread(target=self._emit_loop, daemon=True).start()

    def increment(self, name: str, value: int = 1) -> None:
        with self.lock:
            self.minute_counts[name] += value
            if name == "pairs_scanned":
                self.last_pair_seen = time.time()
                self.zero_throughput_alerted = False
            if name in self.event_history:
                now = time.time()
                hist = self.event_history[name]
                for _ in range(value):
                    hist.append(now)
                self._trim_history(name, now)
                self.daily_totals[name] += value

    def record_rpc_call(self, latency_ms: float, *, error: bool = False) -> None:
        with self.lock:
            self.rpc_latencies.append(latency_ms)
            self.minute_counts["rpc_calls"] += 1
            if error:
                self.minute_counts["rpc_errors"] += 1

    def record_api_call(self, *, error: bool) -> None:
        with self.lock:
            self.api_calls += 1
            if error:
                self.api_errors += 1

    def record_exception(self) -> None:
        with self.lock:
            self.minute_counts["exceptions"] += 1
            self.error_events.append(time.time())

    def set_queue_depth_callback(self, callback) -> None:
        self.queue_depth_callback = callback

    def _trim_history(self, name: str, now: float, window: int = 3600) -> None:
        hist = self.event_history.get(name)
        if not hist:
            return
        while hist and now - hist[0] > window:
            hist.popleft()

    def _emit_loop(self) -> None:
        while not self.stop_event.wait(60):
            self.emit_metrics()

    def emit_metrics(self) -> None:
        now = time.time()
        with self.lock:
            elapsed = max(now - self.last_emit, 1)
            counts = dict(self.minute_counts)
            self.minute_counts.clear()
            rpc_latencies = list(self.rpc_latencies)
            self.rpc_latencies.clear()
            api_calls = self.api_calls
            api_errors = self.api_errors
            self.api_calls = 0
            self.api_errors = 0
            self.error_events = deque(
                ts for ts in self.error_events if now - ts <= self.error_window
            )
            last_pair_seen = self.last_pair_seen
            zero_alerted = self.zero_throughput_alerted
            self.last_emit = now

        avg_rpc_latency = (
            sum(rpc_latencies) / len(rpc_latencies) if rpc_latencies else 0.0
        )
        pairs_per_min = counts.get("pairs_scanned", 0) / (elapsed / 60)
        exceptions_per_min = counts.get("exceptions", 0) / (elapsed / 60)
        api_error_rate = (api_errors / api_calls) if api_calls else 0.0
        queue_depths = (
            self.queue_depth_callback() if self.queue_depth_callback else {}
        )

        context = {
            "pairs_scanned_per_min": round(pairs_per_min, 3),
            "passes": counts.get("passes", 0),
            "trades_placed": counts.get("trades_placed", 0),
            "exceptions_per_min": round(exceptions_per_min, 3),
            "queue_depth": queue_depths,
            "api_error_rate": round(api_error_rate, 4),
            "api_calls": api_calls,
            "average_rpc_latency_ms": round(avg_rpc_latency, 2),
        }

        log_event(logging.INFO, "metrics", "runtime_metrics", context=context)

        if (not zero_alerted) and (now - last_pair_seen >= self.zero_throughput_window):
            self.zero_throughput_alerted = True
            log_event(
                logging.WARNING,
                "alert",
                "No pairs scanned in the last 10 minutes",
                error="zero_throughput",
            )

        error_count = len(self.error_events)
        if error_count >= self.error_threshold and now - self.last_error_alert >= self.error_window:
            self.last_error_alert = now
            log_event(
                logging.WARNING,
                "alert",
                "Error spike detected",
                error="error_spike",
                context={"errors_last_5m": error_count},
            )

    def get_recent_counts(self, window: int = 3600) -> Dict[str, int]:
        now = time.time()
        with self.lock:
            result: Dict[str, int] = {}
            for name in self.event_history:
                self._trim_history(name, now, window)
                result[name] = len(self.event_history[name])
            return result

    def get_queue_depths(self) -> Dict[str, int]:
        callback = self.queue_depth_callback
        if not callback:
            return {}
        try:
            data = callback() or {}
        except Exception:
            logger.exception("queue_depth_callback_error")
            return {}
        return dict(data)

    def snapshot_daily_totals(self) -> Tuple[datetime, Dict[str, int]]:
        with self.lock:
            period_start = self.daily_period_start
            totals = dict(self.daily_totals)
            self.daily_totals = Counter()
            self.daily_period_start = datetime.utcnow()
            return period_start, totals


class RuntimeReporter:
    """Send periodic Telegram summaries about the bot runtime."""

    def __init__(self, metrics_collector: MetricsCollector):
        self.metrics = metrics_collector
        self.stop_event = threading.Event()
        self.next_hour = self._next_hour_boundary()
        self.next_midnight = self._next_midnight_boundary()
        threading.Thread(target=self._run, daemon=True).start()

    @staticmethod
    def _next_hour_boundary(reference: Optional[datetime] = None) -> datetime:
        ref = reference or datetime.utcnow()
        truncated = ref.replace(minute=0, second=0, microsecond=0)
        return truncated + timedelta(hours=1)

    @staticmethod
    def _next_midnight_boundary(reference: Optional[datetime] = None) -> datetime:
        ref = reference or datetime.utcnow()
        tomorrow = (ref + timedelta(days=1)).date()
        return datetime.combine(tomorrow, datetime.min.time())

    def _run(self) -> None:
        while not self.stop_event.wait(30):
            now = datetime.utcnow()
            if now >= self.next_hour:
                try:
                    self.send_hourly_update(now)
                except Exception:
                    logger.exception("hourly_runtime_report_error")
                self.next_hour = self._next_hour_boundary(now)
            if now >= self.next_midnight:
                try:
                    self.send_daily_summary(now)
                except Exception:
                    logger.exception("daily_runtime_report_error")
                self.next_midnight = self._next_midnight_boundary(now)

    def send_hourly_update(self, now: datetime) -> None:
        counts = self.metrics.get_recent_counts(3600)
        queue_depths = self.metrics.get_queue_depths()
        pending_rechecks = queue_depths.get("pending_rechecks", 0)
        volume_pending = queue_depths.get("volume_checks", 0)
        msg = (
            "⏱️ Hourly runtime report\n"
            f"Pairs scanned (1h): {counts.get('pairs_scanned', 0)}\n"
            f"Pairs passed (1h): {counts.get('passes', 0)}\n"
            f"Pairs pending recheck: {pending_rechecks}\n"
            f"Volume checks pending: {volume_pending}"
        )
        send_telegram_message(msg)

    def send_daily_summary(self, now: datetime) -> None:
        period_start, totals = self.metrics.snapshot_daily_totals()
        period_label = period_start.date().isoformat()
        msg = (
            "🗓️ Daily summary\n"
            f"Period starting {period_label}\n"
            f"Total pairs scanned: {totals.get('pairs_scanned', 0)}\n"
            f"Total pairs passed: {totals.get('passes', 0)}"
        )
        send_telegram_message(msg)


metrics = MetricsCollector()
runtime_reporter = RuntimeReporter(metrics)


PAIR_RECORD_LOCK = threading.Lock()


def init_excel(path: str):
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.title = "history"
        ws.append(
            [
                "pair",
                "token0",
                "token1",
                "detected_at",
                "passes",
                "total",
                "liquidity",
                "volume",
                "marketcap",
                "first_sell_address",
                "first_sell_block",
            ]
        )
        wb.save(path)


init_excel(EXCEL_FILE)


def store_pair_record(
    pair_addr: str,
    token0: str,
    token1: str,
    passes: int,
    total: int,
    extra: Optional[Dict] = None,
) -> None:
    """Persist a snapshot of pair evaluation results to the Excel history file."""

    extra = extra or {}
    detected_at = datetime.utcnow().replace(microsecond=0).isoformat()
    liquidity = extra.get("liquidityUsd") or extra.get("liquidity")
    volume = extra.get("volume24h") or extra.get("volume")
    marketcap = extra.get("marketCap") or extra.get("marketcap")
    first_sell_info = extra.get("firstSell") or {}

    row = [
        pair_addr,
        token0,
        token1,
        detected_at,
        passes,
        total,
        liquidity,
        volume,
        marketcap,
        first_sell_info.get("address"),
        first_sell_info.get("blockNumber"),
    ]

    try:
        with PAIR_RECORD_LOCK:
            if not os.path.exists(EXCEL_FILE):
                init_excel(EXCEL_FILE)
            workbook = load_workbook(EXCEL_FILE)
            worksheet = (
                workbook["history"]
                if "history" in workbook.sheetnames
                else workbook.active
            )
            worksheet.append(row)
            workbook.save(EXCEL_FILE)
    except Exception as exc:
        log_event(
            logging.ERROR,
            "pair_record_store",
            "Failed to persist pair record",
            pair=pair_addr,
            error=str(exc),
            context={
                "passes": passes,
                "total": total,
            },
        )


###########################################################
# 2. WEB3 & SESSION
###########################################################

try:
    w3_read, _READ_PROVIDER_INDEX = _init_read_provider(ALCHEMY_PROVIDER_URLS)
except Exception as exc:  # pragma: no cover - defensive startup guard
    _READ_PROVIDER_INDEX = 0
    w3_read = Web3(HTTPProvider(ALCHEMY_PROVIDER_URLS[0]))
    log_event(
        logging.ERROR,
        "rpc_provider_init",
        "reader provider initialization exception",
        error=str(exc),
        context={"url": _current_read_provider_url()},
    )

_EVENT_PROVIDER_URLS = _unique_urls(
    [INFURA_URL, INFURA_URL_BACKUP, *INFURA_EMERGENCY_URLS]
)
_EVENT_V3_PROVIDER_URLS = _unique_urls(
    [INFURA_URL_V3, INFURA_URL_V3_BACKUP, *INFURA_EMERGENCY_URLS]
)
_EVENT_PROVIDER_INDEX = -1
_EVENT_V3_PROVIDER_INDEX = -1


def _connect_provider(urls: List[str], label: str) -> Tuple[Web3, int]:
    for idx, url in enumerate(urls):
        try:
            provider = Web3(HTTPProvider(url))
            if provider.is_connected():
                if idx > 0:
                    log_event(
                        logging.WARNING,
                        "rpc_provider_init",
                        f"{label} provider fallback engaged",
                        context={"url": url, "index": idx},
                    )
                return provider, idx
        except Exception as exc:  # pragma: no cover - defensive connection guard
            log_event(
                logging.ERROR,
                "rpc_provider_init",
                f"{label} provider connection error",
                error=str(exc),
                context={"url": url, "index": idx},
            )
    if w3_read.is_connected():
        log_event(
            logging.ERROR,
            "rpc_provider_init",
            f"All {label} providers unavailable; using reader",
            context={"url": _current_read_provider_url()},
        )
        return w3_read, -1
    raise RuntimeError(f"Unable to establish {label} provider")


w3_event, _EVENT_PROVIDER_INDEX = _connect_provider(
    _EVENT_PROVIDER_URLS, "uniswap_v2"
)
w3_event_v3, _EVENT_V3_PROVIDER_INDEX = _connect_provider(
    _EVENT_V3_PROVIDER_URLS, "uniswap_v3"
)

FETCH_TIMEOUT = 30


def _rotate_rpc_provider(is_v3: bool, cause: Exception) -> Optional[Web3]:
    global w3_event, w3_event_v3, _EVENT_PROVIDER_INDEX, _EVENT_V3_PROVIDER_INDEX

    urls = _EVENT_V3_PROVIDER_URLS if is_v3 else _EVENT_PROVIDER_URLS
    index = _EVENT_V3_PROVIDER_INDEX if is_v3 else _EVENT_PROVIDER_INDEX
    label = "uniswap_v3" if is_v3 else "uniswap_v2"

    if not urls:
        return None

    if index is None or index < 0:
        order = list(range(len(urls)))
    else:
        order = list(range(index + 1, len(urls))) + list(range(0, index))

    for next_idx in order:
        url = urls[next_idx]
        try:
            provider = Web3(HTTPProvider(url))
            if provider.is_connected():
                if is_v3:
                    w3_event_v3 = provider
                    _EVENT_V3_PROVIDER_INDEX = next_idx
                else:
                    w3_event = provider
                    _EVENT_PROVIDER_INDEX = next_idx
                log_event(
                    logging.WARNING,
                    "rpc_provider_rotate",
                    f"Switched {label} provider",
                    context={"url": url, "reason": str(cause)},
                )
                return provider
        except Exception as exc:  # pragma: no cover - defensive connection guard
            log_event(
                logging.ERROR,
                "rpc_provider_rotate",
                f"Failed switching {label} provider",
                error=str(exc),
                context={"url": url},
            )

    if is_v3:
        _EVENT_V3_PROVIDER_INDEX = -1
    else:
        _EVENT_PROVIDER_INDEX = -1
    return None


async def _etherscan_get_async(params: dict, timeout: int = FETCH_TIMEOUT) -> dict:
    """Instrumented wrapper around the tracker Etherscan helper."""

    start = time.perf_counter()
    base_params = dict(params)
    prepared_params = _prepare_etherscan_params(base_params)
    success = False
    result: dict = {}
    last_error: Optional[Exception] = None
    max_attempts = ETHERSCAN_MAX_RETRIES + 1
    delay = ETHERSCAN_RETRY_BACKOFF_SECONDS

    for attempt in range(1, max_attempts + 1):
        attempt_start = time.perf_counter()
        try:
            result = await _perform_etherscan_request(prepared_params, timeout)
        except (aiohttp.ClientError, asyncio.TimeoutError, OSError, requests.RequestException) as exc:
            last_error = exc
            latency_ms = round((time.perf_counter() - attempt_start) * 1000, 2)
            metrics.record_exception()
            should_retry = attempt < max_attempts
            log_event(
                logging.WARNING if should_retry else logging.ERROR,
                "etherscan_api",
                "Etherscan request failed" + ("; retrying" if should_retry else ""),
                error=str(exc),
                context={
                    "module": prepared_params.get("module"),
                    "action": prepared_params.get("action"),
                    "attempt": attempt,
                },
                latency_ms=latency_ms,
            )
            if not should_retry:
                break
        else:
            success = str(result.get("status")) == "1"
            latency_ms = round((time.perf_counter() - attempt_start) * 1000, 2)
            retriable = not success and attempt < max_attempts and _is_retriable_etherscan_response(result)
            message = result.get("message")
            if not success:
                log_event(
                    logging.WARNING,
                    "etherscan_api",
                    "Etherscan returned non-success status" + ("; retrying" if retriable else ""),
                    error=message,
                    context={
                        "module": prepared_params.get("module"),
                        "action": prepared_params.get("action"),
                        "result": result.get("result"),
                        "attempt": attempt,
                    },
                    latency_ms=latency_ms,
                )
            if success or not retriable:
                break

        if attempt >= max_attempts:
            break

        if "apikey" in base_params and ETHERSCAN_API_KEY_LIST:
            base_params["apikey"] = get_next_etherscan_key()
        prepared_params = _prepare_etherscan_params(base_params)

        if delay > 0:
            await asyncio.sleep(delay)
            delay = min(delay * 2, ETHERSCAN_RETRY_BACKOFF_MAX_SECONDS)

    metrics.record_api_call(error=not success)

    if not success and not result and last_error is not None:
        raise last_error

    return result


def safe_block_number(is_v3: bool = False) -> int:
    """Return latest block number with automatic provider fallback."""
    global w3_event, w3_event_v3, _EVENT_PROVIDER_INDEX, _EVENT_V3_PROVIDER_INDEX
    provider = w3_event_v3 if is_v3 else w3_event
    network = "uniswap_v3" if is_v3 else "uniswap_v2"
    start = time.perf_counter()
    try:
        block = provider.eth.block_number
        latency_ms = round((time.perf_counter() - start) * 1000, 2)
        metrics.record_rpc_call(latency_ms)
        return block
    except Exception as e:
        latency_ms = round((time.perf_counter() - start) * 1000, 2)
        metrics.record_rpc_call(latency_ms, error=True)
        metrics.record_exception()
        msg = str(e).lower()
        if "429" in msg or "rate" in msg:
            log_event(
                logging.WARNING,
                "rpc_rate_limit",
                f"Rate limit on {network} provider",
                error=str(e),
                context={"network": network},
                latency_ms=latency_ms,
            )
        else:
            log_event(
                logging.ERROR,
                "rpc_failure",
                f"{network} block number error: {e}",
                error=str(e),
                context={"network": network},
                latency_ms=latency_ms,
            )

        fallback_provider = _rotate_rpc_provider(is_v3, e)
        if fallback_provider is not None:
            backup_start = time.perf_counter()
            try:
                block = fallback_provider.eth.block_number
                backup_latency = round(
                    (time.perf_counter() - backup_start) * 1000, 2
                )
                metrics.record_rpc_call(backup_latency)
                return block
            except Exception as e2:
                backup_latency = round(
                    (time.perf_counter() - backup_start) * 1000, 2
                )
                metrics.record_rpc_call(backup_latency, error=True)
                metrics.record_exception()
                log_event(
                    logging.ERROR,
                    "rpc_backup_failure",
                    f"Fallback {network} block number error: {e2}",
                    error=str(e2),
                    context={"network": network},
                    latency_ms=backup_latency,
                )

        if is_v3:
            w3_event_v3 = w3_read
            _EVENT_V3_PROVIDER_INDEX = -1
        else:
            w3_event = w3_read
            _EVENT_PROVIDER_INDEX = -1
        fallback_start = time.perf_counter()
        block = w3_read.eth.block_number
        fallback_latency = round((time.perf_counter() - fallback_start) * 1000, 2)
        metrics.record_rpc_call(fallback_latency)
        return block


def safe_get_logs(filter_params: dict, is_v3: bool = False) -> List[dict]:
    """Get logs with fallback to backup provider on rate limit."""
    global w3_event, w3_event_v3, _EVENT_PROVIDER_INDEX, _EVENT_V3_PROVIDER_INDEX
    provider = w3_event_v3 if is_v3 else w3_event
    network = "uniswap_v3" if is_v3 else "uniswap_v2"
    start = time.perf_counter()
    try:
        logs = provider.eth.get_logs(filter_params)
        latency_ms = round((time.perf_counter() - start) * 1000, 2)
        metrics.record_rpc_call(latency_ms)
        return logs
    except Exception as e:
        latency_ms = round((time.perf_counter() - start) * 1000, 2)
        metrics.record_rpc_call(latency_ms, error=True)
        metrics.record_exception()
        msg = str(e).lower()
        if "429" in msg or "rate" in msg:
            log_event(
                logging.WARNING,
                "rpc_rate_limit",
                f"Rate limit on {network} get_logs",
                error=str(e),
                context={"network": network},
                latency_ms=latency_ms,
            )
        else:
            log_event(
                logging.ERROR,
                "rpc_failure",
                f"{network} get_logs error: {e}",
                error=str(e),
                context={"network": network},
                latency_ms=latency_ms,
            )

        fallback_provider = _rotate_rpc_provider(is_v3, e)
        if fallback_provider is not None:
            backup_start = time.perf_counter()
            try:
                logs = fallback_provider.eth.get_logs(filter_params)
                backup_latency = round(
                    (time.perf_counter() - backup_start) * 1000, 2
                )
                metrics.record_rpc_call(backup_latency)
                return logs
            except Exception as e2:
                backup_latency = round(
                    (time.perf_counter() - backup_start) * 1000, 2
                )
                metrics.record_rpc_call(backup_latency, error=True)
                metrics.record_exception()
                log_event(
                    logging.ERROR,
                    "rpc_backup_failure",
                    f"Fallback {network} get_logs error: {e2}",
                    error=str(e2),
                    context={"network": network},
                    latency_ms=backup_latency,
                )

        if is_v3:
            w3_event_v3 = w3_read
            _EVENT_V3_PROVIDER_INDEX = -1
        else:
            w3_event = w3_read
            _EVENT_PROVIDER_INDEX = -1
        fallback_start = time.perf_counter()
        logs = w3_read.eth.get_logs(filter_params)
        fallback_latency = round((time.perf_counter() - fallback_start) * 1000, 2)
        metrics.record_rpc_call(fallback_latency)
        return logs



###########################################################
# 3. TELEGRAM
###########################################################

async def async_send_telegram_message(text: str, parse_mode: str = "HTML") -> bool:
    """Send a message to Telegram asynchronously."""
    if not TELEGRAM_BASE_URL or not TELEGRAM_CHAT_ID:
        return False
    url = f"{TELEGRAM_BASE_URL}/sendMessage"
    data = {
        "chat_id": TELEGRAM_CHAT_ID,
        "text": text,
        "parse_mode": parse_mode,
        "disable_web_page_preview": True,
    }
    try:
        async with create_aiohttp_session() as session:
            async with session.post(url, data=data, timeout=FETCH_TIMEOUT) as resp:
                payload = await resp.json()
                return bool(payload.get("ok"))
    except Exception as e:
        logger.error(f"Telegram send failed: {e}")
        return False

def send_telegram_message(text: str, parse_mode: str = "HTML") -> bool:
    """Safe to call from inside or outside an event loop."""
    try:
        loop = asyncio.get_running_loop()
    except RuntimeError:
        return asyncio.run(async_send_telegram_message(text, parse_mode))
    else:
        loop.create_task(async_send_telegram_message(text, parse_mode))
        return True

# Wire wallet tracker notifications to Telegram (Patch A)
try:
    set_notifier(send_telegram_message, async_send_telegram_message)
except Exception:
    pass




# ---------------------------------------------------------
# Telegram command listener (polling): /monitor_on, /monitor_off, /monitor_off_all, /monitor_list
# ---------------------------------------------------------
def _telegram_command_listener():
    if not TELEGRAM_BASE_URL or not TELEGRAM_CHAT_ID:
        logger.warning("Telegram command listener disabled (no credentials)")
        return
    offset = 0
    import time as _time, requests
    session = requests.Session()
    while True:
        try:
            resp = session.get(
                f"{TELEGRAM_BASE_URL}/getUpdates",
                params={"timeout": 25, "offset": offset},
                timeout=FETCH_TIMEOUT + 5,
            )
            data = resp.json() if resp.ok else {}
            for upd in data.get("result", []):
                offset = upd.get("update_id", 0) + 1
                msg = upd.get("message") or upd.get("edited_message") or {}
                chat_id = str(msg.get("chat", {}).get("id", ""))
                # Restrict to configured chat unless left blank
                if TELEGRAM_CHAT_ID and chat_id != str(TELEGRAM_CHAT_ID):
                    continue
                text = (msg.get("text") or "").strip()
                if not text.startswith("/"):
                    continue
                parts = text.split()
                cmd = parts[0].lower()
                arg = parts[1] if len(parts) > 1 else ""
                if cmd in ("/monitor_on", "/monitoron"):
                    token = _resolve_main_token_from_arg(arg)
                    if not token:
                        send_telegram_message("Usage: /monitor_on <token-or-pair-address>")
                        continue
                    start_wallet_monitor(token)
                    send_telegram_message(f"✅ Monitoring started for <code>{token}</code>")
                elif cmd in ("/monitor_off", "/monitoroff"):
                    if not arg:
                        send_telegram_message("Usage: /monitor_off <token-or-pair-address>")
                        continue
                    ok = stop_wallet_monitor(arg)
                    if ok:
                        send_telegram_message(f"🛑 Monitoring stopped for <code>{arg}</code>")
                    else:
                        send_telegram_message(f"ℹ️ Not monitoring <code>{arg}</code>")
                elif cmd in ("/monitor_off_all", "/monitoroffall"):
                    n = stop_all_wallet_monitors()
                    send_telegram_message(f"🧹 Stopped {n} wallet monitor(s).")
                elif cmd in ("/monitor_list", "/monitorlist"):
                    lst = list_wallet_monitors()
                    if not lst:
                        send_telegram_message("No active wallet monitors.")
                    else:
                        body = "\n".join(f"- <code>{x}</code>" for x in lst)
                        send_telegram_message(f"Active monitors ({len(lst)}):\n{body}")
        except Exception as e:
            logger.debug(f"telegram listener error: {e}")
        finally:
            _time.sleep(1)


# Start the Telegram command listener in a background thread
_threading.Thread(target=_telegram_command_listener, daemon=True).start()

###########################################################
# 4. FAIL BUFFER
###########################################################

FAILED_RECHECKS_BUFFER: List[str] = []


def add_failed_recheck_message(msg: str):
    FAILED_RECHECKS_BUFFER.append(msg)


def maybe_flush_failed_rechecks():
    if FAILED_RECHECKS_BUFFER:
        log_msg = "[FailedRechecks]\n" + "\n".join(FAILED_RECHECKS_BUFFER)
        logger.info(log_msg)
    FAILED_RECHECKS_BUFFER.clear()


###########################################################
# 5. HONEYPOT + RENOUNCE
###########################################################

HONEYPOT_API_URL = "https://api.honeypot.is/v2/IsHoneypot"


def check_honeypot_is(
    token_addr: str, pair_addr: str = None, chain_id: int = None
) -> bool:
    """Return True if Honeypot.is flags the token as a honeypot."""

    async def _check() -> bool:
        params = {"address": token_addr}
        if chain_id is not None:
            params["chainID"] = chain_id
        if pair_addr:
            params["pair"] = pair_addr

        try:
            async with create_aiohttp_session() as session:
                async with session.get(HONEYPOT_API_URL, params=params, timeout=15) as resp:
                    j = await resp.json()
            hp = j.get("honeypotResult", {}).get("isHoneypot")
            if hp is True:
                return True
            if hp is False:
                return False
        except Exception as e:
            logger.debug(f"honeypot api error: {e}")

        try:
            url = f"https://honeypot.is/ethereum?address={token_addr}"
            async with create_aiohttp_session() as session:
                async with session.get(url, timeout=15) as r:
                    text = await r.text()
            low = text.lower()
            if "honeypot detected" in low or "this token appears to be a honeypot" in low:
                return True
        except Exception as e:
            logger.debug(f"honeypot html error: {e}")
        return False

    return asyncio.run(_check())


def check_is_renounced(token_addr: str) -> bool:
    """Check if ownership has been renounced using multiple heuristics."""
    owner_addr, _, _ = get_owner_info(token_addr)
    if owner_addr and owner_addr.lower() == ZERO_ADDRESS.lower():
        return True
    if check_renounced_by_event(token_addr):
        return True
    return False


###########################################################
# 8. DEXSCREENER
###########################################################

DEXSCREENER_SEARCH_URL = "https://api.dexscreener.com/latest/dex/search"
DEXSCREENER_PAIR_URL = "https://api.dexscreener.com/latest/dex/pairs/ethereum"
DEXSCREENER_CACHE: Dict[str, Tuple[float, dict]] = {}
DEXSCREENER_CACHE_TTL = 300  # 5 minutes
DEXSCREENER_PAIR_TTL = 10  # short TTL for pair endpoint

UNCX_API_BASE_URL = os.getenv("UNCX_API_BASE_URL", "https://api.uncx.network/api/v1")
UNCX_LOCKS_ENDPOINT = os.getenv(
    "UNCX_LOCKS_ENDPOINT", f"{UNCX_API_BASE_URL.rstrip('/')}/locks/token"
)
UNCX_GRAPH_API_KEY = os.getenv(
    "UNCX_GRAPH_API_KEY",
    "5efb383687746448f735da8954047c4f",
)
UNCX_GRAPH_SUBGRAPH_ID = os.getenv(
    "UNCX_GRAPH_SUBGRAPH_ID",
    "5gByjbCu558gLVwzvWiYD8JPQC8KLk6PSe9AVFy8LC69",
)
UNCX_GRAPH_ENDPOINT_TEMPLATE = os.getenv(
    "UNCX_GRAPH_ENDPOINT_TEMPLATE",
    "https://gateway-arbitrum.network.thegraph.com/api/{api_key}/subgraphs/id/{subgraph_id}",
)

BURN_ADDRESSES: Set[str] = {
    ZERO_ADDRESS.lower(),
    "0x000000000000000000000000000000000000dead",
    "0x0000000000000000000000000000000000000001",
}

LOCKER_ADDRESS_HINTS: Dict[str, str] = {}
LOCKER_KEYWORDS: Tuple[str, ...] = (
    "locker",
    "liquiditylock",
    "liquidity lock",
    "unicrypt",
    "team finance",
    "teamfinance",
    "pinklock",
    "pink lock",
    "pinksale",
    "mudra",
    "deeplock",
    "gempad",
    "safeswap",
)
LP_HOLDER_SNAPSHOT_LIMIT = 20
LOCKER_FRESHNESS_BUFFER_SECONDS = 60


@dataclass
class LPHolderAnalysis:
    address: str
    balance: int
    status: str
    reason: str
    unlock_timestamp: Optional[int] = None


@dataclass
class LiquidityLockDetails:
    source: str
    coverage_pct: Optional[float] = None
    locked_at: Optional[int] = None
    unlock_at: Optional[int] = None

    def as_dict(self) -> Dict[str, Optional[Union[str, float, int]]]:
        duration: Optional[int] = None
        if self.locked_at and self.unlock_at and self.unlock_at > self.locked_at:
            duration = int(self.unlock_at - self.locked_at)
        return {
            "source": self.source,
            "coveragePct": self.coverage_pct,
            "lockedAt": self.locked_at,
            "unlockAt": self.unlock_at,
            "lockDurationSeconds": duration,
        }


def _normalize_timestamp_seconds(value: Any) -> Optional[int]:
    if value in (None, ""):
        return None
    try:
        if isinstance(value, str):
            trimmed = value.strip()
            if not trimmed:
                return None
            ts = float(Decimal(trimmed))
        else:
            ts = float(value)
    except (InvalidOperation, ValueError, TypeError):
        return None

    if ts <= 0:
        return None

    if ts > 1e12:  # milliseconds
        ts = ts / 1000.0

    try:
        return int(ts)
    except (ValueError, TypeError):
        return None


def _normalize_percent_ratio(value: Any) -> Optional[float]:
    if value in (None, ""):
        return None
    try:
        if isinstance(value, str):
            pct = float(Decimal(value.strip()))
        else:
            pct = float(value)
    except (InvalidOperation, ValueError, TypeError):
        return None

    if pct > 1:
        pct = pct / 100.0

    if pct < 0 or pct > 1:
        return None

    return pct


_LP_CONTRACT_NAME_CACHE: Dict[str, Optional[str]] = {}
_LP_IS_CONTRACT_CACHE: Dict[str, bool] = {}


async def _fetch_lp_total_supply_async(pair_addr: str) -> Optional[int]:
    try:
        checksum = to_checksum_address(pair_addr)
    except ValueError:
        return None

    abi = [
        {
            "constant": True,
            "inputs": [],
            "name": "totalSupply",
            "outputs": [{"name": "", "type": "uint256"}],
            "stateMutability": "view",
            "type": "function",
        }
    ]

    contract = w3_read.eth.contract(checksum, abi=abi)
    loop = asyncio.get_running_loop()
    try:
        total_supply = await loop.run_in_executor(
            None, lambda: contract.functions.totalSupply().call()
        )
    except Exception as exc:
        logger.debug(f"lp totalSupply fetch failed for %s: %s", pair_addr, exc)
        return None

    if isinstance(total_supply, int) and total_supply > 0:
        return total_supply
    return None


PAIR_TOKEN_ABI = [
    {
        "constant": True,
        "inputs": [],
        "name": "token0",
        "outputs": [{"name": "", "type": "address"}],
        "stateMutability": "view",
        "type": "function",
    },
    {
        "constant": True,
        "inputs": [],
        "name": "token1",
        "outputs": [{"name": "", "type": "address"}],
        "stateMutability": "view",
        "type": "function",
    },
]


async def _fetch_pair_tokens_async(pair_addr: str) -> Tuple[Optional[str], Optional[str]]:
    cached = known_pairs.get(pair_addr.lower())
    if cached:
        return cached

    try:
        checksum = to_checksum_address(pair_addr)
    except ValueError:
        return None, None

    contract = w3_read.eth.contract(checksum, abi=PAIR_TOKEN_ABI)
    loop = asyncio.get_running_loop()

    try:
        token0 = await loop.run_in_executor(
            None, lambda: contract.functions.token0().call()
        )
        token1 = await loop.run_in_executor(
            None, lambda: contract.functions.token1().call()
        )
        token0 = to_checksum_address(token0)
        token1 = to_checksum_address(token1)
        known_pairs[pair_addr.lower()] = (token0, token1)
        return token0, token1
    except Exception as exc:
        logger.debug(f"pair token fetch failed for %s: %s", pair_addr, exc)
        return None, None


def _extract_int_from_result(value: Any) -> Optional[int]:
    if isinstance(value, int):
        return value
    if isinstance(value, str):
        try:
            return int(value, 16) if value.startswith("0x") else int(value)
        except ValueError:
            return None
    if isinstance(value, HexBytes):
        return int.from_bytes(value, "big")
    if isinstance(value, (list, tuple)):
        extracted: List[int] = []
        for item in value:
            inner = _extract_int_from_result(item)
            if inner is not None:
                extracted.append(inner)
        if extracted:
            return max(extracted)
    return None


async def _is_contract_async(address: str) -> bool:
    lowered = address.lower()
    cached = _LP_IS_CONTRACT_CACHE.get(lowered)
    if cached is not None:
        return cached

    try:
        checksum = to_checksum_address(address)
    except ValueError:
        _LP_IS_CONTRACT_CACHE[lowered] = False
        return False

    loop = asyncio.get_running_loop()
    try:
        code = await loop.run_in_executor(None, lambda: w3_read.eth.get_code(checksum))
        is_contract = bool(code and code != HexBytes(b""))
    except Exception as exc:
        logger.debug(f"is_contract check failed for %s: %s", address, exc)
        is_contract = False

    _LP_IS_CONTRACT_CACHE[lowered] = is_contract
    return is_contract


async def _get_contract_name_async(address: str) -> Optional[str]:
    lowered = address.lower()
    if lowered in _LP_CONTRACT_NAME_CACHE:
        return _LP_CONTRACT_NAME_CACHE[lowered]

    info = await _fetch_contract_source_etherscan_async(address)
    name = (info.get("contractName") or info.get("ContractName") or "").strip()
    if not name:
        name = None
    _LP_CONTRACT_NAME_CACHE[lowered] = name
    return name


def _generate_locker_arguments(
    arg_types: Tuple[str, ...],
    pair_addr: Optional[str],
    main_token: Optional[str],
) -> List[List[Any]]:
    if not arg_types:
        return [[]]

    options: List[List[Any]] = []
    for arg_type in arg_types:
        if arg_type == "address":
            candidates: List[str] = []
            if pair_addr:
                candidates.append(pair_addr)
            if main_token:
                lower_set = {c.lower() for c in candidates}
                if main_token.lower() not in lower_set:
                    candidates.append(main_token)
            if not candidates:
                return []
            options.append(candidates)
        elif arg_type.startswith("uint"):
            options.append([0])
        else:
            return []

    combos: List[List[Any]] = []
    for combination in product(*options):
        combos.append(list(combination))
    return combos


async def _call_locker_uint_function_async(
    contract_addr: str, fn_name: str, arg_types: Tuple[str, ...], args: List[Any]
) -> Optional[int]:
    try:
        checksum = to_checksum_address(contract_addr)
    except ValueError:
        return None

    abi_inputs = []
    for idx, arg_type in enumerate(arg_types):
        abi_inputs.append({"name": f"arg{idx}", "type": arg_type})

    abi = [
        {
            "constant": True,
            "inputs": abi_inputs,
            "name": fn_name,
            "outputs": [{"name": "", "type": "uint256"}],
            "stateMutability": "view",
            "type": "function",
        }
    ]

    prepared_args: List[Any] = []
    for arg_type, raw_value in zip(arg_types, args):
        if arg_type == "address":
            try:
                prepared_args.append(to_checksum_address(raw_value))
            except ValueError:
                return None
        else:
            prepared_args.append(raw_value)

    contract = w3_read.eth.contract(checksum, abi=abi)
    loop = asyncio.get_running_loop()
    try:
        result = await loop.run_in_executor(
            None, lambda: getattr(contract.functions, fn_name)(*prepared_args).call()
        )
    except Exception:
        return None

    return _extract_int_from_result(result)


LOCKER_TIME_METHODS: Tuple[Tuple[str, Tuple[str, ...]], ...] = (
    ("unlockDate", tuple()),
    ("unlockTime", tuple()),
    ("getUnlockTime", tuple()),
    ("getUnlockTime", ("address",)),
    ("getUnlockTime", ("address", "address")),
    ("releaseTime", tuple()),
    ("lockReleaseTime", tuple()),
    ("unlockTimestamp", tuple()),
    ("unlockAt", tuple()),
    ("lockTime", tuple()),
    ("endTime", tuple()),
)


async def _get_locker_unlock_timestamp_async(
    locker_addr: str, pair_addr: str, main_token: Optional[str]
) -> Optional[int]:
    for method_name, arg_types in LOCKER_TIME_METHODS:
        argument_sets = _generate_locker_arguments(arg_types, pair_addr, main_token)
        for args in argument_sets:
            ts = await _call_locker_uint_function_async(locker_addr, method_name, arg_types, args)
            if ts is None:
                continue
            if ts > 0:
                return ts
    return None


async def _analyze_lp_holder_async(
    holder_addr: str,
    balance: int,
    pair_addr: str,
    main_token: Optional[str],
) -> LPHolderAnalysis:
    lowered = holder_addr.lower()
    if lowered in BURN_ADDRESSES:
        return LPHolderAnalysis(holder_addr, balance, "locked", "burn_address")

    if lowered == pair_addr.lower():
        return LPHolderAnalysis(holder_addr, balance, "unknown", "pair_contract")

    hinted = LOCKER_ADDRESS_HINTS.get(lowered)

    if hinted:
        unlock_ts = await _get_locker_unlock_timestamp_async(holder_addr, pair_addr, main_token)
        if unlock_ts is not None and unlock_ts <= int(time.time()) + LOCKER_FRESHNESS_BUFFER_SECONDS:
            return LPHolderAnalysis(
                holder_addr,
                balance,
                "unlocked",
                f"{hinted}_locker_expired",
                unlock_ts,
            )
        return LPHolderAnalysis(
            holder_addr,
            balance,
            "locked",
            f"{hinted}_locker",
            unlock_ts,
        )

    is_contract = await _is_contract_async(holder_addr)
    if not is_contract:
        return LPHolderAnalysis(holder_addr, balance, "unlocked", "externally_owned_account")

    contract_name = await _get_contract_name_async(holder_addr)
    lowered_name = contract_name.lower() if contract_name else ""

    if lowered_name:
        if any(keyword in lowered_name for keyword in LOCKER_KEYWORDS):
            unlock_ts = await _get_locker_unlock_timestamp_async(
                holder_addr, pair_addr, main_token
            )
            if unlock_ts is not None and unlock_ts <= int(time.time()) + LOCKER_FRESHNESS_BUFFER_SECONDS:
                return LPHolderAnalysis(
                    holder_addr,
                    balance,
                    "unlocked",
                    "locker_unlock_expired",
                    unlock_ts,
                )
            return LPHolderAnalysis(
                holder_addr,
                balance,
                "locked",
                "locker_detected",
                unlock_ts,
            )
        if "burn" in lowered_name:
            return LPHolderAnalysis(holder_addr, balance, "locked", "burn_contract")

    return LPHolderAnalysis(holder_addr, balance, "unknown", "unclassified_contract")


async def _check_liquidity_locked_holder_analysis(
    pair_addr: str,
) -> Tuple[Optional[bool], Optional[LiquidityLockDetails]]:
    total_supply = await _fetch_lp_total_supply_async(pair_addr)
    if not total_supply:
        return None, None

    holders_raw = await _fetch_holder_distribution_async(
        pair_addr, LP_HOLDER_SNAPSHOT_LIMIT
    )
    if not holders_raw:
        return None, None

    token0, token1 = await _fetch_pair_tokens_async(pair_addr)
    main_token: Optional[str] = None
    if token0 and token1:
        try:
            main_token = get_non_weth_token(token0, token1)
        except Exception:
            main_token = token0

    analyses: List[LPHolderAnalysis] = []
    for entry in holders_raw:
        address = entry.get("address")
        if not address:
            continue

        balance_value = entry.get("balance")
        balance_int: Optional[int] = None
        if balance_value is not None:
            try:
                balance_int = int(balance_value)
            except (TypeError, ValueError):
                balance_int = _parse_holder_balance(balance_value)

        if balance_int is None:
            share = entry.get("share")
            if share is not None:
                try:
                    balance_int = int(Decimal(total_supply) * Decimal(str(share)))
                except Exception:
                    balance_int = None

        if balance_int is None or balance_int <= 0:
            continue

        try:
            analysis = await _analyze_lp_holder_async(
                address, balance_int, pair_addr, main_token
            )
        except Exception as exc:
            logger.debug(
                "lp holder analysis failed for %s on %s: %s",
                address,
                pair_addr,
                exc,
            )
            analysis = LPHolderAnalysis(address, balance_int, "unknown", "analysis_error")

        analyses.append(analysis)

    if not analyses:
        return None, None

    locked_balance = sum(a.balance for a in analyses if a.status == "locked")
    unlocked_balance = sum(a.balance for a in analyses if a.status == "unlocked")
    unknown_balance = sum(a.balance for a in analyses if a.status == "unknown")

    covered = locked_balance + unlocked_balance + unknown_balance
    if covered < total_supply:
        unknown_balance += total_supply - covered

    snapshot = [
        {
            "address": a.address,
            "status": a.status,
            "balance": a.balance,
            "reason": a.reason,
            "unlock": a.unlock_timestamp,
        }
        for a in analyses
    ]
    logger.debug(
        "LP holder snapshot for %s => locked=%s unlocked=%s unknown=%s total=%s details=%s",
        pair_addr,
        locked_balance,
        unlocked_balance,
        unknown_balance,
        total_supply,
        snapshot,
    )

    coverage_pct: Optional[float] = None
    if total_supply:
        try:
            coverage_pct = locked_balance / float(total_supply)
        except ZeroDivisionError:
            coverage_pct = None

    future_unlocks: List[int] = []
    for analysis in analyses:
        if analysis.status != "locked":
            continue
        ts = _normalize_timestamp_seconds(analysis.unlock_timestamp)
        if ts:
            future_unlocks.append(ts)
    unlock_ts = min(future_unlocks) if future_unlocks else None

    if locked_balance * 100 >= total_supply * 95:
        logger.debug("LP supply for %s considered locked via holder analysis", pair_addr)
        details = LiquidityLockDetails(
            source="holder_analysis",
            coverage_pct=coverage_pct,
            unlock_at=unlock_ts,
        )
        return True, details

    if unlocked_balance * 100 >= total_supply * 5:
        logger.debug(
            "LP supply for %s considered unlocked via holder analysis", pair_addr
        )
        return False, None

    return None, None


async def _check_liquidity_locked_etherscan_async(
    pair_addr: str,
) -> Tuple[bool, Optional[LiquidityLockDetails]]:
    holder_based, holder_details = await _check_liquidity_locked_holder_analysis(pair_addr)
    if holder_based is not None:
        return holder_based, holder_details

    uncx_based, uncx_details = await _check_liquidity_locked_uncx_async(pair_addr)
    if uncx_based is not None:
        return uncx_based, uncx_details

    if not ETHERSCAN_LOOKUPS_ENABLED:
        return False, None
    api_key = get_next_etherscan_key()
    if not api_key:
        return False, None
    params = {
        "module": "account",
        "action": "tokentx",
        "contractaddress": pair_addr,
        "page": 1,
        "offset": 20,
        "sort": "asc",
        "apikey": api_key,
    }
    try:
        j = await _etherscan_get_async(params, FETCH_TIMEOUT)
        if j.get("status") != "1":
            return False, None

        inspected_contracts: Dict[str, str] = {}
        for tx in j.get("result", []):
            to_addr = (tx.get("to") or "").lower()
            from_addr = (tx.get("from") or "").lower()
            func_name = (tx.get("functionName") or "").lower()
            lock_ts = _normalize_timestamp_seconds(tx.get("timeStamp"))

            if to_addr in {
                ZERO_ADDRESS.lower(),
                "0x000000000000000000000000000000000000dead",
            }:
                details = LiquidityLockDetails(
                    source="etherscan_tokentx",
                    locked_at=lock_ts,
                )
                return True, details

            if "lock" in func_name and "unlock" not in func_name:
                details = LiquidityLockDetails(
                    source="etherscan_tokentx",
                    locked_at=lock_ts,
                )
                return True, details

            if from_addr == pair_addr.lower():
                continue

            if not to_addr:
                continue

            cached_name = inspected_contracts.get(to_addr)
            if cached_name is None:
                info = fetch_contract_source_etherscan(to_addr)
                cached_name = (info.get("contractName") or "").lower()
                inspected_contracts[to_addr] = cached_name

            if any(
                keyword in cached_name for keyword in ["lock", "locker", "unicrypt", "team", "pink"]
            ):
                details = LiquidityLockDetails(
                    source="etherscan_tokentx",
                    locked_at=lock_ts,
                )
                return True, details
    except (aiohttp.ClientError, asyncio.TimeoutError, OSError) as e:
        disable_etherscan_lookups(f"lock lookup failed: {e}")
    except Exception as e:
        logger.debug(f"etherscan lock check error: {e}")
    return False, None

def check_liquidity_locked_etherscan(pair_addr: str) -> bool:
    locked, _ = asyncio.run(_check_liquidity_locked_etherscan_async(pair_addr))
    return locked


async def _fetch_dexscreener_data_async(
    token_addr: str, pair_addr: str
) -> Tuple[Optional[dict], Optional[str]]:
    now = time.time()
    pair_info = None
    pdata: Optional[dict] = None
    reason: Optional[str] = None

    async def _get_json(url: str) -> Tuple[Optional[dict], Optional[str]]:
        try:
            async with create_aiohttp_session() as session:
                async with session.get(url, timeout=FETCH_TIMEOUT) as resp:
                    resp.raise_for_status()
                    data = await resp.json()
            metrics.record_api_call(error=False)
            return data, None
        except aiohttp.ClientResponseError as exc:
            metrics.record_api_call(error=True)
            if exc.status == 404:
                return None, "not_listed"
            if exc.status == 429:
                return None, "rate_limited"
            return None, f"http_{exc.status}"
        except (aiohttp.ClientError, asyncio.TimeoutError, OSError):
            metrics.record_api_call(error=True)
            return None, "network_error"
        except Exception as exc:  # pragma: no cover - unexpected parser issues
            metrics.record_api_call(error=True)
            logger.debug(f"dexscreener fetch error: {exc}")
            return None, "unexpected_error"

    # --- token-based lookup (preferred) ---
    key = token_addr.lower()
    cached = DEXSCREENER_CACHE.get(key)
    if cached and now - cached[0] < DEXSCREENER_CACHE_TTL:
        jdata = cached[1]
        reason = None
    else:
        url = f"{DEXSCREENER_SEARCH_URL}?q={token_addr}"
        jdata, reason = await _get_json(url)
        if jdata:
            DEXSCREENER_CACHE[key] = (now, jdata)

    pairs = jdata.get("pairs", []) if jdata else []
    for p in pairs:
        if p.get("pairAddress", "").lower() == pair_addr.lower():
            pair_info = p
            break

    # --- fallback to pair endpoint when token search fails ---
    if not pair_info and reason != "rate_limited":
        pkey = f"pair:{pair_addr.lower()}"
        cached_pair = DEXSCREENER_CACHE.get(pkey)
        if cached_pair and now - cached_pair[0] < DEXSCREENER_PAIR_TTL:
            pdata = cached_pair[1]
            pair_reason = None
        else:
            url = f"{DEXSCREENER_PAIR_URL}/{pair_addr}"
            pdata, pair_reason = await _get_json(url)
            if pdata:
                DEXSCREENER_CACHE[pkey] = (now, pdata)
            reason = pair_reason or reason

    plist = pdata.get("pairs", []) if pdata else []
    if plist:
        pair_info = plist[0]
    elif not reason:
        reason = "not_listed"

    locked, lock_details = await _check_liquidity_locked_etherscan_async(pair_addr)

    if not pair_info:
        return None, reason or "not_listed"

    price_usd = float(pair_info.get("priceUsd", 0) or 0)
    liq_usd = float(pair_info.get("liquidity", {}).get("usd", 0) or 0)
    vol_24h = float(pair_info.get("volume", {}).get("h24", 0) or 0)
    fdv = float(pair_info.get("fdv", 0) or 0)
    mc = float(pair_info.get("marketCap", 0) or 0)

    tx24 = pair_info.get("txns", {}).get("h24", {})
    buys_24 = int(tx24.get("buys", 0))
    sells_24 = int(tx24.get("sells", 0))

    base_token = pair_info.get("baseToken", {})
    base_name = base_token.get("name", "").strip()
    base_symbol = base_token.get("symbol", "").strip()
    info_section = pair_info.get("info", {})
    pair_created_at = _normalize_timestamp_seconds(pair_info.get("pairCreatedAt"))
    logo_url = info_section.get("imageUrl", "")
    websites = [w.get("url") for w in info_section.get("websites", []) if w.get("url")]
    socials = [s.get("url") for s in info_section.get("socials", []) if s.get("url")]
    # Determine locked liquidity prioritising our internal checks before DexScreener labels
    labels = pair_info.get("labels", [])
    if not locked and "locked" in labels:
        locked = True
        if not lock_details:
            lock_details = LiquidityLockDetails(source="dexscreener_label")

    # Detect paid promotions/trending on Dex platforms
    dex_paid = any(lbl.lower() in {"promoted", "boosted", "paid"} for lbl in labels)

    return {
        "priceUsd": price_usd,
        "liquidityUsd": liq_usd,
        "volume24h": vol_24h,
        "fdv": fdv,
        "marketCap": mc,
        "buys": buys_24,
        "sells": sells_24,
        "baseTokenName": base_name,
        "baseTokenSymbol": base_symbol,
        "baseTokenLogo": logo_url,
        "socialLinks": websites + socials,
        "lockedLiquidity": locked,
        "lockedLiquidityDetails": lock_details.as_dict() if lock_details else None,
        "pairCreatedAt": pair_created_at,
        "dexPaid": dex_paid,
    }, None


def fetch_dexscreener_data(
    token_addr: str, pair_addr: str, *, with_reason: bool = False
) -> Union[Optional[dict], Tuple[Optional[dict], Optional[str]]]:
    data, reason = asyncio.run(_fetch_dexscreener_data_async(token_addr, pair_addr))
    if with_reason:
        return data, reason
    return data


async def _check_recent_liquidity_removal_async(pair_addr: str, timeframe_sec: int = 600) -> bool:
    if not ETHERSCAN_LOOKUPS_ENABLED:
        return False
    api_key = get_next_etherscan_key()
    if not api_key:
        return False
    params = {
        "module": "account",
        "action": "tokentx",
        "contractaddress": pair_addr,
        "page": 1,
        "offset": 10,
        "sort": "desc",
        "apikey": api_key,
    }
    try:
        params = _prepare_etherscan_params(params)
        async with create_aiohttp_session() as session:
            async with session.get(ETHERSCAN_API_URL, params=params, timeout=FETCH_TIMEOUT) as r:
                j = await r.json()
        metrics.record_api_call(error=False)
        if j.get("status") != "1":
            return False
        now_ts = time.time()
        for tx in j.get("result", []):
            if tx.get("to", "").lower() == ZERO_ADDRESS.lower():
                ts = int(tx.get("timeStamp", "0"))
                if now_ts - ts <= timeframe_sec:
                    return True
    except (aiohttp.ClientError, asyncio.TimeoutError, OSError) as e:
        metrics.record_api_call(error=True)
        disable_etherscan_lookups(f"liquidity removal lookup failed: {e}")
    except Exception as e:
        logger.debug(f"Etherscan removal check error: {e}")
    return False

def check_recent_liquidity_removal(pair_addr: str, timeframe_sec: int = 600) -> bool:
    return asyncio.run(_check_recent_liquidity_removal_async(pair_addr, timeframe_sec))


###########################################################
# 9. ADVANCED CONTRACT VERIFICATION
#    (Fetch from Etherscan, analyze code, detect high risk)
###########################################################

# Example: Set your Etherscan API keys if you want to do real contract checks
ETHERSCAN_API_KEY_LIST = load_etherscan_keys()
_etherscan_key_getter = make_key_getter(ETHERSCAN_API_KEY_LIST)


def get_next_etherscan_key() -> str:
    return _etherscan_key_getter()


ETHERSCAN_API_URL_CANDIDATES = load_etherscan_base_urls()
ETHERSCAN_API_URL = (
    ETHERSCAN_API_URL_CANDIDATES[0] if ETHERSCAN_API_URL_CANDIDATES else ""
)
ETHERSCAN_CHAIN_ID = os.getenv("ETHERSCAN_CHAIN_ID", "1")

ETHERSCAN_MAX_RETRIES = int(os.getenv("ETHERSCAN_MAX_RETRIES", "2") or "0")
ETHERSCAN_RETRY_BACKOFF_SECONDS = float(
    os.getenv("ETHERSCAN_RETRY_BACKOFF_SECONDS", "0.5") or "0"
)
ETHERSCAN_RETRY_BACKOFF_MAX_SECONDS = float(
    os.getenv("ETHERSCAN_RETRY_BACKOFF_MAX_SECONDS", "5") or "0"
)
_ETHERSCAN_RETRYABLE_TERMS = (
    "max rate limit",
    "rate limit reached",
    "rate limit exceeded",
    "busy",
    "timeout",
    "temporarily unavailable",
    "please try again later",
)


def _prepare_etherscan_params(params: dict, url: Optional[str] = None) -> dict:
    """Return request parameters augmented with ``chainid`` when required."""

    prepared = dict(params)
    target_url = (url or ETHERSCAN_API_URL or "").lower()
    if "/v2/" in target_url and "chainid" not in prepared:
        prepared["chainid"] = ETHERSCAN_CHAIN_ID
    return prepared


def _is_retriable_etherscan_response(payload: dict) -> bool:
    """Return ``True`` when the payload suggests a transient Etherscan issue."""

    message = str(payload.get("message") or "").lower()
    result_field = payload.get("result")
    if isinstance(result_field, str):
        result_text = result_field.lower()
    elif isinstance(result_field, list) and result_field and isinstance(result_field[0], str):
        result_text = " ".join(str(item).lower() for item in result_field)
    else:
        result_text = str(result_field or "").lower()

    combined = f"{message} {result_text}".strip()
    return any(term in combined for term in _ETHERSCAN_RETRYABLE_TERMS)


async def _perform_direct_etherscan_request(
    params: dict, timeout: int
) -> dict:
    async with create_aiohttp_session() as session:
        async with session.get(
            ETHERSCAN_API_URL, params=params, timeout=timeout
        ) as resp:
            resp.raise_for_status()
            return await resp.json()


async def _perform_etherscan_request(params: dict, timeout: int) -> dict:
    """Execute a single Etherscan request using the tracker when available."""

    if tracker_etherscan_get_async is not None:
        result = await tracker_etherscan_get_async(params, timeout)
        message = str(result.get("message") or "").lower()
        if (
            str(result.get("status")) != "1"
            and "etherscan lookups disabled" in message
        ):
            return await _perform_direct_etherscan_request(params, timeout)
        return result

    return await _perform_direct_etherscan_request(params, timeout)


def _env_flag(name: str, default: bool = True) -> bool:
    val = os.getenv(name)
    if val is None:
        return default
    return val.strip().lower() not in {"0", "false", "no", "off"}


ETHERSCAN_LOOKUPS_ENABLED = _env_flag("ENABLE_ETHERSCAN_LOOKUPS", True)
USE_ETHERSCAN_TOKEN_HOLDERS = _env_flag("ENABLE_ETHERSCAN_TOKEN_HOLDERS", True)
ETHERSCAN_DISABLED_REASON: Optional[str] = None
ETHERSCAN_RECOVERY_DELAY_SECONDS = int(
    os.getenv("ETHERSCAN_RECOVERY_DELAY_SECONDS", "300") or "0"
)
_ETHERSCAN_RECOVERY_LOCK = threading.Lock()
_ETHERSCAN_RECOVERY_TIMER: Optional[threading.Timer] = None


def _attempt_etherscan_recovery() -> None:
    """Try to re-enable Etherscan lookups after a cooldown."""

    global ETHERSCAN_LOOKUPS_ENABLED, ETHERSCAN_DISABLED_REASON, _ETHERSCAN_RECOVERY_TIMER

    with _ETHERSCAN_RECOVERY_LOCK:
        _ETHERSCAN_RECOVERY_TIMER = None

    if ETHERSCAN_LOOKUPS_ENABLED:
        return

    previous_reason = ETHERSCAN_DISABLED_REASON
    logger.info("Attempting to re-enable Etherscan lookups after cooldown")

    ETHERSCAN_LOOKUPS_ENABLED = True
    ETHERSCAN_DISABLED_REASON = None

    try:
        set_tracker_etherscan_enabled(True, "Etherscan recovery attempt")
    except Exception:  # pragma: no cover - defensive
        logger.debug("Failed to propagate Etherscan recovery to tracker", exc_info=True)

    try:
        ensure_etherscan_connectivity()
    except Exception as exc:  # pragma: no cover - defensive
        logger.warning("Etherscan recovery health check raised: %s", exc, exc_info=True)

    if ETHERSCAN_LOOKUPS_ENABLED:
        log_event(
            logging.INFO,
            "etherscan_recovery",
            "Etherscan lookups re-enabled",
            context={"previous_reason": previous_reason or ""},
        )
    else:
        logger.warning("Etherscan recovery attempt failed; rescheduling")
        schedule_etherscan_recovery()


def schedule_etherscan_recovery() -> None:
    """Schedule a cooldown task to retry enabling Etherscan lookups."""

    if ETHERSCAN_RECOVERY_DELAY_SECONDS <= 0:
        return

    with _ETHERSCAN_RECOVERY_LOCK:
        global _ETHERSCAN_RECOVERY_TIMER
        if _ETHERSCAN_RECOVERY_TIMER is not None:
            _ETHERSCAN_RECOVERY_TIMER.cancel()
        timer = threading.Timer(ETHERSCAN_RECOVERY_DELAY_SECONDS, _attempt_etherscan_recovery)
        timer.daemon = True
        _ETHERSCAN_RECOVERY_TIMER = timer

    timer.start()


def disable_etherscan_lookups(reason: str) -> None:
    global ETHERSCAN_LOOKUPS_ENABLED, ETHERSCAN_DISABLED_REASON
    if ETHERSCAN_LOOKUPS_ENABLED:
        ETHERSCAN_LOOKUPS_ENABLED = False
        ETHERSCAN_DISABLED_REASON = reason
        logger.warning("Disabling Etherscan lookups: %s", reason)
        try:
            set_tracker_etherscan_enabled(False, reason)
        except Exception:  # pragma: no cover - defensive
            logger.debug("Failed to propagate Etherscan disable to tracker", exc_info=True)
    else:
        ETHERSCAN_DISABLED_REASON = reason

    schedule_etherscan_recovery()


###########################################################
# UNCX LOCK LOOKUPS
###########################################################

UNCX_LOOKUPS_ENABLED = _env_flag("ENABLE_UNCX_LOOKUPS", True)
UNCX_DISABLED_REASON: Optional[str] = None
UNCX_RECOVERY_DELAY_SECONDS = int(os.getenv("UNCX_RECOVERY_DELAY_SECONDS", "300") or "0")
_UNCX_RECOVERY_LOCK = threading.Lock()
_UNCX_RECOVERY_TIMER: Optional[threading.Timer] = None


def _attempt_uncx_recovery() -> None:
    """Attempt to re-enable Uncx lookups after a cooldown."""

    global UNCX_LOOKUPS_ENABLED, UNCX_DISABLED_REASON, _UNCX_RECOVERY_TIMER

    with _UNCX_RECOVERY_LOCK:
        _UNCX_RECOVERY_TIMER = None

    if UNCX_LOOKUPS_ENABLED:
        return

    logger.info("Re-enabling Uncx lookups after cooldown")
    UNCX_LOOKUPS_ENABLED = True
    UNCX_DISABLED_REASON = None


def schedule_uncx_recovery() -> None:
    """Schedule a background task to re-enable Uncx lookups."""

    if UNCX_RECOVERY_DELAY_SECONDS <= 0:
        return

    with _UNCX_RECOVERY_LOCK:
        global _UNCX_RECOVERY_TIMER
        if _UNCX_RECOVERY_TIMER is not None:
            _UNCX_RECOVERY_TIMER.cancel()
        timer = threading.Timer(UNCX_RECOVERY_DELAY_SECONDS, _attempt_uncx_recovery)
        timer.daemon = True
        _UNCX_RECOVERY_TIMER = timer

    timer.start()


def disable_uncx_lookups(reason: str) -> None:
    """Disable Uncx lookups following persistent errors."""

    global UNCX_LOOKUPS_ENABLED, UNCX_DISABLED_REASON
    if UNCX_LOOKUPS_ENABLED:
        UNCX_LOOKUPS_ENABLED = False
        UNCX_DISABLED_REASON = reason
        logger.warning("Disabling Uncx lookups: %s", reason)
    else:
        UNCX_DISABLED_REASON = reason

    schedule_uncx_recovery()


def enable_uncx_lookups(reason: str = "") -> None:
    """Manually re-enable Uncx lookups."""

    global UNCX_LOOKUPS_ENABLED, UNCX_DISABLED_REASON
    if not UNCX_LOOKUPS_ENABLED:
        UNCX_LOOKUPS_ENABLED = True
        UNCX_DISABLED_REASON = None
        if reason:
            logger.info("Re-enabled Uncx lookups: %s", reason)


def _coerce_uncx_timestamp(value: Any) -> Optional[float]:
    """Convert a variety of timestamp representations to seconds."""

    if value is None:
        return None

    if isinstance(value, (int, float)):
        return float(value)

    if isinstance(value, str):
        trimmed = value.strip()
        if not trimmed:
            return None
        try:
            return float(Decimal(trimmed))
        except (InvalidOperation, ValueError):
            try:
                return datetime.fromisoformat(trimmed.replace("Z", "+00:00")).timestamp()
            except ValueError:
                return None

    try:
        return float(Decimal(str(value)))
    except (InvalidOperation, ValueError, TypeError):
        return None


def _extract_uncx_unlock_timestamp(lock: dict) -> Optional[float]:
    """Return the unlock timestamp (in seconds) for an Uncx lock entry."""

    timestamp_keys = (
        "unlockDate",
        "unlock_date",
        "unlockTimestamp",
        "unlockTs",
        "unlock_time",
        "unlockTime",
        "unlockAt",
        "unlock_at",
        "unlockTimeStamp",
        "unlock",  # sometimes used as shorthand
    )

    for key in timestamp_keys:
        if key in lock:
            ts = _coerce_uncx_timestamp(lock.get(key))
            if ts is not None:
                return ts

    # Some responses may include ISO8601 strings under alternative keys
    for key in ("unlockDateISO", "unlock_date_iso", "unlockTimeISO"):
        if key in lock:
            ts = _coerce_uncx_timestamp(lock.get(key))
            if ts is not None:
                return ts

    return None


def _extract_uncx_lock_timestamp(lock: dict) -> Optional[float]:
    timestamp_keys = (
        "lockDate",
        "lock_date",
        "lockTimestamp",
        "lockTs",
        "lock_time",
        "lockTime",
        "lockAt",
        "lock_at",
        "lockTimeStamp",
    )

    for key in timestamp_keys:
        if key in lock:
            ts = _coerce_uncx_timestamp(lock.get(key))
            if ts is not None:
                return ts

    return None


def _extract_uncx_percent(lock: dict) -> Optional[float]:
    percent_keys = (
        "lockedPercent",
        "percent",
        "percentLocked",
        "lockPercent",
        "lpPercent",
        "share",
    )

    for key in percent_keys:
        if key not in lock:
            continue
        pct = _normalize_percent_ratio(lock.get(key))
        if pct is not None:
            return pct

    return None


def _extract_uncx_amount(lock: dict) -> Optional[Decimal]:
    """Return the locked token amount from an Uncx lock entry."""

    amount_keys = (
        "amount",
        "amountLocked",
        "locked_amount",
        "tokensLocked",
        "tokenAmount",
        "lockAmount",
        "amount_total",
        "amount_locked",
        "amountToken",
        "lpTokensLocked",
    )

    for key in amount_keys:
        if key not in lock:
            continue
        raw_value = lock.get(key)
        if raw_value in (None, ""):
            continue
        try:
            amount = Decimal(str(raw_value))
        except (InvalidOperation, ValueError, TypeError):
            continue
        if amount <= 0:
            continue

        # Normalize values that include a decimals hint when supplied as integers
        decimals_hint = lock.get("decimals") or lock.get("tokenDecimals")
        if (
            isinstance(decimals_hint, (int, float))
            or (isinstance(decimals_hint, str) and decimals_hint.isdigit())
        ):
            try:
                decimals_int = int(decimals_hint)
            except (TypeError, ValueError):
                decimals_int = None
            else:
                if decimals_int and amount == amount.to_integral_value():
                    try:
                        scaled = amount / (Decimal(10) ** decimals_int)
                    except (InvalidOperation, OverflowError):
                        scaled = None
                    else:
                        if scaled is not None and scaled > 0:
                            amount = scaled

        return amount

    formatted = lock.get("amountFormatted") or lock.get("amount_formatted")
    if formatted:
        try:
            amount = Decimal(str(formatted))
        except (InvalidOperation, ValueError, TypeError):
            return None
        return amount if amount > 0 else None

    return None


def _extract_uncx_locks(payload: Any) -> Optional[List[dict]]:
    """Extract the list of lock entries from a Uncx API response."""

    if payload is None:
        return []

    if isinstance(payload, list):
        return payload

    if isinstance(payload, dict):
        for key in ("locks", "result", "data", "items", "rows"):
            if key not in payload:
                continue
            nested = payload.get(key)
            if isinstance(nested, list):
                return nested
            if isinstance(nested, dict):
                extracted = _extract_uncx_locks(nested)
                if extracted is not None:
                    return extracted
        return []

    return None


async def _check_liquidity_locked_uncx_rest_async(
    pair_addr: str,
) -> Tuple[Optional[bool], Optional[LiquidityLockDetails]]:
    """Query the legacy Uncx REST API for liquidity locks."""

    if not UNCX_LOOKUPS_ENABLED:
        return None, None

    endpoint = UNCX_LOCKS_ENDPOINT.rstrip("/")
    url = f"{endpoint}/{pair_addr}"

    try:
        async with create_aiohttp_session() as session:
            async with session.get(url, timeout=FETCH_TIMEOUT) as resp:
                if resp.status == 404:
                    metrics.record_api_call(error=False)
                    logger.debug("Uncx reports no locks for %s", pair_addr)
                    return False, None
                if resp.status == 429:
                    metrics.record_api_call(error=True)
                    disable_uncx_lookups("Uncx rate limited")
                    return None, None
                if resp.status >= 500:
                    metrics.record_api_call(error=True)
                    disable_uncx_lookups(f"Uncx service error: {resp.status}")
                    return None, None

                resp.raise_for_status()
                data = await resp.json()
        metrics.record_api_call(error=False)
    except (aiohttp.ClientError, asyncio.TimeoutError, OSError) as exc:
        metrics.record_api_call(error=True)
        disable_uncx_lookups(f"Uncx lookup failed: {exc}")
        return None, None
    except Exception as exc:  # pragma: no cover - unexpected parser issues
        metrics.record_api_call(error=True)
        logger.debug("Uncx lock check error: %s", exc, exc_info=True)
        return None, None

    locks = _extract_uncx_locks(data)
    if locks is None:
        return None, None

    if not locks:
        return False, None

    now_ts = time.time()
    active_amount = Decimal(0)
    total_amount = Decimal(0)
    inconclusive = False
    entries = 0
    lock_times: List[int] = []
    unlock_times: List[int] = []
    coverage_values: List[float] = []

    for lock in locks:
        if not isinstance(lock, dict):
            inconclusive = True
            continue

        entries += 1

        amount = _extract_uncx_amount(lock)
        if amount is None:
            inconclusive = True
            continue

        status_text = str(lock.get("status") or "").lower()
        unlocked_flag = lock.get("isUnlocked")
        is_locked_flag = lock.get("isLocked")

        if unlocked_flag is True or "withdrawn" in status_text:
            continue

        total_amount += amount

        if is_locked_flag is True or ("lock" in status_text and "unlock" not in status_text):
            active_amount += amount
            continue

        unlock_ts = _extract_uncx_unlock_timestamp(lock)
        if unlock_ts is None:
            inconclusive = True
            continue

        normalized_unlock = _normalize_timestamp_seconds(unlock_ts)
        if normalized_unlock:
            unlock_times.append(normalized_unlock)

        if normalized_unlock is None:
            active_amount += amount
            continue

        if normalized_unlock > now_ts:
            active_amount += amount

        lock_ts = _extract_uncx_lock_timestamp(lock)
        normalized_lock = _normalize_timestamp_seconds(lock_ts)
        if normalized_lock:
            lock_times.append(normalized_lock)

        pct = _extract_uncx_percent(lock)
        if pct is not None:
            coverage_values.append(pct)

    if active_amount > 0:
        logger.debug(
            "Uncx lock snapshot for %s => active_amount=%s total=%s entries=%s",
            pair_addr,
            str(active_amount),
            str(total_amount) if total_amount else "0",
            entries,
        )
        details = LiquidityLockDetails(
            source="uncx_rest",
            coverage_pct=max(coverage_values) if coverage_values else None,
            locked_at=min(lock_times) if lock_times else None,
            unlock_at=max(unlock_times) if unlock_times else None,
        )
        return True, details

    if inconclusive:
        return None, None

    return False, None


def _format_uncx_graph_endpoint() -> Optional[str]:
    api_key = (UNCX_GRAPH_API_KEY or "").strip()
    subgraph_id = (UNCX_GRAPH_SUBGRAPH_ID or "").strip()
    template = (UNCX_GRAPH_ENDPOINT_TEMPLATE or "").strip()

    if not api_key or not subgraph_id or "{api_key}" not in template or "{subgraph_id}" not in template:
        return None

    try:
        return template.format(api_key=api_key, subgraph_id=subgraph_id)
    except Exception:  # pragma: no cover - defensive format issues
        return None


def _decimal_from_graph(value: Any) -> Optional[Decimal]:
    if value is None:
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError, TypeError):
        return None


def _int_from_graph(value: Any) -> Optional[int]:
    if value in (None, ""):
        return None
    try:
        return int(str(value), 0)
    except (ValueError, TypeError):
        return None


async def _check_liquidity_locked_uncx_graph_async(
    pair_addr: str,
) -> Tuple[Optional[bool], Optional[LiquidityLockDetails]]:
    endpoint = _format_uncx_graph_endpoint()
    if not endpoint:
        return None, None

    payload = {
        "query": """
            query ($poolId: ID!, $limit: Int!) {
              pool(id: $poolId) {
                id
                lockedPools {
                  id
                  lockedLiquidity
                  lockedPercent
                  lockedCoreUSD
                  lockedAmount0
                  lockedAmount1
                  numberOfLocks
                }
              }
              locks(where: { pool: $poolId }, first: $limit) {
                id
                lockedLiquidity
                lockedPercent
                lockedCoreUSD
                lockedAmount0
                lockedAmount1
                lockDate
                unlockDate
              }
            }
        """,
        "variables": {"poolId": pair_addr.lower(), "limit": 200},
    }

    try:
        async with create_aiohttp_session() as session:
            async with session.post(endpoint, json=payload, timeout=FETCH_TIMEOUT) as resp:
                resp.raise_for_status()
                result = await resp.json()
        metrics.record_api_call(error=False)
    except (aiohttp.ClientError, asyncio.TimeoutError, OSError) as exc:
        metrics.record_api_call(error=True)
        logger.debug("Uncx graph lookup failed for %s: %s", pair_addr, exc)
        return None, None
    except Exception as exc:  # pragma: no cover - unexpected parser issues
        metrics.record_api_call(error=True)
        logger.debug("Uncx graph processing error for %s: %s", pair_addr, exc, exc_info=True)
        return None, None

    if not isinstance(result, dict):
        return None, None

    if result.get("errors"):
        logger.debug("Uncx graph returned errors for %s: %s", pair_addr, result["errors"])
        return None

    data = result.get("data") or {}
    pool_info = data.get("pool") or {}
    lock_entries = data.get("locks") or []

    now_ts = int(time.time())
    active_liquidity = Decimal(0)
    total_liquidity = Decimal(0)
    parsed_any = False
    lock_times: List[int] = []
    unlock_times: List[int] = []
    coverage_values: List[float] = []

    for entry in lock_entries:
        if not isinstance(entry, dict):
            continue
        locked_liq = _decimal_from_graph(entry.get("lockedLiquidity"))
        if locked_liq is None:
            continue
        parsed_any = True
        if locked_liq <= 0:
            continue

        total_liquidity += locked_liq
        unlock_ts = _int_from_graph(entry.get("unlockDate"))
        normalized_unlock = _normalize_timestamp_seconds(unlock_ts)
        if normalized_unlock:
            unlock_times.append(normalized_unlock)
        if normalized_unlock in (None, 0) or normalized_unlock > now_ts:
            active_liquidity += locked_liq

        lock_ts = _int_from_graph(entry.get("lockDate"))
        normalized_lock = _normalize_timestamp_seconds(lock_ts)
        if normalized_lock:
            lock_times.append(normalized_lock)

        pct = _normalize_percent_ratio(entry.get("lockedPercent"))
        if pct is not None:
            coverage_values.append(pct)

    if active_liquidity > 0:
        logger.debug(
            "Uncx graph snapshot for %s => active_liquidity=%s total_liquidity=%s entries=%s",
            pair_addr,
            str(active_liquidity),
            str(total_liquidity) if total_liquidity else "0",
            len(lock_entries),
        )
        details = LiquidityLockDetails(
            source="uncx_graph",
            coverage_pct=max(coverage_values) if coverage_values else None,
            locked_at=min(lock_times) if lock_times else None,
            unlock_at=max(unlock_times) if unlock_times else None,
        )
        return True, details

    if parsed_any:
        return False, None

    locked_pools = pool_info.get("lockedPools") or []
    for locked_pool in locked_pools:
        if not isinstance(locked_pool, dict):
            continue
        locked_liq = _decimal_from_graph(locked_pool.get("lockedLiquidity"))
        locked_pct = _normalize_percent_ratio(locked_pool.get("lockedPercent"))
        if locked_liq is not None and locked_liq > 0:
            details = LiquidityLockDetails(
                source="uncx_graph",
                coverage_pct=locked_pct,
            )
            return True, details
        if locked_pct is not None and locked_pct > 0:
            details = LiquidityLockDetails(
                source="uncx_graph",
                coverage_pct=locked_pct,
            )
            return True, details

    return (False, None) if pool_info else (None, None)


async def _check_liquidity_locked_uncx_async(
    pair_addr: str,
) -> Tuple[Optional[bool], Optional[LiquidityLockDetails]]:
    """Return True/False when Uncx provides a definitive liquidity status."""

    rest_result: Optional[bool] = None
    rest_details: Optional[LiquidityLockDetails] = None
    if UNCX_LOOKUPS_ENABLED:
        rest_result, rest_details = await _check_liquidity_locked_uncx_rest_async(pair_addr)
        if rest_result is True:
            return True, rest_details

    graph_result, graph_details = await _check_liquidity_locked_uncx_graph_async(pair_addr)
    if graph_result is not None:
        return graph_result, graph_details

    return rest_result, rest_details


def ensure_etherscan_connectivity() -> None:
    """Validate Etherscan API keys and select a reachable endpoint."""

    global ETHERSCAN_API_URL

    if not ETHERSCAN_LOOKUPS_ENABLED:
        return

    if not ETHERSCAN_API_URL_CANDIDATES:
        log_event(
            logging.ERROR,
            "etherscan_endpoint",
            "No Etherscan API URLs configured",
        )
        disable_etherscan_lookups("No Etherscan API URLs configured")
        return

    if not any(key.strip() for key in ETHERSCAN_API_KEY_LIST):
        log_event(
            logging.ERROR,
            "etherscan_endpoint",
            "No Etherscan API keys configured",
        )
        disable_etherscan_lookups("No Etherscan API keys configured")
        return

    async def _select_endpoint():
        attempts: List[Tuple[str, str]] = []
        timeout = aiohttp.ClientTimeout(total=10)
        async with create_aiohttp_session(timeout=timeout) as session:
            for url in ETHERSCAN_API_URL_CANDIDATES:
                start = time.perf_counter()
                try:
                    params = _prepare_etherscan_params(
                        {
                            "module": "proxy",
                            "action": "eth_blockNumber",
                            "apikey": get_next_etherscan_key(),
                        },
                        url,
                    )
                    async with session.get(url, params=params) as resp:
                        resp.raise_for_status()
                        await resp.text()
                        latency_ms = round((time.perf_counter() - start) * 1000, 2)
                        log_event(
                            logging.INFO,
                            "etherscan_endpoint",
                            "Verified Etherscan endpoint",
                            context={"url": url},
                            latency_ms=latency_ms,
                        )
                        return url, attempts
                except Exception as exc:  # pragma: no cover - network errors
                    latency_ms = round((time.perf_counter() - start) * 1000, 2)
                    attempts.append((url, str(exc)))
                    log_event(
                        logging.WARNING,
                        "etherscan_endpoint",
                        "Failed to reach Etherscan endpoint",
                        error=str(exc),
                        context={"url": url},
                        latency_ms=latency_ms,
                    )
        return None, attempts

    try:
        selected_url, failures = asyncio.run(_select_endpoint())
    except Exception as exc:  # pragma: no cover - asyncio misconfiguration
        log_event(
            logging.ERROR,
            "etherscan_endpoint",
            "Etherscan endpoint verification failed",
            error=str(exc),
        )
        disable_etherscan_lookups(f"Etherscan verification failed: {exc}")
        return

    if selected_url:
        previous_primary = ETHERSCAN_API_URL_CANDIDATES[0]
        ETHERSCAN_API_URL = selected_url
        if selected_url != previous_primary:
            log_event(
                logging.INFO,
                "etherscan_endpoint",
                "Using fallback Etherscan endpoint",
                context={"url": selected_url},
            )
        return

    for url, error in failures:
        log_event(
            logging.ERROR,
            "etherscan_endpoint",
            "Etherscan endpoint unreachable",
            error=error,
            context={"url": url},
        )
    disable_etherscan_lookups("All configured Etherscan endpoints unreachable")


# Initialize wallet tracker now that helper functions are defined
wallet_tracker = get_shared_tracker(w3_read, get_next_etherscan_key)
start_market_mode_monitor()


class ContractVerificationStatus:
    UNVERIFIED = "unverified"
    VERIFIED = "verified"
    ERROR = "error"


async def _fetch_contract_source_etherscan_async(token_addr: str) -> dict:
    """Return verified source code information from Etherscan with retries."""

    if not ETHERSCAN_LOOKUPS_ENABLED:
        return {
            "status": ContractVerificationStatus.ERROR,
            "source": [],
            "compilerVersion": "",
            "contractName": "",
            "error": ETHERSCAN_DISABLED_REASON or "Etherscan lookups disabled",
        }

    token_addr = token_addr.lower()
    base_params = {
        "module": "contract",
        "action": "getsourcecode",
        "address": token_addr,
    }
    urls = [url for url in ETHERSCAN_API_URL_CANDIDATES if url]
    if ETHERSCAN_API_URL and ETHERSCAN_API_URL not in urls:
        urls.insert(0, ETHERSCAN_API_URL)
    if not urls:
        return {
            "status": ContractVerificationStatus.ERROR,
            "source": [],
            "compilerVersion": "",
            "contractName": "",
            "error": "No Etherscan API URLs configured",
        }

    errors: List[str] = []
    max_attempts = 3

    for base_url in urls:
        for attempt in range(max_attempts):
            attempt_params = dict(base_params)
            attempt_params["apikey"] = get_next_etherscan_key()
            attempt_params = _prepare_etherscan_params(attempt_params, base_url)

            try:
                async with create_aiohttp_session() as session:
                    async with session.get(base_url, params=attempt_params, timeout=20) as response:
                        if response.status >= 500:
                            body = (await response.text())[:120].strip()
                            errors.append(
                                f"{base_url} {response.status}: {body or 'server error'}"
                            )
                            await asyncio.sleep(min(2 ** attempt, 5))
                            continue

                        try:
                            payload = await response.json(content_type=None)
                        except (aiohttp.ContentTypeError, json.JSONDecodeError, ValueError) as exc:
                            body = (await response.text())[:120].strip()
                            errors.append(
                                f"{base_url} invalid JSON: {exc} :: {body or 'no body'}"
                            )
                            await asyncio.sleep(min(2 ** attempt, 5))
                            continue
            except (aiohttp.ClientError, asyncio.TimeoutError, OSError) as exc:
                errors.append(f"{base_url} attempt {attempt + 1}: {exc}")
                await asyncio.sleep(min(2 ** attempt, 5))
                continue
            except Exception as exc:  # pragma: no cover - defensive guard
                logger.warning("fetch_contract_source_etherscan unexpected error: %s", exc)
                errors.append(f"{base_url} unexpected error: {exc}")
                await asyncio.sleep(min(2 ** attempt, 5))
                continue

            status = payload.get("status")
            result = payload.get("result", [])
            message = (payload.get("message") or "").lower()

            if status == "1" and result:
                source_code = result[0].get("SourceCode", "")
                if not source_code:
                    return {
                        "status": ContractVerificationStatus.UNVERIFIED,
                        "source": [],
                        "compilerVersion": "",
                        "contractName": "",
                    }

                compiler = result[0].get("CompilerVersion", "")
                contract_name = result[0].get("ContractName", "")
                sources_list: List[dict] = []
                stripped = source_code.strip()
                if stripped.startswith("{"):
                    try:
                        data = json.loads(source_code)
                        for fname, info in data.get("sources", {}).items():
                            content = info.get("content", "")
                            sources_list.append({"filename": fname, "content": content})
                    except Exception:
                        sources_list.append(
                            {
                                "filename": contract_name or "contract.sol",
                                "content": source_code,
                            }
                        )
                else:
                    sources_list.append(
                        {"filename": contract_name or "contract.sol", "content": source_code}
                    )

                return {
                    "status": ContractVerificationStatus.VERIFIED,
                    "source": sources_list,
                    "compilerVersion": compiler,
                    "contractName": contract_name,
                }

            if message == "contract source code not verified":
                return {
                    "status": ContractVerificationStatus.UNVERIFIED,
                    "source": [],
                    "compilerVersion": "",
                    "contractName": "",
                }

            errors.append(
                f"{base_url} unexpected payload: status={status} message={payload.get('message')}"
            )
            await asyncio.sleep(min(2 ** attempt, 5))

    if errors:
        logger.warning(
            "Etherscan source fetch failed for %s: %s", token_addr, errors[-1]
        )

    return {
        "status": ContractVerificationStatus.ERROR,
        "source": [],
        "compilerVersion": "",
        "contractName": "",
        "error": "; ".join(errors) if errors else "unknown error",
    }

def fetch_contract_source_etherscan(token_addr: str) -> dict:
    return asyncio.run(_fetch_contract_source_etherscan_async(token_addr))


def get_contract_creator(token_addr: str) -> Optional[str]:
    """Return the deployer address via Etherscan as a fallback."""
    if not ETHERSCAN_LOOKUPS_ENABLED:
        return None
    api_key = get_next_etherscan_key()
    if not api_key:
        return None
    params = {
        "module": "contract",
        "action": "getcontractcreation",
        "contractaddresses": token_addr,
        "apikey": api_key,
    }
    try:
        params = _prepare_etherscan_params(params)
        resp = requests.get(ETHERSCAN_API_URL, params=params, timeout=20)
        data = resp.json()
        if data.get("status") == "1" and data.get("result"):
            return data["result"][0].get("contractCreator")
    except requests.RequestException as exc:
        disable_etherscan_lookups(f"contract creator lookup failed: {exc}")
    except Exception:
        logger.debug("contract creator fetch failed", exc_info=True)
    return None


def get_owner_info(token_addr: str):
    """Return (owner_or_creator, eth_balance, token_balance)."""
    owner_addr = None
    for fn in ["owner", "getOwner", "ownerAddress", "admin", "administrator"]:
        abi = [
            {
                "constant": True,
                "inputs": [],
                "name": fn,
                "outputs": [{"name": "", "type": "address"}],
                "type": "function",
            }
        ]
        try:
            c = w3_read.eth.contract(to_checksum_address(token_addr), abi=abi)
            owner_addr = getattr(c.functions, fn)().call()
            break
        except Exception:
            continue

    if not owner_addr:
        owner_addr = get_contract_creator(token_addr)

    if owner_addr:
        try:
            bal_eth = w3_read.eth.get_balance(owner_addr)
            token_abi = [
                {
                    "constant": True,
                    "inputs": [{"name": "", "type": "address"}],
                    "name": "balanceOf",
                    "outputs": [{"name": "", "type": "uint256"}],
                    "type": "function",
                }
            ]
            token = w3_read.eth.contract(to_checksum_address(token_addr), abi=token_abi)
            bal_token = token.functions.balanceOf(owner_addr).call()
            return owner_addr, w3_read.from_wei(bal_eth, "ether"), bal_token
        except Exception:
            pass
        try:
            bal_eth = w3_read.eth.get_balance(owner_addr)
            return owner_addr, w3_read.from_wei(bal_eth, "ether"), None
        except Exception:
            return owner_addr, None, None
    return None, None, None


async def _check_owner_wallet_activity_async(token_addr: str, owner_addr: str) -> bool:
    """Return True if suspicious owner transactions detected recently."""
    if not owner_addr:
        return False
    if not ETHERSCAN_LOOKUPS_ENABLED:
        return False
    api_key = get_next_etherscan_key()
    if not api_key:
        return False
    params = {
        "module": "account",
        "action": "tokentx",
        "address": owner_addr,
        "contractaddress": token_addr,
        "page": 1,
        "offset": 10,
        "sort": "desc",
        "apikey": api_key,
    }
    try:
        j = await _etherscan_get_async(params, FETCH_TIMEOUT)
        if j.get("status") != "1":
            return False
        total_supply = None
        try:
            abi = [{"constant": True, "inputs": [], "name": "totalSupply", "outputs": [{"name": "", "type": "uint256"}], "type": "function"}]
            c = w3_read.eth.contract(to_checksum_address(token_addr), abi=abi)
            total_supply = c.functions.totalSupply().call()
        except Exception:
            pass
        for tx in j.get("result", []):
            val = int(tx.get("value", "0"))
            if total_supply and val > total_supply * 0.05:
                return True
        return False
    except (aiohttp.ClientError, asyncio.TimeoutError, OSError) as e:
        disable_etherscan_lookups(f"owner activity lookup failed: {e}")
        return False
    except Exception as e:
        logger.debug(f"owner activity check error: {e}")
        return False


def check_owner_wallet_activity(token_addr: str, owner_addr: str) -> bool:
    return asyncio.run(_check_owner_wallet_activity_async(token_addr, owner_addr))


def _coerce_goplus_flag(value: Optional[Union[str, int, bool]]) -> Optional[bool]:
    """Normalize GoPlus boolean-like values to real booleans."""

    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    if isinstance(value, str):
        cleaned = value.strip().lower()
        if cleaned in {"1", "true", "yes", "y"}:
            return True
        if cleaned in {"0", "false", "no", "n"}:
            return False
    return None


def _extract_goplus_security(entry: dict) -> dict:
    """Return a trimmed security payload with normalized flags."""

    security: dict = {"raw": entry}
    score = entry.get("total_score")
    try:
        security["total_score"] = int(score) if score is not None else None
    except (TypeError, ValueError):
        security["total_score"] = None
    whitelist_flag = entry.get("is_whitelist")
    if whitelist_flag is None:
        whitelist_flag = entry.get("is_whitelisted")
    security["is_whitelist"] = _coerce_goplus_flag(whitelist_flag)
    return security


async def _fetch_third_party_security_async(token_addr: str) -> Optional[dict]:
    url = (
        "https://api.gopluslabs.io/api/v1/token_security/1?contract_addresses="
        f"{token_addr}"
    )
    try:
        async with create_aiohttp_session() as session:
            async with session.get(url, timeout=FETCH_TIMEOUT) as resp:
                data = await resp.json()
        entry = data.get("result", {}).get(token_addr.lower())
        if entry:
            return _extract_goplus_security(entry)
    except Exception as e:
        logger.debug(f"third-party score error: {e}")
    return None


def get_third_party_security(token_addr: str) -> Optional[dict]:
    return asyncio.run(_fetch_third_party_security_async(token_addr))


def get_third_party_risk_score(token_addr: str) -> Optional[int]:
    security = get_third_party_security(token_addr)
    if security:
        return security.get("total_score")
    return None


def get_lp_total_supply(pair_addr: str) -> Optional[int]:
    """Return current totalSupply for the LP token."""
    try:
        c = w3_read.eth.contract(to_checksum_address(pair_addr), abi=PAIR_ABI)
        return c.functions.totalSupply().call()
    except Exception as e:
        logger.debug(f"lp supply fetch error: {e}")
        return None


def _node_contains_identifiers(node, keywords) -> bool:
    """Recursively search an AST node for identifiers containing keywords."""
    if node is None:
        return False
    if isinstance(node, dict):
        if node.get("type") == "Identifier":
            name = str(node.get("name", "")).lower()
            for kw in keywords:
                if kw in name:
                    return True
        for child in node.values():
            if _node_contains_identifiers(child, keywords):
                return True
    elif isinstance(node, list):
        for elem in node:
            if _node_contains_identifiers(elem, keywords):
                return True
    return False


def _iter_modifiers(node):
    if isinstance(node, dict):
        if node.get("type") == "ModifierDefinition":
            yield node
        for child in node.values():
            if isinstance(child, (dict, list)):
                yield from _iter_modifiers(child)
    elif isinstance(node, list):
        for elem in node:
            yield from _iter_modifiers(elem)


def analyze_solidity_source(source_text: str) -> dict:
    """
    Returns dictionary of risk flags & a total riskScore.
    Example:
      {
        "ownerFunctions": bool,
        "canSetFees": bool,
        "maxTaxPossible": "unbounded" or "some guess",
        "canBlacklist": bool,
        "canPauseTrading": bool,
        "upgradeableProxy": bool,
        "renounceOwnerImplemented": bool,
        "botWhitelist": bool,
        "score": int
      }
    """
    text_lower = source_text.lower()
    flags = {
        "ownerFunctions": False,
        "canSetFees": False,
        "maxTaxPossible": "unknown",
        "canBlacklist": False,
        "canPauseTrading": False,
        "upgradeableProxy": False,
        "renounceOwnerImplemented": False,
        "transferBlockingModifier": False,
        "botWhitelist": False,
        "canModifyLimits": False,
        "canMint": False,
        "ownerActivity": False,
        "walletDrainer": False,
        "delegatecall": False,
        "selfDestruct": False,
        "tokenomicsPatterns": False,
        "autoLiquidityAdd": False,
        "ownerPrivileges": False,
        "thirdPartyScore": None,
        "thirdPartyWhitelist": False,
        "vestingOrTimelock": False,
        "privateSaleFunctions": False,
    }

    # parse AST to detect transfer/sell usage inside modifiers
    try:
        with contextlib.redirect_stderr(io.StringIO()):
            ast = parser.parse(source_text)
        for mod in _iter_modifiers(ast):
            body = mod.get("body")
            if _node_contains_identifiers(body, {"transfer", "sell"}):
                flags["transferBlockingModifier"] = True
                break
    except Exception:
        pass

    # 1) "onlyOwner" or "Ownable"
    if "onlyowner" in text_lower or "ownable" in text_lower:
        flags["ownerFunctions"] = True

    # 2) set fee or tax
    set_fee_pattern = re.compile(r"function\s+set\w*fee", re.IGNORECASE)
    if set_fee_pattern.search(source_text):
        flags["canSetFees"] = True
        # Possibly guess a max tax
        flags["maxTaxPossible"] = "potentially 100%"

    # 3) modify trading limits or tax
    if re.search(r"function\s+(setmaxwallet|setmaxtx|setmaxtransaction|updatetax|settax)", source_text, re.IGNORECASE):
        flags["canModifyLimits"] = True

    # 4) minting ability
    if re.search(r"function\s+mint", source_text, re.IGNORECASE):
        flags["canMint"] = True

    if re.search(r"liquidityfee|marketingfee|reflection", text_lower):
        flags["tokenomicsPatterns"] = True
    if re.search(r"addliquidity|swapandliquify|autoliquidity", text_lower):
        flags["autoLiquidityAdd"] = True
    if re.search(r"setowner\s*\(|transferownership\s*\(", text_lower):
        flags["ownerPrivileges"] = True
    if re.search(r"vesting|timelock", text_lower):
        flags["vestingOrTimelock"] = True
    if re.search(r"claim|unlock|release", text_lower):
        flags["privateSaleFunctions"] = True

    # 5) blacklisting
    if "blacklist(" in text_lower or "addtoblacklist(" in text_lower:
        flags["canBlacklist"] = True

    # 6) can pause trading
    if "enabletrading(" in text_lower or "settradingenabled(" in text_lower:
        flags["canPauseTrading"] = True

    # 7) upgradeable proxy
    if "proxy" in text_lower or "upgradeable" in text_lower:
        flags["upgradeableProxy"] = True

    # 8) renounceOwner
    if "function renounceownership" in text_lower:
        flags["renounceOwnerImplemented"] = True

    # 9) bot or whitelist detection
    if re.search(r"\b(bot\w*|whitelist\w*)", text_lower):
        flags["botWhitelist"] = True

    # 10) wallet drainer patterns
    if _has_wallet_drainer_pattern(source_text):
        flags["walletDrainer"] = True
    if re.search(r"delegatecall\s*\(", source_text):
        flags["delegatecall"] = True
    if re.search(r"(selfdestruct|suicide)\s*\(", source_text, re.IGNORECASE):
        flags["selfDestruct"] = True

    # risk scoring
    risk_score = 0
    if flags["ownerFunctions"]:
        risk_score += 2
    if flags["canSetFees"]:
        risk_score += 2
    if flags["maxTaxPossible"] == "potentially 100%":
        risk_score += 3
    if flags["canBlacklist"]:
        risk_score += 3
    if flags["canPauseTrading"]:
        risk_score += 3
    if flags["upgradeableProxy"]:
        risk_score += 4
    if not flags["renounceOwnerImplemented"]:
        risk_score += 2
    if flags["transferBlockingModifier"]:
        risk_score += 4
    if flags["botWhitelist"]:
        risk_score += 3
    if flags["canModifyLimits"]:
        risk_score += 2
    if flags["canMint"]:
        risk_score += 3
    if flags["walletDrainer"]:
        risk_score += 5
    if flags["delegatecall"]:
        risk_score += 4
    if flags["selfDestruct"]:
        risk_score += 4
    if flags["tokenomicsPatterns"]:
        risk_score += 1
    if flags["autoLiquidityAdd"]:
        risk_score += 2
    if flags["ownerPrivileges"]:
        risk_score += 3
    if flags["privateSaleFunctions"]:
        risk_score += 1
    if flags["thirdPartyScore"] is not None:
        risk_score += max(0, (100 - flags["thirdPartyScore"]) // 20)
    if flags["thirdPartyWhitelist"]:
        risk_score += 4

    flags["score"] = risk_score
    return flags


def detect_proxy_implementation(addr: str) -> Optional[str]:
    """Return implementation address if contract is a proxy (EIP1967)."""
    try:
        raw = w3_read.eth.get_storage_at(to_checksum_address(addr), EIP1967_IMPL_SLOT)
        if int.from_bytes(raw, "big") != 0:
            return to_checksum_address("0x" + raw.hex()[-40:])
    except Exception:
        pass
    return None


async def _check_renounced_by_event_async(addr: str) -> bool:
    if not ETHERSCAN_LOOKUPS_ENABLED:
        return False
    topic0 = Web3.keccak(text="OwnershipTransferred(address,address)").hex()
    zero_topic = "0x" + "0" * 64
    params = {
        "module": "logs",
        "action": "getLogs",
        "fromBlock": "0",
        "toBlock": "latest",
        "address": addr,
        "topic0": topic0,
        "topic2": zero_topic,
        "apikey": get_next_etherscan_key(),
    }
    try:
        params = _prepare_etherscan_params(params)
        async with create_aiohttp_session() as session:
            async with session.get(ETHERSCAN_API_URL, params=params, timeout=20) as r:
                j = await r.json()
        if j.get("status") == "1" and j.get("result"):
            return True
    except (aiohttp.ClientError, asyncio.TimeoutError, OSError) as e:
        disable_etherscan_lookups(f"renounce lookup failed: {e}")
    except Exception:
        pass
    return False


def check_renounced_by_event(addr: str) -> bool:
    return asyncio.run(_check_renounced_by_event_async(addr))


def run_slither_analysis(source: object) -> dict:
    """Run Slither on the given source code and return issue count.

    ``source`` can be a string (single Solidity file) or a list of
    ``{"filename": str, "content": str}`` dictionaries representing a project.
    """
    import tempfile, subprocess, json as _json, shutil

    result = {"slitherIssues": None}
    if shutil.which("slither") is None:
        logger.warning("slither executable not found; skipping analysis")
        result["slitherIssues"] = "not_installed"
        return result
    try:
        with tempfile.TemporaryDirectory() as tmpd:
            if isinstance(source, str):
                src_path = os.path.join(tmpd, "contract.sol")
                with open(src_path, "w", encoding="utf-8") as f:
                    f.write(source)
                target = src_path
            else:
                for part in source:
                    fpath = os.path.join(tmpd, part.get("filename", "contract.sol"))
                    os.makedirs(os.path.dirname(fpath), exist_ok=True)
                    with open(fpath, "w", encoding="utf-8") as f:
                        f.write(part.get("content", ""))
                target = tmpd

            out_json = os.path.join(tmpd, "slither.json")
            cmd = ["slither", target, "--json", out_json, "--solc-disable-warnings"]
            proc = subprocess.run(
                cmd,
                check=True,
                timeout=60,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
            )
            data = _json.load(open(out_json, encoding="utf-8"))
            issues = data.get("results", {}).get("detectors", [])
            result["slitherIssues"] = len(issues)
    except subprocess.CalledProcessError as e:
        stdout = (
            e.stdout.decode("utf-8", errors="replace")
            if isinstance(e.stdout, bytes)
            else str(e.stdout)
        )
        stderr = (
            e.stderr.decode("utf-8", errors="replace")
            if isinstance(e.stderr, bytes)
            else str(e.stderr)
        )
        logger.warning(f"slither analysis failed: {stdout}\n{stderr}")
    except Exception as e:
        logger.warning(f"slither analysis error: {e}")
    return result


def advanced_contract_check(token_addr: str) -> dict:
    """Fetch source and analyze risk. Handles proxies and renounce checks."""
    impl = detect_proxy_implementation(token_addr)
    target = impl or token_addr
    info = fetch_contract_source_etherscan(target)
    owner_addr, owner_bal, owner_token_bal = get_owner_info(target)
    suspicious_activity = False
    if owner_addr:
        suspicious_activity = check_owner_wallet_activity(target, owner_addr)
    renounced = False
    if owner_addr and owner_addr.lower() == ZERO_ADDRESS.lower():
        renounced = True
    else:
        renounced = check_renounced_by_event(target)
    third_party_security = get_third_party_security(target)
    third_score = None
    third_party_whitelist = None
    if third_party_security:
        third_score = third_party_security.get("total_score")
        third_party_whitelist = third_party_security.get("is_whitelist")
    private_sale = detect_private_sale_indicators(target)
    onchain_metrics = fetch_onchain_metrics(target)

    if info["status"] == ContractVerificationStatus.VERIFIED:
        combined = "".join(part["content"] + "\n" for part in info["source"])
        flags = analyze_solidity_source(combined)
        slither_res = run_slither_analysis(info["source"])
        if slither_res.get("slitherIssues") is not None:
            flags["slitherIssues"] = slither_res["slitherIssues"]
        else:
            flags["slitherIssues"] = "error"
        if impl:
            flags["upgradeableProxy"] = True
        if renounced:
            flags["renounced"] = True
        else:
            flags["renounced"] = False
        flags["thirdPartyScore"] = third_score
        if third_party_whitelist:
            flags["thirdPartyWhitelist"] = True
        score = flags.get("score", 0)
        if third_score is not None:
            score += max(0, (100 - third_score) // 20)
        owner_flag = flags.get("ownerActivity") or suspicious_activity
        if owner_flag:
            flags["ownerActivity"] = True
            score += 3
        if isinstance(flags.get("slitherIssues"), int):
            score += min(flags["slitherIssues"], 5)
        if not renounced:
            score += 2
        if score >= 10:
            st = "HIGH_RISK"
        else:
            st = "OK"
        result = {
            "verified": True,
            "riskScore": score,
            "riskFlags": flags,
            "status": st,
            "owner": owner_addr,
            "ownerBalanceEth": owner_bal,
            "ownerTokenBalance": owner_token_bal,
            "implementation": impl,
            "renounced": renounced,
            "slitherIssues": flags.get("slitherIssues"),
            "ownerActivity": suspicious_activity,
            "privateSale": private_sale,
            "onChainMetrics": onchain_metrics,
            "thirdPartySecurity": third_party_security,
        }
        return result
    elif info["status"] == ContractVerificationStatus.UNVERIFIED:
        return {
            "verified": False,
            "riskScore": 9999,
            "riskFlags": {"ownerActivity": suspicious_activity},
            "status": "UNVERIFIED",
            "owner": owner_addr,
            "ownerBalanceEth": owner_bal,
            "ownerTokenBalance": owner_token_bal,
            "implementation": impl,
            "renounced": renounced,
            "slitherIssues": None,
            "thirdPartySecurity": third_party_security,
        }
    else:
        return {
            "verified": False,
            "riskScore": None,
            "riskFlags": {"ownerActivity": suspicious_activity},
            "status": "ERROR",
            "owner": owner_addr,
            "ownerBalanceEth": owner_bal,
            "ownerTokenBalance": owner_token_bal,
            "implementation": impl,
            "renounced": renounced,
            "slitherIssues": None,
            "thirdPartySecurity": third_party_security,
            "error": info.get("error"),
        }


###########################################################
# Additional Metrics & Bull Season Helpers
###########################################################

def _parse_holder_balance(raw_value: Union[str, int, float, None]) -> Optional[int]:
    if raw_value is None:
        return None
    if isinstance(raw_value, int):
        return raw_value
    if isinstance(raw_value, float):
        if not np.isfinite(raw_value):
            return None
        return int(raw_value)
    if isinstance(raw_value, str):
        value = raw_value.strip()
        if not value:
            return None
        try:
            if value.startswith("0x"):
                return int(value, 16)
            if "." in value:
                return int(float(value))
            return int(value)
        except ValueError:
            return None
    return None


def _parse_holder_share(raw_value: Union[str, int, float, None]) -> Optional[float]:
    if raw_value is None:
        return None
    try:
        if isinstance(raw_value, (int, float)):
            share = float(raw_value)
        elif isinstance(raw_value, str):
            cleaned = raw_value.replace("%", "").strip()
            if not cleaned:
                return None
            share = float(cleaned)
        else:
            return None
    except (TypeError, ValueError):
        return None

    if share > 1:
        share = share / 100.0
    if share < 0:
        return None
    return min(share, 1.0)


def _normalise_holder_entry(entry: dict) -> Optional[dict]:
    address = entry.get("address") or entry.get("TokenHolderAddress")
    if not address:
        return None
    try:
        address = to_checksum_address(address)
    except ValueError:
        # keep original if checksum conversion fails
        address = address

    balance = _parse_holder_balance(
        entry.get("balance")
        or entry.get("TokenHolderQuantity")
        or entry.get("TokenHolderBalance")
        or entry.get("rawBalance")
    )
    share = _parse_holder_share(
        entry.get("share")
        or entry.get("TokenHolderPercentage")
        or entry.get("TokenHolderShare")
    )
    result = {"address": address, "balance": balance}
    if share is not None:
        result["share"] = share
    return result


async def _fetch_ethplorer_top_holders(token_addr: str, limit: int = 10) -> List[dict]:
    base_url = ETHPLORER_BASE_URL.rstrip("/")
    url = f"{base_url}/getTopTokenHolders/{token_addr}"
    timeout = aiohttp.ClientTimeout(total=FETCH_TIMEOUT)
    max_attempts = max(len(ETHPLORER_API_KEYS), 1)

    for attempt in range(max_attempts):
        api_key = get_next_ethplorer_key()
        if not api_key:
            api_key = _DEFAULT_ETHPLORER_KEYS[0]
        params = {"apiKey": api_key, "limit": limit}
        try:
            async with create_aiohttp_session(timeout=timeout) as session:
                async with session.get(url, params=params) as resp:
                    resp.raise_for_status()
                    payload = await resp.json()
        except (aiohttp.ClientError, asyncio.TimeoutError, OSError) as exc:
            log_event(
                logging.WARNING,
                "ethplorer_api",
                "Failed to fetch holder distribution",
                error=str(exc),
                context={"token": token_addr, "attempt": attempt + 1, "api_key": "***"},
            )
            continue
        except Exception as exc:
            logger.debug(f"ethplorer holder fetch error: {exc}")
            continue

        holders = []
        for item in payload.get("holders", []):
            normalised = _normalise_holder_entry(item)
            if normalised:
                holders.append(normalised)
        if holders:
            log_event(
                logging.INFO,
                "ethplorer_api",
                "Fetched holder distribution via Ethplorer",
                context={"token": token_addr, "count": len(holders)},
            )
        return holders

    return []


async def _fetch_holder_distribution_async(token_addr: str, limit: int = 10) -> List[dict]:
    holders = await _fetch_ethplorer_top_holders(token_addr, limit)
    if holders:
        return holders

    if not ETHERSCAN_LOOKUPS_ENABLED or not USE_ETHERSCAN_TOKEN_HOLDERS:
        return holders

    api_key = get_next_etherscan_key()
    if not api_key:
        return holders
    params = {
        "module": "token",
        "action": "tokenholderlist",
        "contractaddress": token_addr,
        "page": 1,
        "offset": limit,
        "apikey": api_key,
    }
    try:
        data = await _etherscan_get_async(params, FETCH_TIMEOUT)
        if isinstance(data, dict):
            result = data.get("result", [])
            if isinstance(result, list):
                holders = []
                for entry in result:
                    normalised = _normalise_holder_entry(entry)
                    if normalised:
                        holders.append(normalised)
                return holders
        return holders
    except (aiohttp.ClientError, asyncio.TimeoutError, OSError) as e:
        disable_etherscan_lookups(f"holder distribution lookup failed: {e}")
        return holders
    except Exception as e:
        logger.debug(f"holder distribution error: {e}")
        return holders


def fetch_holder_distribution(token_addr: str, limit: int = 10) -> List[dict]:
    return asyncio.run(_fetch_holder_distribution_async(token_addr, limit))


async def _analyze_transfer_history_async(token_addr: str, limit: int = 100) -> dict:
    if not ETHERSCAN_LOOKUPS_ENABLED:
        return {
            "uniqueBuyers": 0,
            "uniqueSellers": 0,
            "smartMoneyCount": 0,
        }
    api_key = get_next_etherscan_key()
    metrics = {
        "uniqueBuyers": 0,
        "uniqueSellers": 0,
        "smartMoneyCount": 0,
    }
    if not api_key:
        return metrics
    params = {
        "module": "account",
        "action": "tokentx",
        "contractaddress": token_addr,
        "page": 1,
        "offset": limit,
        "sort": "asc",
        "apikey": api_key,
    }
    try:
        data = await _etherscan_get_async(params, FETCH_TIMEOUT)
        if data.get("status") != "1":
            return metrics
        buyers = set()
        sellers = set()
        for tx in data.get("result", []):
            frm = tx.get("from", "").lower()
            to = tx.get("to", "").lower()
            if frm in {ZERO_ADDRESS.lower(), token_addr.lower()}:
                buyers.add(to)
            elif to in {ZERO_ADDRESS.lower(), token_addr.lower()}:
                sellers.add(frm)
            if frm in SMART_MONEY_WALLETS or to in SMART_MONEY_WALLETS:
                metrics["smartMoneyCount"] += 1
        metrics["uniqueBuyers"] = len(buyers)
        metrics["uniqueSellers"] = len(sellers)
        return metrics
    except (aiohttp.ClientError, asyncio.TimeoutError, OSError) as e:
        disable_etherscan_lookups(f"transfer history lookup failed: {e}")
        return metrics
    except Exception as e:
        logger.debug(f"transfer history error: {e}")
        return metrics


def analyze_transfer_history(token_addr: str, limit: int = 100) -> dict:
    return asyncio.run(_analyze_transfer_history_async(token_addr, limit))


async def _detect_private_sale_async(token_addr: str) -> dict:
    if not ETHERSCAN_LOOKUPS_ENABLED:
        return {"hasPresale": False, "largeTransfers": []}
    api_key = get_next_etherscan_key()
    result = {"hasPresale": False, "largeTransfers": []}
    if not api_key:
        return result
    params = {
        "module": "account",
        "action": "tokentx",
        "contractaddress": token_addr,
        "page": 1,
        "offset": 20,
        "sort": "asc",
        "apikey": api_key,
    }
    try:
        data = await _etherscan_get_async(params, FETCH_TIMEOUT)
        if data.get("status") != "1":
            return result
        total_supply = None
        try:
            abi = [{"constant": True, "inputs": [], "name": "totalSupply", "outputs": [{"name": "", "type": "uint256"}], "type": "function"}]
            c = w3_read.eth.contract(to_checksum_address(token_addr), abi=abi)
            total_supply = c.functions.totalSupply().call()
        except Exception:
            pass
        for tx in data.get("result", []):
            val = int(tx.get("value", "0"))
            if total_supply and val > total_supply * 0.02:
                result["hasPresale"] = True
                result["largeTransfers"].append(tx.get("to"))
        return result
    except (aiohttp.ClientError, asyncio.TimeoutError, OSError) as e:
        disable_etherscan_lookups(f"private sale lookup failed: {e}")
        return result
    except Exception as e:
        logger.debug(f"private sale detect error: {e}")
        return result


def detect_private_sale_indicators(token_addr: str) -> dict:
    return asyncio.run(_detect_private_sale_async(token_addr))


def has_private_sale(token_addr: str) -> bool:
    info = detect_private_sale_indicators(token_addr)
    return info.get("hasPresale", False)


def fetch_onchain_metrics(token_addr: str) -> dict:
    holders = fetch_holder_distribution(token_addr)
    if not isinstance(holders, list):
        holders = []
    transfers = analyze_transfer_history(token_addr)
    holder_share = None
    total_supply = 0
    try:
        abi = [{
            "constant": True,
            "inputs": [],
            "name": "totalSupply",
            "outputs": [{"name": "", "type": "uint256"}],
            "type": "function",
        }]
        c = w3_read.eth.contract(to_checksum_address(token_addr), abi=abi)
        total_supply = c.functions.totalSupply().call()
    except Exception as e:
        logger.debug(f"totalSupply fetch error: {e}")

    share_values: List[float] = []
    for holder in holders:
        if not isinstance(holder, dict):
            continue
        share = holder.get("share")
        if share is None:
            share = _parse_holder_share(
                holder.get("TokenHolderPercentage")
                or holder.get("TokenHolderShare")
            )
        if share is not None:
            share_values.append(share)
            continue
        if total_supply > 0:
            bal_val = holder.get("balance")
            if bal_val is None:
                bal_val = (
                    holder.get("TokenHolderQuantity")
                    or holder.get("TokenHolderBalance")
                    or holder.get("rawBalance")
                )
            bal = _parse_holder_balance(bal_val)
            if bal is not None:
                share_values.append(bal / total_supply)
    if share_values:
        holder_share = max(min(s, 1.0) for s in share_values)
    ratio = None
    if transfers["uniqueSellers"] > 0:
        ratio = transfers["uniqueBuyers"] / transfers["uniqueSellers"]
    return {
        "holderConcentration": holder_share,
        "uniqueBuyerSellerRatio": ratio,
        "smartMoneyCount": transfers["smartMoneyCount"],
    }


def get_current_liquidity(token_addr: str, pair_addr: str) -> float:
    """Helper to fetch current liquidity for a pair.

    The previous implementation mistakenly queried DexScreener with the pair
    address for both parameters which always returned empty data. By accepting
    the token address separately we ensure the lookup matches the pair and
    provides meaningful liquidity figures."""
    ds = fetch_dexscreener_data(token_addr, pair_addr)
    if ds:
        return ds.get("liquidityUsd", 0)
    return 0


def get_wallet_report(token_addr: str) -> dict:
    """Return cached wallet report or generate a new one."""
    key = token_addr.lower()
    now = time.time()
    data = WALLET_REPORT_CACHE.get(key)
    if data and now - data.get("ts", 0) < WALLET_REPORT_TTL:
        return data["report"]
    try:
        fut = asyncio.run_coroutine_threadsafe(
            wallet_tracker.generate_wallet_report(token_addr), wallet_event_loop
        )
        report = fut.result()
        WALLET_REPORT_CACHE[key] = {"ts": now, "report": report}
        return report
    except Exception as e:
        logger.error(f"wallet report error for {token_addr}: {e}")
        return {
            "risk_assessment": {"overall_risk": 100, "red_flags": ["error"]},
            "marketing_analysis": {"activity_score": 0, "total_spend_eth": 0},
            "developer_analysis": {"holding_percentage": 0},
            "wallet_summary": {},
        }



# ---------------------------------------------------------
# Wallet monitoring helpers (start/stop/list + resolver)
# ---------------------------------------------------------

def _resolve_main_token_from_arg(arg: str) -> Optional[str]:
    """Accepts a token or pair address; returns main token address or None."""
    if not arg:
        return None
    a = arg.strip()
    a_low = a.lower()
    # Direct token address (heuristic)
    if a_low.startswith("0x") and len(a_low) == 42:
        # If it's a known pair, convert to token
        if a_low in known_pairs:
            t0, t1 = known_pairs[a_low]
            try:
                return get_non_weth_token(t0, t1)
            except Exception:
                return t1 if t0.lower() == WETH_ADDRESS.lower() else t0
        # Otherwise assume token
        return a
    # Try to match known pairs by prefix
    for p, (t0, t1) in list(known_pairs.items()):
        if p.lower().startswith(a_low) or p.lower().endswith(a_low):
            try:
                return get_non_weth_token(t0, t1)
            except Exception:
                return t1 if t0.lower() == WETH_ADDRESS.lower() else t0
    return None


def start_wallet_monitor(token_addr: str):
    """Start background wallet monitoring if under limit."""
    key = token_addr.lower()
    if key in wallet_monitor_tasks or len(wallet_monitor_tasks) >= MAX_WALLET_MONITORS:
        return
    stop_event = asyncio.Event()
    wallet_monitor_stops[key] = stop_event
    try:
        fut = asyncio.run_coroutine_threadsafe(
            wallet_tracker.monitor_wallet_realtime(token_addr, wallet_activity_callback, stop_event),
            wallet_event_loop,
        )
        wallet_monitor_tasks[key] = fut
    except Exception as e:
        logger.error(f"wallet monitor error for {token_addr}: {e}")


def stop_wallet_monitor(token_or_pair_addr: str) -> bool:
    """Stop a running monitor. Returns True if stopped."""
    token = _resolve_main_token_from_arg(token_or_pair_addr) or token_or_pair_addr
    key = token.lower()
    fut = wallet_monitor_tasks.pop(key, None)
    ev = wallet_monitor_stops.pop(key, None)
    if ev is not None:
        try:
            ev.set()
        except Exception:
            pass
    if fut is not None:
        try:
            fut.cancel()
        except Exception:
            pass
        return True
    return False


def stop_all_wallet_monitors() -> int:
    keys = list(wallet_monitor_tasks.keys())
    count = 0
    for k in keys:
        try:
            if stop_wallet_monitor(k):
                count += 1
        except Exception:
            continue
    return count


def list_wallet_monitors() -> List[str]:
    """Return list of tokens currently being monitored."""
    return list(wallet_monitor_tasks.keys())



###########################################################
# 9. FIRST SELL DETECTION HELPERS
###########################################################


def _address_in_source(token_addr: str, wallet: str) -> bool:
    info = fetch_contract_source_etherscan(token_addr)
    if info.get("status") != ContractVerificationStatus.VERIFIED:
        return False
    w = wallet.lower().replace("0x", "")
    for part in info.get("source", []):
        if w in part.get("content", "").lower():
            return True
    return False


def analyze_seller_wallet(token_addr: str, wallet: str) -> Tuple[str, int]:
    """Return descriptive flags and a simple risk score for a seller wallet."""
    flags = []
    risk = 0
    try:
        code = w3_read.eth.get_code(to_checksum_address(wallet))
        if code and len(code) > 0:
            flags.append("contract")
            risk += 5
    except Exception:
        pass
    try:
        if _address_in_source(token_addr, wallet):
            flags.append("listed_in_contract")
            risk += 2
    except Exception:
        pass
    try:
        txc = w3_read.eth.get_transaction_count(to_checksum_address(wallet))
        if txc < 5:
            flags.append("new_wallet")
            risk += 1
    except Exception:
        pass
    try:
        bal = w3_read.eth.get_balance(to_checksum_address(wallet))
        if bal < Web3.to_wei(0.05, "ether"):
            flags.append("low_balance")
            risk += 1
    except Exception:
        pass
    if not flags:
        flags.append("EOA")
    return ", ".join(flags), risk


def detect_first_sell(pair_addr: str, token0: str, token1: str, from_block: int) -> Optional[str]:
    try:
        logs = w3_read.eth.get_logs(
            {
                "address": pair_addr,
                "fromBlock": from_block,
                "toBlock": "latest",
                "topics": [SWAP_TOPIC_V2],
            }
        )
        for lg in logs:
            data_field = lg["data"]
            if isinstance(data_field, HexBytes):
                data_field = data_field.hex()
            if data_field.startswith("0x"):
                data_field = data_field[2:]
            a0_in, a1_in, a0_out, a1_out = decode(
                ["uint256", "uint256", "uint256", "uint256"],
                bytes.fromhex(data_field),
            )
            is_sell = False
            if token0.lower() == WETH_ADDRESS.lower():
                is_sell = a0_out > 0
            elif token1.lower() == WETH_ADDRESS.lower():
                is_sell = a1_out > 0
            if is_sell:
                try:
                    tx = w3_read.eth.get_transaction(lg["transactionHash"])
                    return tx["from"]
                except Exception:
                    sender = "0x" + lg["topics"][1].hex()[-40:]
                    return sender
    except Exception as e:
        logger.debug(f"first sell v2 error: {e}")

    try:
        logs = w3_read.eth.get_logs(
            {
                "address": pair_addr,
                "fromBlock": from_block,
                "toBlock": "latest",
                "topics": [SWAP_TOPIC_V3],
            }
        )
        for lg in logs:
            data_field = lg["data"]
            if isinstance(data_field, HexBytes):
                data_field = data_field.hex()
            if data_field.startswith("0x"):
                data_field = data_field[2:]
            a0, a1, _, _, _ = decode(
                ["int256", "int256", "uint160", "uint128", "int24"],
                bytes.fromhex(data_field),
            )
            is_sell = False
            if token0.lower() == WETH_ADDRESS.lower():
                is_sell = a0 < 0
            elif token1.lower() == WETH_ADDRESS.lower():
                is_sell = a1 < 0
            if is_sell:
                try:
                    tx = w3_read.eth.get_transaction(lg["transactionHash"])
                    return tx["from"]
                except Exception:
                    sender = "0x" + lg["topics"][1].hex()[-40:]
                    return sender
    except Exception as e:
        logger.debug(f"first sell v3 error: {e}")
    return None


###########################################################
# 9. MAIN PAIR CRITERIA (extended with advanced checks)
###########################################################

detected_at: Dict[str, float] = {}
# Track which pairs have already been processed to avoid duplicate work.
SEEN_PAIRS: Set[str] = set()


def should_retry_dexscreener(pair_addr: str, reason: str) -> Tuple[bool, Optional[float]]:
    """Decide whether to keep retrying DexScreener for the given failure reason."""

    if reason != "not_listed":
        return True, None

    first_seen = detected_at.get(pair_addr.lower())
    if first_seen is None:
        # Should not happen, but err on the side of retrying
        return True, None

    age = time.time() - first_seen
    return age <= DEXSCREENER_NOT_LISTED_REQUEUE_WINDOW, age


def check_pair_criteria(
    pair_addr: str, token0: str, token1: str
) -> Tuple[int, int, dict]:
    """Evaluate core safety criteria for a newly created pair.

    The legacy implementation returned ``None`` in many scenarios which in turn
    caused ``handle_new_pair`` to crash when unpacking the result.  The new
    implementation always returns the expected ``(passes, total, extra)`` tuple
    together with a detailed ``extra`` payload describing every datapoint used
    in the decision.
    """

    main_token = get_non_weth_token(token0, token1)
    counter_token = token1 if main_token.lower() == token0.lower() else token0

    expected_total_checks = 14

    dex_data, dex_reason = fetch_dexscreener_data(
        main_token, pair_addr, with_reason=True
    )

    extra: Dict[str, object] = {
        "pairAddress": pair_addr,
        "token0": token0,
        "token1": token1,
        "tokenAddress": main_token,
        "dexscreener_missing": False,
        "should_requeue": False,
        "dexscreener_reason": None,
        "transient_failure": False,
    }

    if not dex_data:
        reason = dex_reason or "unknown"
        should_requeue, age = should_retry_dexscreener(pair_addr, reason)
        extra.update(
            {
                "dexscreener_missing": True,
                "dexscreener_reason": reason,
                "should_requeue": should_requeue,
                "dexscreener_not_listed_age": age,
                "dexscreener_retry_window": DEXSCREENER_NOT_LISTED_REQUEUE_WINDOW,
                "dexscreener_retry_window_expired": (
                    age is not None and age > DEXSCREENER_NOT_LISTED_REQUEUE_WINDOW
                ),
                "transient_failure": reason
                in {"network_error", "rate_limited", "unexpected_error"},
            }
        )
        extra["checkBreakdown"] = {}
        return 0, expected_total_checks, extra

    extra.update(dex_data)
    token_name = dex_data.get("baseTokenName") or ""
    token_symbol = dex_data.get("baseTokenSymbol") or ""
    extra["tokenName"] = token_name or token_symbol or main_token
    if token_symbol:
        extra["tokenSymbol"] = token_symbol

    buys = int(dex_data.get("buys", 0) or 0)
    sells = int(dex_data.get("sells", 0) or 0)
    trades = buys + sells
    extra["trades24h"] = trades
    extra["clogPercent"] = (sells / trades) * 100 if trades else None

    try:
        contract_info = advanced_contract_check(main_token)
    except Exception as exc:  # pragma: no cover - defensive guard
        logger.error("contract analysis failed for %s: %s", main_token, exc)
        contract_info = {
            "verified": False,
            "riskScore": None,
            "status": "ERROR",
            "owner": None,
            "ownerBalanceEth": None,
            "ownerTokenBalance": None,
            "implementation": None,
            "renounced": None,
            "slitherIssues": None,
            "privateSale": {},
            "onChainMetrics": {},
            "error": str(exc),
        }

    extra.update(
        {
            "verified": contract_info.get("verified"),
            "contractCheckStatus": contract_info.get("status"),
            "riskScore": contract_info.get("riskScore"),
            "riskFlags": contract_info.get("riskFlags", {}),
            "owner": contract_info.get("owner"),
            "ownerBalanceEth": contract_info.get("ownerBalanceEth"),
            "ownerTokenBalance": contract_info.get("ownerTokenBalance"),
            "implementation": contract_info.get("implementation"),
            "contractRenounced": contract_info.get("renounced"),
            "slitherIssues": contract_info.get("slitherIssues"),
            "privateSale": contract_info.get("privateSale", {}),
            "onChainMetrics": contract_info.get("onChainMetrics", {}),
            "contractAnalysisError": contract_info.get("error"),
        }
    )

    honeypot_main = check_honeypot_is(main_token, pair_addr=pair_addr)
    honeypot_counter = False
    if counter_token.lower() != WETH_ADDRESS.lower():
        honeypot_counter = check_honeypot_is(counter_token, pair_addr=pair_addr)
    extra["honeypotMain"] = honeypot_main
    extra["honeypotCounter"] = honeypot_counter

    recent_liq_removal = check_recent_liquidity_removal(pair_addr)
    extra["recentLiquidityRemoval"] = recent_liq_removal

    checks: List[Tuple[str, bool]] = []

    def add_check(name: str, passed: bool) -> None:
        checks.append((name, bool(passed)))

    add_check("liquidity", float(dex_data.get("liquidityUsd", 0)) >= MIN_LIQUIDITY_USD)
    add_check("volume", float(dex_data.get("volume24h", 0)) >= MIN_VOLUME_USD)
    add_check("fdv", float(dex_data.get("fdv", 0)) >= MIN_FDV_USD)
    add_check("marketcap", float(dex_data.get("marketCap", 0)) >= MIN_MARKETCAP_USD)
    add_check("buys", buys >= MIN_BUYS_FIRST_HOUR)
    add_check("trades", trades >= MIN_TRADES_REQUIRED)
    add_check("locked_liquidity", bool(dex_data.get("lockedLiquidity")))
    add_check("no_recent_liq_removal", not recent_liq_removal)
    add_check("honeypot_main", not honeypot_main)
    add_check(
        "honeypot_counter",
        counter_token.lower() == WETH_ADDRESS.lower() or not honeypot_counter,
    )
    add_check("contract_verified", contract_info.get("verified") is True)
    risk_score = contract_info.get("riskScore")
    add_check(
        "risk_score",
        isinstance(risk_score, (int, float)) and risk_score < 10,
    )
    add_check("renounced", contract_info.get("renounced") is True)
    private_sale = contract_info.get("privateSale", {})
    add_check(
        "no_private_sale",
        not bool(private_sale.get("hasPresale")) if isinstance(private_sale, dict) else True,
    )

    passes = sum(1 for _, ok in checks if ok)
    total_checks = len(checks) or expected_total_checks
    extra["checkBreakdown"] = {name: ok for name, ok in checks}

    return passes, total_checks, extra

# Initialize wallet tracker now that helper functions are defined
wallet_tracker = get_shared_tracker(w3_read, get_next_etherscan_key)
start_market_mode_monitor()


class ContractVerificationStatus:
    UNVERIFIED = "unverified"
    VERIFIED = "verified"
    ERROR = "error"


async def _fetch_contract_source_etherscan_async(token_addr: str) -> dict:
    """Return verified source code information from Etherscan with retries."""

    if not ETHERSCAN_LOOKUPS_ENABLED:
        return {
            "status": ContractVerificationStatus.ERROR,
            "source": [],
            "compilerVersion": "",
            "contractName": "",
            "error": ETHERSCAN_DISABLED_REASON or "Etherscan lookups disabled",
        }

    token_addr = token_addr.lower()
    base_params = {
        "module": "contract",
        "action": "getsourcecode",
        "address": token_addr,
    }
    urls = [url for url in ETHERSCAN_API_URL_CANDIDATES if url]
    if ETHERSCAN_API_URL and ETHERSCAN_API_URL not in urls:
        urls.insert(0, ETHERSCAN_API_URL)
    if not urls:
        return {
            "status": ContractVerificationStatus.ERROR,
            "source": [],
            "compilerVersion": "",
            "contractName": "",
            "error": "No Etherscan API URLs configured",
        }

    errors: List[str] = []
    max_attempts = 3

    for base_url in urls:
        for attempt in range(max_attempts):
            attempt_params = dict(base_params)
            attempt_params["apikey"] = get_next_etherscan_key()
            attempt_params = _prepare_etherscan_params(attempt_params, base_url)

            try:
                async with create_aiohttp_session() as session:
                    async with session.get(base_url, params=attempt_params, timeout=20) as response:
                        if response.status >= 500:
                            body = (await response.text())[:120].strip()
                            errors.append(
                                f"{base_url} {response.status}: {body or 'server error'}"
                            )
                            await asyncio.sleep(min(2 ** attempt, 5))
                            continue

                        try:
                            payload = await response.json(content_type=None)
                        except (aiohttp.ContentTypeError, json.JSONDecodeError, ValueError) as exc:
                            body = (await response.text())[:120].strip()
                            errors.append(
                                f"{base_url} invalid JSON: {exc} :: {body or 'no body'}"
                            )
                            await asyncio.sleep(min(2 ** attempt, 5))
                            continue
            except (aiohttp.ClientError, asyncio.TimeoutError, OSError) as exc:
                errors.append(f"{base_url} attempt {attempt + 1}: {exc}")
                await asyncio.sleep(min(2 ** attempt, 5))
                continue
            except Exception as exc:  # pragma: no cover - defensive guard
                logger.warning("fetch_contract_source_etherscan unexpected error: %s", exc)
                errors.append(f"{base_url} unexpected error: {exc}")
                await asyncio.sleep(min(2 ** attempt, 5))
                continue

            status = payload.get("status")
            result = payload.get("result", [])
            message = (payload.get("message") or "").lower()

            if status == "1" and result:
                source_code = result[0].get("SourceCode", "")
                if not source_code:
                    return {
                        "status": ContractVerificationStatus.UNVERIFIED,
                        "source": [],
                        "compilerVersion": "",
                        "contractName": "",
                    }

                compiler = result[0].get("CompilerVersion", "")
                contract_name = result[0].get("ContractName", "")
                sources_list: List[dict] = []
                stripped = source_code.strip()
                if stripped.startswith("{"):
                    try:
                        data = json.loads(source_code)
                        for fname, info in data.get("sources", {}).items():
                            content = info.get("content", "")
                            sources_list.append({"filename": fname, "content": content})
                    except Exception:
                        sources_list.append(
                            {
                                "filename": contract_name or "contract.sol",
                                "content": source_code,
                            }
                        )
                else:
                    sources_list.append(
                        {"filename": contract_name or "contract.sol", "content": source_code}
                    )

                return {
                    "status": ContractVerificationStatus.VERIFIED,
                    "source": sources_list,
                    "compilerVersion": compiler,
                    "contractName": contract_name,
                }

            if message == "contract source code not verified":
                return {
                    "status": ContractVerificationStatus.UNVERIFIED,
                    "source": [],
                    "compilerVersion": "",
                    "contractName": "",
                }

            errors.append(
                f"{base_url} unexpected payload: status={status} message={payload.get('message')}"
            )
            await asyncio.sleep(min(2 ** attempt, 5))

    if errors:
        logger.warning(
            "Etherscan source fetch failed for %s: %s", token_addr, errors[-1]
        )

    return {
        "status": ContractVerificationStatus.ERROR,
        "source": [],
        "compilerVersion": "",
        "contractName": "",
        "error": "; ".join(errors) if errors else "unknown error",
    }

def fetch_contract_source_etherscan(token_addr: str) -> dict:
    return asyncio.run(_fetch_contract_source_etherscan_async(token_addr))


def get_contract_creator(token_addr: str) -> Optional[str]:
    """Return the deployer address via Etherscan as a fallback."""
    if not ETHERSCAN_LOOKUPS_ENABLED:
        return None
    api_key = get_next_etherscan_key()
    if not api_key:
        return None
    params = {
        "module": "contract",
        "action": "getcontractcreation",
        "contractaddresses": token_addr,
        "apikey": api_key,
    }
    try:
        params = _prepare_etherscan_params(params)
        resp = requests.get(ETHERSCAN_API_URL, params=params, timeout=20)
        data = resp.json()
        if data.get("status") == "1" and data.get("result"):
            return data["result"][0].get("contractCreator")
    except requests.RequestException as exc:
        disable_etherscan_lookups(f"contract creator lookup failed: {exc}")
    except Exception:
        logger.debug("contract creator fetch failed", exc_info=True)
    return None


def get_owner_info(token_addr: str):
    """Return (owner_or_creator, eth_balance, token_balance)."""
    owner_addr = None
    for fn in ["owner", "getOwner", "ownerAddress", "admin", "administrator"]:
        abi = [
            {
                "constant": True,
                "inputs": [],
                "name": fn,
                "outputs": [{"name": "", "type": "address"}],
                "type": "function",
            }
        ]
        try:
            c = w3_read.eth.contract(to_checksum_address(token_addr), abi=abi)
            owner_addr = getattr(c.functions, fn)().call()
            break
        except Exception:
            continue

    if not owner_addr:
        owner_addr = get_contract_creator(token_addr)

    if owner_addr:
        try:
            bal_eth = w3_read.eth.get_balance(owner_addr)
            token_abi = [
                {
                    "constant": True,
                    "inputs": [{"name": "", "type": "address"}],
                    "name": "balanceOf",
                    "outputs": [{"name": "", "type": "uint256"}],
                    "type": "function",
                }
            ]
            token = w3_read.eth.contract(to_checksum_address(token_addr), abi=token_abi)
            bal_token = token.functions.balanceOf(owner_addr).call()
            return owner_addr, w3_read.from_wei(bal_eth, "ether"), bal_token
        except Exception:
            pass
        try:
            bal_eth = w3_read.eth.get_balance(owner_addr)
            return owner_addr, w3_read.from_wei(bal_eth, "ether"), None
        except Exception:
            return owner_addr, None, None
    return None, None, None


async def _check_owner_wallet_activity_async(token_addr: str, owner_addr: str) -> bool:
    """Return True if suspicious owner transactions detected recently."""
    if not owner_addr:
        return False
    if not ETHERSCAN_LOOKUPS_ENABLED:
        return False
    api_key = get_next_etherscan_key()
    if not api_key:
        return False
    params = {
        "module": "account",
        "action": "tokentx",
        "address": owner_addr,
        "contractaddress": token_addr,
        "page": 1,
        "offset": 10,
        "sort": "desc",
        "apikey": api_key,
    }
    try:
        j = await _etherscan_get_async(params, FETCH_TIMEOUT)
        if j.get("status") != "1":
            return False
        total_supply = None
        try:
            abi = [{"constant": True, "inputs": [], "name": "totalSupply", "outputs": [{"name": "", "type": "uint256"}], "type": "function"}]
            c = w3_read.eth.contract(to_checksum_address(token_addr), abi=abi)
            total_supply = c.functions.totalSupply().call()
        except Exception:
            pass
        for tx in j.get("result", []):
            val = int(tx.get("value", "0"))
            if total_supply and val > total_supply * 0.05:
                return True
        return False
    except (aiohttp.ClientError, asyncio.TimeoutError, OSError) as e:
        disable_etherscan_lookups(f"owner activity lookup failed: {e}")
        return False
    except Exception as e:
        logger.debug(f"owner activity check error: {e}")
        return False


def check_owner_wallet_activity(token_addr: str, owner_addr: str) -> bool:
    return asyncio.run(_check_owner_wallet_activity_async(token_addr, owner_addr))


async def _fetch_third_party_security_async(token_addr: str) -> Optional[dict]:
    url = (
        "https://api.gopluslabs.io/api/v1/token_security/1?contract_addresses="
        f"{token_addr}"
    )
    try:
        async with create_aiohttp_session() as session:
            async with session.get(url, timeout=FETCH_TIMEOUT) as resp:
                data = await resp.json()
        entry = data.get("result", {}).get(token_addr.lower())
        if entry:
            return _extract_goplus_security(entry)
    except Exception as e:
        logger.debug(f"third-party score error: {e}")
    return None


def get_third_party_security(token_addr: str) -> Optional[dict]:
    return asyncio.run(_fetch_third_party_security_async(token_addr))


def get_third_party_risk_score(token_addr: str) -> Optional[int]:
    security = get_third_party_security(token_addr)
    if security:
        return security.get("total_score")
    return None


def get_lp_total_supply(pair_addr: str) -> Optional[int]:
    """Return current totalSupply for the LP token."""
    try:
        c = w3_read.eth.contract(to_checksum_address(pair_addr), abi=PAIR_ABI)
        return c.functions.totalSupply().call()
    except Exception as e:
        logger.debug(f"lp supply fetch error: {e}")
        return None


def _node_contains_identifiers(node, keywords) -> bool:
    """Recursively search an AST node for identifiers containing keywords."""
    if node is None:
        return False
    if isinstance(node, dict):
        if node.get("type") == "Identifier":
            name = str(node.get("name", "")).lower()
            for kw in keywords:
                if kw in name:
                    return True
        for child in node.values():
            if _node_contains_identifiers(child, keywords):
                return True
    elif isinstance(node, list):
        for elem in node:
            if _node_contains_identifiers(elem, keywords):
                return True
    return False


def _iter_modifiers(node):
    if isinstance(node, dict):
        if node.get("type") == "ModifierDefinition":
            yield node
        for child in node.values():
            if isinstance(child, (dict, list)):
                yield from _iter_modifiers(child)
    elif isinstance(node, list):
        for elem in node:
            yield from _iter_modifiers(elem)


def analyze_solidity_source(source_text: str) -> dict:
    """
    Returns dictionary of risk flags & a total riskScore.
    Example:
      {
        "ownerFunctions": bool,
        "canSetFees": bool,
        "maxTaxPossible": "unbounded" or "some guess",
        "canBlacklist": bool,
        "canPauseTrading": bool,
        "upgradeableProxy": bool,
        "renounceOwnerImplemented": bool,
        "botWhitelist": bool,
        "score": int
      }
    """
    text_lower = source_text.lower()
    flags = {
        "ownerFunctions": False,
        "canSetFees": False,
        "maxTaxPossible": "unknown",
        "canBlacklist": False,
        "canPauseTrading": False,
        "upgradeableProxy": False,
        "renounceOwnerImplemented": False,
        "transferBlockingModifier": False,
        "botWhitelist": False,
        "canModifyLimits": False,
        "canMint": False,
        "ownerActivity": False,
        "walletDrainer": False,
        "delegatecall": False,
        "selfDestruct": False,
        "tokenomicsPatterns": False,
        "autoLiquidityAdd": False,
        "ownerPrivileges": False,
        "thirdPartyScore": None,
        "thirdPartyWhitelist": False,
        "vestingOrTimelock": False,
        "privateSaleFunctions": False,
    }

    # parse AST to detect transfer/sell usage inside modifiers
    try:
        with contextlib.redirect_stderr(io.StringIO()):
            ast = parser.parse(source_text)
        for mod in _iter_modifiers(ast):
            body = mod.get("body")
            if _node_contains_identifiers(body, {"transfer", "sell"}):
                flags["transferBlockingModifier"] = True
                break
    except Exception:
        pass

    # 1) "onlyOwner" or "Ownable"
    if "onlyowner" in text_lower or "ownable" in text_lower:
        flags["ownerFunctions"] = True

    # 2) set fee or tax
    set_fee_pattern = re.compile(r"function\s+set\w*fee", re.IGNORECASE)
    if set_fee_pattern.search(source_text):
        flags["canSetFees"] = True
        # Possibly guess a max tax
        flags["maxTaxPossible"] = "potentially 100%"

    # 3) modify trading limits or tax
    if re.search(r"function\s+(setmaxwallet|setmaxtx|setmaxtransaction|updatetax|settax)", source_text, re.IGNORECASE):
        flags["canModifyLimits"] = True

    # 4) minting ability
    if re.search(r"function\s+mint", source_text, re.IGNORECASE):
        flags["canMint"] = True

    if re.search(r"liquidityfee|marketingfee|reflection", text_lower):
        flags["tokenomicsPatterns"] = True
    if re.search(r"addliquidity|swapandliquify|autoliquidity", text_lower):
        flags["autoLiquidityAdd"] = True
    if re.search(r"setowner\s*\(|transferownership\s*\(", text_lower):
        flags["ownerPrivileges"] = True
    if re.search(r"vesting|timelock", text_lower):
        flags["vestingOrTimelock"] = True
    if re.search(r"claim|unlock|release", text_lower):
        flags["privateSaleFunctions"] = True

    # 5) blacklisting
    if "blacklist(" in text_lower or "addtoblacklist(" in text_lower:
        flags["canBlacklist"] = True

    # 6) can pause trading
    if "enabletrading(" in text_lower or "settradingenabled(" in text_lower:
        flags["canPauseTrading"] = True

    # 7) upgradeable proxy
    if "proxy" in text_lower or "upgradeable" in text_lower:
        flags["upgradeableProxy"] = True

    # 8) renounceOwner
    if "function renounceownership" in text_lower:
        flags["renounceOwnerImplemented"] = True

    # 9) bot or whitelist detection
    if re.search(r"\b(bot\w*|whitelist\w*)", text_lower):
        flags["botWhitelist"] = True

    # 10) wallet drainer patterns
    if _has_wallet_drainer_pattern(source_text):
        flags["walletDrainer"] = True
    if re.search(r"delegatecall\s*\(", source_text):
        flags["delegatecall"] = True
    if re.search(r"(selfdestruct|suicide)\s*\(", source_text, re.IGNORECASE):
        flags["selfDestruct"] = True

    # risk scoring
    risk_score = 0
    if flags["ownerFunctions"]:
        risk_score += 2
    if flags["canSetFees"]:
        risk_score += 2
    if flags["maxTaxPossible"] == "potentially 100%":
        risk_score += 3
    if flags["canBlacklist"]:
        risk_score += 3
    if flags["canPauseTrading"]:
        risk_score += 3
    if flags["upgradeableProxy"]:
        risk_score += 4
    if not flags["renounceOwnerImplemented"]:
        risk_score += 2
    if flags["transferBlockingModifier"]:
        risk_score += 4
    if flags["botWhitelist"]:
        risk_score += 3
    if flags["canModifyLimits"]:
        risk_score += 2
    if flags["canMint"]:
        risk_score += 3
    if flags["walletDrainer"]:
        risk_score += 5
    if flags["delegatecall"]:
        risk_score += 4
    if flags["selfDestruct"]:
        risk_score += 4
    if flags["tokenomicsPatterns"]:
        risk_score += 1
    if flags["autoLiquidityAdd"]:
        risk_score += 2
    if flags["ownerPrivileges"]:
        risk_score += 3
    if flags["privateSaleFunctions"]:
        risk_score += 1
    if flags["thirdPartyScore"] is not None:
        risk_score += max(0, (100 - flags["thirdPartyScore"]) // 20)
    if flags["thirdPartyWhitelist"]:
        risk_score += 4

    flags["score"] = risk_score
    return flags


def detect_proxy_implementation(addr: str) -> Optional[str]:
    """Return implementation address if contract is a proxy (EIP1967)."""
    try:
        raw = w3_read.eth.get_storage_at(to_checksum_address(addr), EIP1967_IMPL_SLOT)
        if int.from_bytes(raw, "big") != 0:
            return to_checksum_address("0x" + raw.hex()[-40:])
    except Exception:
        pass
    return None


async def _check_renounced_by_event_async(addr: str) -> bool:
    if not ETHERSCAN_LOOKUPS_ENABLED:
        return False
    topic0 = Web3.keccak(text="OwnershipTransferred(address,address)").hex()
    zero_topic = "0x" + "0" * 64
    params = {
        "module": "logs",
        "action": "getLogs",
        "fromBlock": "0",
        "toBlock": "latest",
        "address": addr,
        "topic0": topic0,
        "topic2": zero_topic,
        "apikey": get_next_etherscan_key(),
    }
    try:
        params = _prepare_etherscan_params(params)
        async with create_aiohttp_session() as session:
            async with session.get(ETHERSCAN_API_URL, params=params, timeout=20) as r:
                j = await r.json()
        if j.get("status") == "1" and j.get("result"):
            return True
    except (aiohttp.ClientError, asyncio.TimeoutError, OSError) as e:
        disable_etherscan_lookups(f"renounce lookup failed: {e}")
    except Exception:
        pass
    return False


def check_renounced_by_event(addr: str) -> bool:
    return asyncio.run(_check_renounced_by_event_async(addr))


def run_slither_analysis(source: object) -> dict:
    """Run Slither on the given source code and return issue count.

    ``source`` can be a string (single Solidity file) or a list of
    ``{"filename": str, "content": str}`` dictionaries representing a project.
    """
    import tempfile, subprocess, json as _json, shutil

    result = {"slitherIssues": None}
    if shutil.which("slither") is None:
        logger.warning("slither executable not found; skipping analysis")
        result["slitherIssues"] = "not_installed"
        return result
    try:
        with tempfile.TemporaryDirectory() as tmpd:
            if isinstance(source, str):
                src_path = os.path.join(tmpd, "contract.sol")
                with open(src_path, "w", encoding="utf-8") as f:
                    f.write(source)
                target = src_path
            else:
                for part in source:
                    fpath = os.path.join(tmpd, part.get("filename", "contract.sol"))
                    os.makedirs(os.path.dirname(fpath), exist_ok=True)
                    with open(fpath, "w", encoding="utf-8") as f:
                        f.write(part.get("content", ""))
                target = tmpd

            out_json = os.path.join(tmpd, "slither.json")
            cmd = ["slither", target, "--json", out_json, "--solc-disable-warnings"]
            proc = subprocess.run(
                cmd,
                check=True,
                timeout=60,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
            )
            data = _json.load(open(out_json, encoding="utf-8"))
            issues = data.get("results", {}).get("detectors", [])
            result["slitherIssues"] = len(issues)
    except subprocess.CalledProcessError as e:
        stdout = (
            e.stdout.decode("utf-8", errors="replace")
            if isinstance(e.stdout, bytes)
            else str(e.stdout)
        )
        stderr = (
            e.stderr.decode("utf-8", errors="replace")
            if isinstance(e.stderr, bytes)
            else str(e.stderr)
        )
        logger.warning(f"slither analysis failed: {stdout}\n{stderr}")
    except Exception as e:
        logger.warning(f"slither analysis error: {e}")
    return result


def advanced_contract_check(token_addr: str) -> dict:
    """Fetch source and analyze risk. Handles proxies and renounce checks."""
    impl = detect_proxy_implementation(token_addr)
    target = impl or token_addr
    info = fetch_contract_source_etherscan(target)
    owner_addr, owner_bal, owner_token_bal = get_owner_info(target)
    suspicious_activity = False
    if owner_addr:
        suspicious_activity = check_owner_wallet_activity(target, owner_addr)
    renounced = False
    if owner_addr and owner_addr.lower() == ZERO_ADDRESS.lower():
        renounced = True
    else:
        renounced = check_renounced_by_event(target)
    third_score = get_third_party_risk_score(target)
    private_sale = detect_private_sale_indicators(target)
    onchain_metrics = fetch_onchain_metrics(target)

    if info["status"] == ContractVerificationStatus.VERIFIED:
        combined = "".join(part["content"] + "\n" for part in info["source"])
        flags = analyze_solidity_source(combined)
        slither_res = run_slither_analysis(info["source"])
        if slither_res.get("slitherIssues") is not None:
            flags["slitherIssues"] = slither_res["slitherIssues"]
        else:
            flags["slitherIssues"] = "error"
        if impl:
            flags["upgradeableProxy"] = True
        if renounced:
            flags["renounced"] = True
        else:
            flags["renounced"] = False
        flags["thirdPartyScore"] = third_score
        score = flags.get("score", 0)
        if third_score is not None:
            score += max(0, (100 - third_score) // 20)
        owner_flag = flags.get("ownerActivity") or suspicious_activity
        if owner_flag:
            flags["ownerActivity"] = True
            score += 3
        if isinstance(flags.get("slitherIssues"), int):
            score += min(flags["slitherIssues"], 5)
        if not renounced:
            score += 2
        if score >= 10:
            st = "HIGH_RISK"
        else:
            st = "OK"
        result = {
            "verified": True,
            "riskScore": score,
            "riskFlags": flags,
            "status": st,
            "owner": owner_addr,
            "ownerBalanceEth": owner_bal,
            "ownerTokenBalance": owner_token_bal,
            "implementation": impl,
            "renounced": renounced,
            "slitherIssues": flags.get("slitherIssues"),
            "ownerActivity": suspicious_activity,
            "privateSale": private_sale,
            "onChainMetrics": onchain_metrics,
        }
        return result
    elif info["status"] == ContractVerificationStatus.UNVERIFIED:
        return {
            "verified": False,
            "riskScore": 9999,
            "riskFlags": {"ownerActivity": suspicious_activity},
            "status": "UNVERIFIED",
            "owner": owner_addr,
            "ownerBalanceEth": owner_bal,
            "ownerTokenBalance": owner_token_bal,
            "implementation": impl,
            "renounced": renounced,
            "slitherIssues": None,
        }
    else:
        return {
            "verified": False,
            "riskScore": None,
            "riskFlags": {"ownerActivity": suspicious_activity},
            "status": "ERROR",
            "owner": owner_addr,
            "ownerBalanceEth": owner_bal,
            "ownerTokenBalance": owner_token_bal,
            "implementation": impl,
            "renounced": renounced,
            "slitherIssues": None,
            "error": info.get("error"),
        }



###########################################################
# 10. PASSING PAIRS
###########################################################

passing_pairs: Dict[str, dict] = {}
PASSED_PAIRS: Set[str] = set()

# Pairs waiting to hit minimum volume/trade requirements before promotion
volume_checks: Dict[str, dict] = {}


def queue_passing_refresh(
    pair_addr: str, token0: str, token1: str, init_mc: float, init_liq: float
):
    if pair_addr not in passing_pairs:
        passing_pairs[pair_addr] = {
            "token0": token0,
            "token1": token1,
            "attempt_index": 0,
            "last_attempt": time.time(),
            "initial_mc": init_mc,
            "mc_milestones_hit": set(),
            "silent_active": True,
            "last_silent_check": time.time(),
            "last_silent_mc": init_mc,
            "last_liq_check": time.time(),
            "last_liquidity": init_liq,
            "no_new_high_count": 0,
            # For quick 1-min honeypot check
            "last_hp_check": 0,
            # To avoid repeated spam logs
            "no_more_attempts_logged": False,
            "lp_supply": get_lp_total_supply(pair_addr),
            "last_lp_check": time.time(),
            "last_lp_block": w3_read.eth.block_number,
            "first_sell_block": w3_read.eth.block_number,
            "first_sell_detected": False,
            "verification_retry_at": 0,
            "verification_warning": False,
        }


def queue_volume_check(
    pair_addr: str,
    token0: str,
    token1: str,
    passes: int,
    total: int,
    extra: dict,
    is_recheck: bool = False,
    attempt_num: int = None,
):
    if pair_addr not in volume_checks:
        volume_checks[pair_addr] = {
            "token0": token0,
            "token1": token1,
            "start": time.time(),
            "last_check": 0,
            "passes": passes,
            "total": total,
            "extra": extra,
            "is_recheck": is_recheck,
            "attempt": attempt_num,
        }


def _format_utc_timestamp(ts: int) -> str:
    return datetime.utcfromtimestamp(ts).strftime("%Y-%m-%d %H:%M:%S UTC")


def _format_duration_brief(seconds: int) -> str:
    if seconds <= 0:
        return ""
    parts: List[str] = []
    days, remainder = divmod(seconds, 86400)
    if days:
        parts.append(f"{days}d")
    hours, remainder = divmod(remainder, 3600)
    if hours:
        parts.append(f"{hours}h")
    minutes, _ = divmod(remainder, 60)
    if minutes or not parts:
        parts.append(f"{minutes}m")
    return " ".join(parts)


def _format_lock_source_name(source: Optional[str]) -> Optional[str]:
    if not source:
        return None
    normalized = source.strip().lower()
    mapping = {
        "holder_analysis": "Holder snapshot",
        "uncx_rest": "UNCX REST",
        "uncx_graph": "UNCX Graph",
        "etherscan_tokentx": "Etherscan token tx",
        "dexscreener_label": "DexScreener label",
        "ethplorer_api": "Ethplorer",
    }
    if normalized in mapping:
        return mapping[normalized]
    return source.replace("_", " ").strip().title()


def _build_lock_info_line(
    created_at: Optional[int], lock_details: Optional[Dict[str, Any]]
) -> Optional[str]:
    segments: List[str] = []
    if created_at:
        segments.append(f"Pair Created: {_format_utc_timestamp(created_at)}")

    details = lock_details or {}
    locked_at = _normalize_timestamp_seconds(details.get("lockedAt")) if details else None
    unlock_at = _normalize_timestamp_seconds(details.get("unlockAt")) if details else None
    raw_duration = details.get("lockDurationSeconds")
    duration_value: Optional[int] = None
    if raw_duration is not None:
        try:
            duration_value = int(raw_duration)
        except (ValueError, TypeError):
            duration_value = None
    if duration_value is None and locked_at and unlock_at and unlock_at > locked_at:
        duration_value = unlock_at - locked_at

    if locked_at:
        segments.append(f"Locked: {_format_utc_timestamp(locked_at)}")
    elif details:
        segments.append("Locked: Unknown")
    if duration_value and duration_value > 0:
        segments.append(f"Duration: {_format_duration_brief(duration_value)}")
    elif details:
        segments.append("Duration: Unknown")

    coverage = details.get("coveragePct")
    if isinstance(coverage, (int, float)):
        segments.append(f"Coverage: {coverage * 100:.2f}%")

    if not segments:
        return None

    return "Lock Info: " + " | ".join(segments)


def send_ui_criteria_message(
    pair_addr: str,
    passes: int,
    total: int,
    is_recheck: bool = False,
    token_name: str = "",
    clog_percent: float = None,
    logo_url: str = None,
    extra_stats: Dict = None,
    recheck_attempt: int = None,
    is_passing_refresh: bool = False,
):
    """Send a concise Telegram update when a pair clears the criteria."""

    if passes < MIN_PASS_THRESHOLD:
        return

    prefix = "[Refresh]" if is_passing_refresh else ("[Recheck]" if is_recheck else "[NewPair]")
    attempt_str = f" (Attempt #{recheck_attempt})" if recheck_attempt else ""
    pass_str = f"{passes}/{total} passes"
    tn = token_name or "Unnamed"

    msg = (
        f"🟢 <b>{tn}</b> {prefix}{attempt_str}\n"
        f"Pair: <code>{pair_addr}</code>\n"
        f"Criteria: <b>{pass_str}</b>"
    )

    global PASSED_PAIRS
    first_time_pass = False
    key = pair_addr.lower()
    if key not in PASSED_PAIRS:
        PASSED_PAIRS.add(key)
        first_time_pass = True

    status_text = (
        "✅ New passing pair ready for trading"
        if first_time_pass
        else "✅ Pair remains in passing status"
    )
    msg += f"\nStatus: {status_text}"

    if clog_percent is not None:
        msg += f"\nClog: {clog_percent:.0f}% sells"

    if extra_stats:
        pr = extra_stats.get("priceUsd")
        li = extra_stats.get("liquidityUsd")
        vo = extra_stats.get("volume24h")
        fdv = extra_stats.get("fdv")
        mc = extra_stats.get("marketCap")
        buys = extra_stats.get("buys")
        sells = extra_stats.get("sells")
        locked = extra_stats.get("lockedLiquidity", False)
        lock_details = extra_stats.get("lockedLiquidityDetails")
        pair_created_at = _normalize_timestamp_seconds(extra_stats.get("pairCreatedAt"))
        owner_addr = extra_stats.get("owner")
        owner_bal = extra_stats.get("ownerBalanceEth")
        owner_tok = extra_stats.get("ownerTokenBalance")
        impl = extra_stats.get("implementation")
        contract_renounced = extra_stats.get("contractRenounced")
        slither_issues = extra_stats.get("slitherIssues")
        psale = extra_stats.get("privateSale", {})
        metrics = extra_stats.get("onChainMetrics", {})

        msg += "\n\n<b>Dex Stats:</b>"
        if pr and pr > 0:
            msg += f"\nPrice: ${pr:,.12f}"
        if li is not None:
            msg += f"\nLiquidity: ${li:,.0f}"
        if vo is not None:
            msg += f"\n24h Volume: ${vo:,.0f}"
        if fdv:
            msg += f"\nFDV: ${fdv:,.0f}"
        if mc:
            msg += f"\nM.Cap: ${mc:,.0f}"
        msg += f"\nBuys/Sells: {buys}/{sells}"
        if locked is not None:
            method_text = ""
            if lock_details:
                source_label = _format_lock_source_name(lock_details.get("source"))
                if source_label:
                    method_text = f" ({source_label})"
            msg += f"\nLiquidity Locked: <b>{'Yes' if locked else 'No'}</b>{method_text}"
            lock_info_line = _build_lock_info_line(pair_created_at, lock_details)
            if lock_info_line:
                msg += f"\n{lock_info_line}"
        rscore = extra_stats.get("riskScore")
        verified_bool = bool(extra_stats.get("verified") is True)
        risk_warning = extra_stats.get("riskWarning")
        status_text = str(extra_stats.get("contractCheckStatus") or "")
        if not risk_warning and status_text.upper() == "ERROR":
            risk_warning = (
                f"Verification error - retrying in {_format_retry_delay(VERIFICATION_RETRY_DELAY)}"
            )
        if rscore is not None:
            msg += f"\nVerified: <b>{'Yes' if verified_bool else 'No'}</b> | Risk Score: {rscore}"
            if risk_warning:
                msg += f" ⚠️ {risk_warning}"
        links = extra_stats.get("socialLinks", [])
        if links:
            msg += "\n\n<b>Links:</b>"
            for lk in links:
                msg += f"\n{lk}"
        if owner_addr:
            msg += f"\nOwner: {owner_addr}"
        if owner_addr and owner_addr.lower() != ZERO_ADDRESS.lower() and owner_bal is not None:
            msg += f"\nOwner ETH: {owner_bal:.4f}"
        if owner_addr and owner_addr.lower() != ZERO_ADDRESS.lower() and owner_tok is not None:
            msg += f"\nOwner Token Bal: {owner_tok}"
        if impl:
            msg += f"\nImpl: {impl}"
        if contract_renounced is not None:
            msg += f"\nContract Renounced: <b>{'Yes' if contract_renounced else 'No'}</b>"
        if psale.get("hasPresale"):
            count = len(psale.get("largeTransfers", []))
            msg += f"\nPrivate Sale: {count} large transfers"
        if metrics:
            hc = metrics.get("holderConcentration")
            ratio = metrics.get("uniqueBuyerSellerRatio")
            sm = metrics.get("smartMoneyCount")
            if hc is not None:
                msg += f"\nHolder Concentration: {hc:.2%}"
            if ratio is not None:
                msg += f"\nBuyer/Seller Ratio: {ratio:.2f}"
            if sm:
                msg += f"\nSmart Money Buys: {sm}"
        if slither_issues not in (None, "error"):
            msg += f"\nSlither Issues: {slither_issues}"
        elif slither_issues == "error":
            msg += "\nSlither: error"

    send_telegram_message(msg)


def evaluate_fail_reasons(extra: Dict) -> List[str]:
    """Return list of failure reasons based on statistics."""

    reasons: List[str] = []

    def _to_float(value: Any, default: float = 0.0) -> float:
        if value is None:
            return default
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, str):
            try:
                return float(value.strip())
            except ValueError:
                return default
        return default

    def _to_int(value: Any, default: int = 0) -> int:
        return int(_to_float(value, float(default)))

    mc = _to_float(extra.get("marketCap"))
    liq = _to_float(extra.get("liquidityUsd"))
    fdv = _to_float(extra.get("fdv"))
    buys = _to_int(extra.get("buys"))
    sells = _to_int(extra.get("sells"))
    risk = _to_float(extra.get("riskScore"))
    renounced_contract = extra.get("contractRenounced")

    slither_raw = extra.get("slitherIssues")
    if isinstance(slither_raw, int):
        slither_issues: Optional[int] = slither_raw
    elif isinstance(slither_raw, str):
        try:
            slither_issues = int(slither_raw.strip())
        except ValueError:
            slither_issues = None
    else:
        slither_issues = None

    if mc >= 100_000 and (buys + sells) < 10:
        reasons.append("High market cap with <10 buys/sells")
    if liq < MIN_LIQUIDITY_USD:
        reasons.append(f"Liquidity below ${MIN_LIQUIDITY_USD}")
    if fdv < MIN_FDV_USD or mc < MIN_MARKETCAP_USD:
        reasons.append("FDV/MC below thresholds")
    if risk >= 10:
        reasons.append("Contract high risk")
    if renounced_contract is False:
        reasons.append("Contract not renounced")
    if slither_issues is not None and slither_issues >= 5:
        reasons.append(f"Slither issues {slither_issues}")

    return reasons


def check_marketcap_milestones(pair_addr: str, current_mc: float):
    if pair_addr not in passing_pairs:
        return
    info = passing_pairs[pair_addr]
    init_mc = info["initial_mc"]
    if init_mc <= 0:
        return
    ratio = current_mc / init_mc
    for milestone in [2, 3, 5, 10]:
        if ratio >= milestone and milestone not in info["mc_milestones_hit"]:
            info["mc_milestones_hit"].add(milestone)
            logger.info(
                f"[MC Milestone] {pair_addr} reached {milestone}x from ${init_mc:,.0f}"
            )
            send_telegram_message(
                f"[MC Milestone] {pair_addr} reached {milestone}x from ${init_mc:,.0f}"
            )


def handle_passing_refreshes():
    now_ts = time.time()
    remove_list = []
    for pair, data in list(passing_pairs.items()):
        if not data.get("first_sell_detected"):
            seller = detect_first_sell(
                pair,
                data["token0"],
                data["token1"],
                data.get("first_sell_block", w3_read.eth.block_number),
            )
            if seller:
                token_addr = get_non_weth_token(data["token0"], data["token1"])
                info, risk = analyze_seller_wallet(token_addr, seller)
                send_telegram_message(
                    f"[FirstSell] {pair} sold by <code>{seller}</code> ({info}, risk {risk})"
                )
                record_first_sell(pair, seller, w3_read.eth.block_number)
                data["first_sell_detected"] = True
                if risk >= 5:
                    remove_list.append(pair)

        idx = data["attempt_index"]

        # normal refresh attempts
        if idx < len(PASSING_REFRESH_DELAYS):
            delay = PASSING_REFRESH_DELAYS[idx]
            if (now_ts - data["last_attempt"]) >= delay:
                retry_at = data.get("verification_retry_at", 0)
                if retry_at and now_ts < retry_at:
                    # Awaiting scheduled verification retry; skip refresh attempt this cycle.
                    pass
                else:
                    data["attempt_index"] += 1
                    data["last_attempt"] = now_ts
                    attempt_num = data["attempt_index"]
                    logger.info(f"[Refresh] Attempt #{attempt_num} for {pair}")

                    passes, total, xtra = recheck_logic_detail(
                        pair, data["token0"], data["token1"], attempt_num, True
                    )
                    fail, reason = critical_verification_failure(xtra)
                    if fail:
                        if reason == "verification error" and passes >= MIN_PASS_THRESHOLD:
                            _schedule_verification_retry(
                                pair,
                                data,
                                xtra,
                                context="passing refresh",
                                attempt_num=attempt_num,
                            )
                            continue
                        logger.info(f"[Remove] {pair} removed: {reason}")
                        if reason not in ("verification error", "risk score 9999"):
                            send_telegram_message(f"[Remove] {pair} removed: {reason}")
                        remove_list.append(pair)
                        continue
                    else:
                        if data.get("verification_warning") and str(
                            xtra.get("contractCheckStatus", "")
                        ).upper() != "ERROR":
                            data["verification_warning"] = False
                            data["verification_retry_at"] = 0
                    data["last_liquidity"] = xtra.get(
                        "liquidityUsd", data.get("last_liquidity")
                    )
                    data["last_liq_check"] = now_ts
                    if passes < MIN_PASS_THRESHOLD:
                        logger.info(f"[Refresh] => removing {pair} from passing.")
                        remove_list.append(pair)
        else:
            # no more attempts
            if not data["no_more_attempts_logged"]:
                logger.info(f"[Refresh] No more scheduled attempts for {pair}")
                data["no_more_attempts_logged"] = True

        if pair in remove_list:
            continue

        # silent check for MC
        if data["silent_active"]:
            detection_time = detected_at.get(pair.lower(), data["last_attempt"])
            elapsed = now_ts - detection_time
            if elapsed > SILENT_CHECK_DURATION:
                data["silent_active"] = False
                logger.info(f"[SilentCheck] => ended for {pair}")
            else:
                if (now_ts - data["last_silent_check"]) >= SILENT_CHECK_INTERVAL:
                    logger.info(f"[SilentCheck] => 10-min for {pair}")
                    p2, t2, xtra = check_pair_criteria(
                        pair, data["token0"], data["token1"]
                    )
                    fail, reason = critical_verification_failure(xtra)
                    if fail:
                        if reason == "verification error" and p2 >= MIN_PASS_THRESHOLD:
                            _schedule_verification_retry(
                                pair,
                                data,
                                xtra,
                                context="silent refresh",
                            )
                            data["last_silent_check"] = now_ts
                            continue
                        logger.info(f"[Remove] {pair} removed: {reason}")
                        if reason not in ("verification error", "risk score 9999"):
                            send_telegram_message(f"[Remove] {pair} removed: {reason}")
                        remove_list.append(pair)
                        continue
                    else:
                        if data.get("verification_warning") and str(
                            xtra.get("contractCheckStatus", "")
                        ).upper() != "ERROR":
                            data["verification_warning"] = False
                            data["verification_retry_at"] = 0
                    cmc = xtra.get("marketCap", 0)
                    data["last_liquidity"] = xtra.get(
                        "liquidityUsd", data.get("last_liquidity")
                    )
                    data["last_liq_check"] = now_ts
                    pmc = data["last_silent_mc"]
                    if elapsed >= 1800 and abs(cmc - pmc) < 1e-5:
                        logger.info(
                            f"[SilentCheck] => Inactive after 30m no MC change {pair}"
                        )
                        data["silent_active"] = False
                    else:
                        if cmc > pmc:
                            data["last_silent_mc"] = cmc
                            data["no_new_high_count"] = 0
                            ratio = 0
                            if data["initial_mc"] > 0:
                                ratio = cmc / data["initial_mc"]
                            logger.info(
                                f"[SilentCheck] {pair} new MC high ~${cmc:,.0f} (x{ratio:.2f})"
                            )
                            send_telegram_message(
                                f"[SilentCheck] {pair} new MC high ~${cmc:,.0f} (x{ratio:.2f})"
                            )
                        else:
                            data["no_new_high_count"] += 1
                            if data["no_new_high_count"] >= 2:
                                logger.info(f"[SilentCheck] => done no new high {pair}")
                                data["silent_active"] = False
                    data["last_silent_check"] = now_ts

        # quick 30s honeypot check
        if pair not in remove_list:
            if (now_ts - data["last_hp_check"]) >= 30:
                data["last_hp_check"] = now_ts
                hp0 = check_honeypot_is(data["token0"], pair_addr=pair)
                hp1 = check_honeypot_is(data["token1"], pair_addr=pair)
                if hp0 or hp1:
                    logger.info(f"[HoneypotDetected] {pair} removed from passing pairs")
                    send_telegram_message(f"[HoneypotDetected] {pair} removed from passing pairs")
                    remove_list.append(pair)

        # liquidity removal check every 5m
        if pair not in remove_list:
            if (now_ts - data.get("last_liq_check", 0)) >= 300:
                ds = fetch_dexscreener_data(data["token0"], pair)
                if ds:
                    new_liq = ds.get("liquidityUsd", 0)
                    old_liq = data.get("last_liquidity", new_liq)
                    if old_liq and new_liq < old_liq * 0.9:
                        send_telegram_message(
                            f"[LiquidityRemoved] {pair} liquidity dropped from ${old_liq:,.0f} to ${new_liq:,.0f} => potential rugpull"
                        )
                        remove_list.append(pair)
                    data["last_liquidity"] = new_liq

                # check LP supply drop
                supply_now = get_lp_total_supply(pair)
                prev_supply = data.get("lp_supply")
                if (
                    supply_now is not None
                    and prev_supply
                    and supply_now < prev_supply * 0.9
                ):
                    send_telegram_message(
                        f"[LiquidityRemoved] {pair} LP supply drop detected"
                    )
                    remove_list.append(pair)
                if supply_now is not None:
                    data["lp_supply"] = supply_now

                # check for Burn events
                last_block = data.get("last_lp_block", w3_read.eth.block_number)
                try:
                    logs = w3_read.eth.get_logs(
                        {
                            "address": pair,
                            "fromBlock": last_block + 1,
                            "toBlock": "latest",
                            "topics": [BURN_TOPIC],
                        }
                    )
                    if logs:
                        send_telegram_message(
                            f"[LiquidityRemoved] {pair} burn event detected"
                        )
                        remove_list.append(pair)
                    data["last_lp_block"] = w3_read.eth.block_number
                except Exception as e:
                    logger.debug(f"burn log error: {e}")


                if check_recent_liquidity_removal(pair):
                    send_telegram_message(
                        f"[LiquidityRemoved] {pair} LP burn detected via Etherscan"
                    )
                    remove_list.append(pair)

                data["last_liq_check"] = now_ts

        if pair not in remove_list:
            # done attempts & silent => remove
            if (
                data["attempt_index"] >= len(PASSING_REFRESH_DELAYS)
                and not data["silent_active"]
            ):
                remove_list.append(pair)

    for rm in remove_list:
        passing_pairs.pop(rm, None)
        logger.info(f"[PassingPairs] => removed {rm}")


def handle_volume_checks():
    now = time.time()
    remove = []
    for pair, data in list(volume_checks.items()):
        if (now - data["last_check"]) >= 5 or (now - data["start"]) >= 300:
            data["last_check"] = now
            ds = fetch_dexscreener_data(data["token0"], pair)
            if not ds:
                if (now - data["start"]) >= 300:
                    remove.append(pair)
                continue
            vol = ds.get("volume24h", 0)
            trades = ds.get("buys", 0) + ds.get("sells", 0)
            if vol >= MIN_VOLUME_USD and trades >= MIN_TRADES_REQUIRED:
                extra = data["extra"]
                extra.update(ds)
                send_ui_criteria_message(
                    pair,
                    data["passes"],
                    data["total"],
                    is_recheck=data.get("is_recheck", False),
                    token_name=extra.get("tokenName", ""),
                    clog_percent=extra.get("clogPercent"),
                    extra_stats=extra,
                    recheck_attempt=data.get("attempt"),
                    is_passing_refresh=False,
                )
                queue_passing_refresh(
                    pair,
                    data["token0"],
                    data["token1"],
                    ds.get("marketCap", 0),
                    ds.get("liquidityUsd", 0),
                )
                remove.append(pair)
            elif (now - data["start"]) >= 300:
                logger.info(f"[VolumeCheck] {pair} failed to reach targets")
                remove.append(pair)

    for rm in remove:
        volume_checks.pop(rm, None)


def handle_wallet_updates():
    now = time.time()
    for token, data in list(WALLET_REPORT_CACHE.items()):
        if now - data.get("ts", 0) >= WALLET_REPORT_TTL:
            try:
                fut = asyncio.run_coroutine_threadsafe(
                    wallet_tracker.generate_wallet_report(token), wallet_event_loop
                )
                WALLET_REPORT_CACHE[token]["report"] = fut.result()
                WALLET_REPORT_CACHE[token]["ts"] = time.time()
                time.sleep(1)
            except Exception as e:
                logger.error(f"wallet update error for {token}: {e}")


###########################################################
# 10.5 recheck_logic_detail
###########################################################


def recheck_logic_detail(
    pair_addr: str, t0: str, t1: str, attempt_num: int, is_passing_refresh: bool
):
    passes, total, extra = check_pair_criteria(pair_addr, t0, t1)
    status = str(extra.get("contractCheckStatus", "")).upper()
    if status == "ERROR":
        extra["riskWarning"] = (
            f"Verification error - retrying in {_format_retry_delay(VERIFICATION_RETRY_DELAY)}"
        )
    mode = "Refresh" if is_passing_refresh else "Recheck"
    logger.info(
        f"[{mode}] => {pair_addr} => {passes}/{total} passes (attempt {attempt_num}). "
        f"Verified={extra.get('contractCheckStatus')} Risk={extra.get('riskScore')}"
    )

    # evaluate for suspicious metrics before notifying
    fail_reasons = evaluate_fail_reasons(extra)
    if fail_reasons:
        passes = 0

    if passes >= MIN_PASS_THRESHOLD and is_passing_refresh:
        send_ui_criteria_message(
            pair_addr,
            passes,
            total,
            is_recheck=False,
            token_name=extra.get("tokenName", ""),
            clog_percent=extra.get("clogPercent"),
            extra_stats=extra,
            recheck_attempt=attempt_num,
            is_passing_refresh=True,
        )

    mc = extra.get("marketCap", 0)
    if mc > 0:
        check_marketcap_milestones(pair_addr, mc)
    main_token = get_non_weth_token(t0, t1)
    wallet_rep = get_wallet_report(main_token)
    if passes >= MIN_PASS_THRESHOLD and not is_passing_refresh:
        metrics.increment("passes")
    if passes >= MIN_PASS_THRESHOLD:
        start_wallet_monitor(main_token)
    return (passes, total, extra)


###########################################################
# 11. RECHECK FAILING
###########################################################

pending_rechecks: Dict[str, dict] = {}


def _collect_queue_depth() -> dict:
    return {
        "passing_pairs": len(passing_pairs),
        "volume_checks": len(volume_checks),
        "pending_rechecks": len(pending_rechecks),
    }


metrics.set_queue_depth_callback(_collect_queue_depth)


def queue_recheck(pair_addr: str, token0: str, token1: str):
    if pair_addr not in pending_rechecks:
        pending_rechecks[pair_addr] = {
            "token0": token0,
            "token1": token1,
            "attempt_index": 0,
            "last_attempt": time.time(),
            "created": time.time(),
            "fail_count": 0,
            "verification_retry_at": 0,
            "verification_warning": False,
        }


def handle_rechecks():
    now_ts = time.time()
    rm_list = []
    for pair, data in list(pending_rechecks.items()):
        idx = data["attempt_index"]
        if idx >= len(RECHECK_DELAYS):
            logger.info(f"[Recheck] => removing {pair}, no more attempts.")
            rm_list.append(pair)
            continue

        delay = RECHECK_DELAYS[idx]
        if (now_ts - data["last_attempt"]) >= delay:
            retry_at = data.get("verification_retry_at", 0)
            if retry_at and now_ts < retry_at:
                continue
            data["attempt_index"] += 1
            data["last_attempt"] = now_ts
            attempt_num = data["attempt_index"]
            logger.info(f"[Recheck] Attempt #{attempt_num} for {pair}")

            passes, total, extra = recheck_logic_detail(
                pair, data["token0"], data["token1"], attempt_num, False
            )
            if isinstance(extra, dict) and extra.get("dexscreener_missing"):
                dex_reason = extra.get("dexscreener_reason", "unknown")
                if extra.get("should_requeue", True):
                    logger.info(
                        f"[Recheck] => {pair} DexScreener unavailable ({dex_reason}); will retry"
                    )
                else:
                    age = extra.get("dexscreener_not_listed_age")
                    retry_window = extra.get("dexscreener_retry_window")
                    if age is not None and extra.get(
                        "dexscreener_retry_window_expired"
                    ):
                        logger.info(
                            f"[Recheck] => removing {pair}, DexScreener still not listed after"
                            f" ~{age:.1f}s (window {retry_window}s)"
                        )
                    else:
                        logger.info(
                            f"[Recheck] => removing {pair}, DexScreener unavailable ({dex_reason})"
                        )
                    rm_list.append(pair)
                continue
            fail, reason = critical_verification_failure(extra)
            if fail:
                if reason == "verification error" and passes >= MIN_PASS_THRESHOLD:
                    _schedule_verification_retry(
                        pair,
                        data,
                        extra,
                        context="recheck",
                        attempt_num=attempt_num,
                    )
                    continue
                logger.info(f"[Remove] {pair} removed: {reason}")
                if reason not in ("verification error", "risk score 9999"):
                    send_telegram_message(f"[Remove] {pair} removed: {reason}")
                rm_list.append(pair)
                continue
            else:
                if data.get("verification_warning"):
                    data["verification_warning"] = False
                    data["verification_retry_at"] = 0
            if passes >= MIN_PASS_THRESHOLD:
                vol_now = extra.get("volume24h", 0)
                trades_now = extra.get("buys", 0) + extra.get("sells", 0)
                if vol_now >= MIN_VOLUME_USD and trades_now >= MIN_TRADES_REQUIRED:
                    logger.info(f"[Recheck] => removing {pair}, now passing.")
                    rm_list.append(pair)
                    mc_now = extra.get("marketCap", 0)
                    liq_now = extra.get("liquidityUsd", 0)
                    send_ui_criteria_message(
                        pair,
                        passes,
                        total,
                        is_recheck=True,
                        token_name=extra.get("tokenName", ""),
                        clog_percent=extra.get("clogPercent"),
                        extra_stats=extra,
                        recheck_attempt=attempt_num,
                        is_passing_refresh=False,
                    )
                    queue_passing_refresh(
                        pair, data["token0"], data["token1"], mc_now, liq_now
                    )
                    main_token = get_non_weth_token(data["token0"], data["token1"])
                    start_wallet_monitor(main_token)
                    continue
                else:
                    logger.info(f"[VolumeCheck] waiting on {pair}")
                    queue_volume_check(
                        pair,
                        data["token0"],
                        data["token1"],
                        passes,
                        total,
                        extra,
                        is_recheck=True,
                        attempt_num=attempt_num,
                    )
                    rm_list.append(pair)
                    continue
            else:
                line = (
                    f"- {extra.get('tokenName','Unnamed')} => {passes}/{total} (Attempt #{attempt_num}) "
                    f"Verified={extra.get('contractCheckStatus')} Risk={extra.get('riskScore')}"
                )
                if not extra.get("transient_failure"):
                    add_failed_recheck_message(line)
                    data["fail_count"] = data.get("fail_count", 0) + 1
                    if data["fail_count"] >= 3:
                        logger.info(f"[Recheck] => removing {pair}, failed 3 times")
                        rm_list.append(pair)
                        continue
                else:
                    logger.info(
                        f"[Recheck] => {pair} retry deferred due to transient issue"
                    )

        # stop after 10 minutes of failures
        elapsed = now_ts - data.get("created", data["last_attempt"])
        if elapsed >= 600:
            logger.info(f"[Recheck] => removing {pair}, failed after 10m")
            rm_list.append(pair)

    for r in rm_list:
        pending_rechecks.pop(r, None)
        logger.info(f"[Recheck] => removed {r}")


###########################################################
# Extra mapping for resolving pair -> (token0, token1)
known_pairs: Dict[str, Tuple[str, str]] = {}

###########################################################
# 12. HANDLE NEW PAIR
###########################################################


def handle_new_pair(pair_addr: str, token0: str, token1: str):
    """Process a newly created pair and evaluate its criteria."""

    start_time = time.perf_counter()
    metrics.increment("pairs_scanned")
    status_message = "processed"
    outcome_context: Dict[str, object] = {}
    logged_outcome = False
    paddr_display = pair_addr

    try:
        paddr = to_checksum_address(pair_addr)
        token0 = to_checksum_address(token0)
        token1 = to_checksum_address(token1)
        paddr_display = paddr
        lower_addr = paddr.lower()
        if lower_addr in SEEN_PAIRS:
            log_event(
                logging.INFO,
                "skip_pair",
                f"{paddr} already processed, skipping",
                pair=paddr,
                context={"reason": "already_processed"},
            )
            return
        SEEN_PAIRS.add(lower_addr)
        detected_at[paddr.lower()] = time.time()
        known_pairs[paddr.lower()] = (token0, token1)

        passes, total, extra = check_pair_criteria(paddr, token0, token1)
        outcome_context.update({"passes": passes, "total": total})
        if extra.get("dexscreener_missing"):
            dex_reason = extra.get("dexscreener_reason", "unknown")
            should_requeue = extra.get("should_requeue", True)
            outcome_context["reason"] = dex_reason
            log_context = {"dex_reason": dex_reason}
            if "dexscreener_not_listed_age" in extra:
                log_context["age_seconds"] = round(
                    float(extra["dexscreener_not_listed_age"]), 2
                )
                log_context["retry_window"] = extra.get(
                    "dexscreener_retry_window"
                )
                if extra.get("dexscreener_retry_window_expired"):
                    log_context["retry_window_expired"] = True
            if should_requeue:
                status_message = "requeue_missing_dexscreener"
                log_event(
                    logging.INFO,
                    "requeue",
                    f"{paddr} missing DexScreener data",
                    pair=paddr,
                    context={"reason": status_message, **log_context},
                )
                queue_recheck(paddr, token0, token1)
            else:
                status_message = "skip_missing_dexscreener"
                log_event(
                    logging.INFO,
                    "skip_pair",
                    f"{paddr} missing DexScreener data (not requeueing)",
                    pair=paddr,
                    context=log_context,
                )
            return
        if not isinstance(extra, dict):
            status_message = "invalid_extra"
            log_event(
                logging.ERROR,
                "invalid_pair_data",
                f"check_pair_criteria returned invalid data for {paddr}",
                pair=paddr,
                error=str(extra),
            )
            extra = {}
        if not extra or not extra.get("tokenName"):
            status_message = "skip_missing_token_name"
            log_event(
                logging.DEBUG,
                "skip_pair",
                f"{paddr} missing DexScreener token name",
                pair=paddr,
            )
            return
        fail, reason = critical_verification_failure(extra)
        verification_error = False
        if fail:
            if reason == "verification error":
                verification_error = True
                status_message = "verification_error_pending"
                outcome_context["verification_warning"] = True
                extra.setdefault(
                    "riskWarning",
                    f"Verification error - retrying in {_format_retry_delay(VERIFICATION_RETRY_DELAY)}",
                )
                log_event(
                    logging.INFO,
                    "verification_warning",
                    f"{paddr} contract verification error detected",
                    pair=paddr,
                    context={"reason": reason},
                )
            else:
                status_message = "removed_by_verification"
                outcome_context["remove_reason"] = reason
                log_event(
                    logging.INFO,
                    "remove_pair",
                    f"{paddr} removed: {reason}",
                    pair=paddr,
                    context={"reason": reason},
                )
                if reason not in ("verification error", "risk score 9999"):
                    send_telegram_message(f"[Remove] {paddr} removed: {reason}")
                return
        latency_ms = round((time.perf_counter() - start_time) * 1000, 2)
        log_event(
            logging.INFO,
            "new_pair",
            f"{paddr} => {passes}/{total} partial passes",
            pair=paddr,
            latency_ms=latency_ms,
            context={
                "verification": extra.get("contractCheckStatus"),
                "risk": extra.get("riskScore"),
                "passes": passes,
                "total": total,
            },
        )
        store_pair_record(paddr, token0, token1, passes, total, extra)

        fail_reasons = evaluate_fail_reasons(extra)
        if fail_reasons:
            outcome_context["fail_reasons"] = fail_reasons
            passes = 0

        mc_now = extra.get("marketCap", 0)
        if mc_now > 0:
            check_marketcap_milestones(paddr, mc_now)

        main_token = get_non_weth_token(token0, token1)

        if passes >= MIN_PASS_THRESHOLD:
            metrics.increment("passes")
            start_wallet_monitor(main_token)

        if passes >= MIN_PASS_THRESHOLD:
            vol_now = extra.get("volume24h", 0)
            trades_now = extra.get("buys", 0) + extra.get("sells", 0)
            if vol_now >= MIN_VOLUME_USD and trades_now >= MIN_TRADES_REQUIRED:
                send_ui_criteria_message(
                    paddr,
                    passes,
                    total,
                    is_recheck=False,
                    token_name=extra.get("tokenName", ""),
                    clog_percent=extra.get("clogPercent"),
                    extra_stats=extra,
                )
                liq_now = extra.get("liquidityUsd", 0)
                queue_passing_refresh(paddr, token0, token1, mc_now, liq_now)
                outcome_context["next_step"] = "passing_refresh"
                if verification_error:
                    data_ref = passing_pairs.get(paddr)
                    if data_ref is not None:
                        _schedule_verification_retry(
                            paddr,
                            data_ref,
                            extra,
                            context="initial pass",
                        )
            else:
                log_event(
                    logging.INFO,
                    "queue_volume_check",
                    f"Waiting for volume targets on {paddr}",
                    pair=paddr,
                    context={
                        "volume24h": vol_now,
                        "trades": trades_now,
                    },
                )
                outcome_context["next_step"] = "volume_check"
                queue_volume_check(paddr, token0, token1, passes, total, extra)
        else:
            log_event(
                logging.INFO,
                "queue_recheck",
                f"Not enough passes for {paddr}, scheduling recheck",
                pair=paddr,
            )
            outcome_context["next_step"] = "recheck"
            queue_recheck(paddr, token0, token1)
    except Exception as e:
        latency_ms = round((time.perf_counter() - start_time) * 1000, 2)
        tb = traceback.extract_tb(e.__traceback__)
        line = tb[-1].lineno if tb else "unknown"
        metrics.record_exception()
        log_event(
            logging.ERROR,
            "handle_new_pair_error",
            f"handle_new_pair failed for {paddr_display} at line {line}",
            pair=paddr_display,
            latency_ms=latency_ms,
            error=str(e),
            context={"line": line},
        )
        logger.debug(''.join(traceback.format_tb(e.__traceback__)))
        logged_outcome = True
    finally:
        if not logged_outcome:
            latency_ms = round((time.perf_counter() - start_time) * 1000, 2)
            log_event(
                logging.INFO,
                "handle_new_pair_complete",
                status_message,
                pair=paddr_display,
                latency_ms=latency_ms,
                context=outcome_context or None,
            )


###########################################################
# 13. OPTIONAL REPORTS
###########################################################

_last_daily_date = None
_last_weekly_date = None
_last_hourly_pnl_time = 0




###########################################################
# 14. MAIN LOOP
###########################################################


def load_last_block(fname: str) -> int:
    if os.path.exists(fname):
        try:
            data = json.load(open(fname, "r"))
            return data.get("last_block", 0)
        except Exception:
            pass
    return 0


def save_last_block(bn: int, fname: str):
    try:
        json.dump({"last_block": bn}, open(fname, "w"))
    except Exception:
        pass


def main():
    global runtime_reporter
    log_event(logging.INFO, "startup", "Starting advanced CryptoBot")
    ensure_etherscan_connectivity()
    if runtime_reporter is None:
        runtime_reporter = RuntimeReporter(metrics)
    send_telegram_message(
        "🤖 Ethereum bot v2 online and scanning for fresh Ethereum pairs."
    )

    last_block_v2 = load_last_block(LAST_BLOCK_FILE_V2)
    if last_block_v2 == 0:
        last_block_v2 = safe_block_number(False)
        save_last_block(last_block_v2, LAST_BLOCK_FILE_V2)
        log_event(
            logging.INFO,
            "init_block",
            "Initialized v2 last block",
            context={"network": "v2", "block": last_block_v2},
        )

    last_block_v3 = load_last_block(LAST_BLOCK_FILE_V3)
    if last_block_v3 == 0:
        last_block_v3 = safe_block_number(True)
        save_last_block(last_block_v3, LAST_BLOCK_FILE_V3)
        log_event(
            logging.INFO,
            "init_block",
            "Initialized v3 last block",
            context={"network": "v3", "block": last_block_v3},
        )

    while True:
        try:
            curr_block_v2 = safe_block_number(False)
            if curr_block_v2 > last_block_v2:
                from_blk = last_block_v2 + 1
                to_blk = curr_block_v2

                filter_params = {
                    "fromBlock": from_blk,
                    "toBlock": to_blk,
                    "address": UNISWAP_V2_FACTORY_ADDRESS,
                    "topics": [PAIR_CREATED_TOPIC_V2],
                }
                logs = safe_get_logs(filter_params)

                for lg in logs:
                    if len(lg["topics"]) < 3:
                        continue
                    token0_hex = "0x" + lg["topics"][1].hex()[-40:]
                    token1_hex = "0x" + lg["topics"][2].hex()[-40:]
                    data_field = lg["data"]
                    if isinstance(data_field, HexBytes):
                        data_field = data_field.hex()
                    if data_field.startswith("0x"):
                        data_field = data_field[2:]
                    raw_bytes = bytes.fromhex(data_field)
                    try:
                        pair_addr, _ = decode(["address", "uint256"], raw_bytes)
                        handle_new_pair(pair_addr, token0_hex, token1_hex)
                    except Exception as e:
                        tb = traceback.extract_tb(e.__traceback__)
                        line = tb[-1].lineno if tb else 'unknown'
                        logger.error(
                            f"handle_new_pair v2 error at line {line}: {e}",
                        )

                last_block_v2 = to_blk
                save_last_block(last_block_v2, LAST_BLOCK_FILE_V2)
                log_event(
                    logging.INFO,
                    "blocks_processed",
                    "Processed v2 blocks",
                    context={
                        "network": "v2",
                        "from_block": from_blk,
                        "to_block": to_blk,
                        "pairs_found": len(logs),
                    },
                )

            curr_block_v3 = safe_block_number(True)
            if curr_block_v3 > last_block_v3:
                from_blk = last_block_v3 + 1
                to_blk = curr_block_v3

                filter_params = {
                    "fromBlock": from_blk,
                    "toBlock": to_blk,
                    "address": UNISWAP_V3_FACTORY_ADDRESS,
                    "topics": [POOL_CREATED_TOPIC_V3],
                }
                logs = safe_get_logs(filter_params, is_v3=True)

                for lg in logs:
                    if len(lg["topics"]) < 4:
                        continue
                    token0_hex = "0x" + lg["topics"][1].hex()[-40:]
                    token1_hex = "0x" + lg["topics"][2].hex()[-40:]
                    data_field = lg["data"]
                    if isinstance(data_field, HexBytes):
                        data_field = data_field.hex()
                    if data_field.startswith("0x"):
                        data_field = data_field[2:]
                    raw_bytes = bytes.fromhex(data_field)
                    try:
                        _, pool_addr = decode(["int24", "address"], raw_bytes)
                        handle_new_pair(pool_addr, token0_hex, token1_hex)
                    except Exception as e:
                        tb = traceback.extract_tb(e.__traceback__)
                        line = tb[-1].lineno if tb else 'unknown'
                        logger.error(
                            f"handle_new_pair v3 error at line {line}: {e}",
                        )

                last_block_v3 = to_blk
                save_last_block(last_block_v3, LAST_BLOCK_FILE_V3)
                log_event(
                    logging.INFO,
                    "blocks_processed",
                    "Processed v3 blocks",
                    context={
                        "network": "v3",
                        "from_block": from_blk,
                        "to_block": to_blk,
                        "pairs_found": len(logs),
                    },
                )

            # handle failing rechecks
            handle_rechecks()

            # check volume milestones for pending pairs
            handle_volume_checks()

            # handle passing refresh
            handle_passing_refreshes()

            # update wallet reports
            handle_wallet_updates()

            # flush fail buffer
            maybe_flush_failed_rechecks()

        except KeyboardInterrupt:
            log_event(logging.INFO, "shutdown", "KeyboardInterrupt => stopping")
            break
        except Exception as e:
            metrics.record_exception()
            log_event(
                logging.ERROR,
                "main_loop_error",
                "Main loop error",
                error=str(e),
            )
            time.sleep(5)

        time.sleep(MAIN_LOOP_SLEEP)


if __name__ == "__main__":
    main()
